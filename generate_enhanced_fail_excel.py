# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import sys

# 設定輸出編碼為 UTF-8
sys.stdout.reconfigure(encoding='utf-8')

input_file = r'MINE\result\FAIL匯總.xlsx'
output_file = r'MINE\result\FAIL匯總_Enhanced.xlsx'

print(f"正在讀取原始檔案: {input_file}")

# 讀取原始資料
wb = openpyxl.load_workbook(input_file)
ws = wb.active

data = []

# 遍歷每一行，提取資料
# 假設資料模式是：
# Column 1: Log 檔名資訊
# Column 2: 錯誤訊息 (可能為空)
for row in range(1, ws.max_row + 1):
    col1_val = ws.cell(row=row, column=1).value
    col2_val = ws.cell(row=row, column=2).value
    
    if not col1_val or not isinstance(col1_val, str):
        continue
        
    # 過濾掉標題、分隔線、連結等無關行
    if "錯誤原因統計" in col1_val or "──" in col1_val or "回到 Summary" in col1_val or "工作表快速連結" in col1_val:
        continue
        
    # 嘗試解析 Log 檔名
    # 範例: 1+4cam stitching1 test-WE253600014-2025-0923-104358-B7PL011-202.log_SFIS_ON 測試總時間:0.0 Sec.
    # 簡單的正則表達式提取
    log_info = col1_val
    error_msg = col2_val if col2_val else ""
    
    # 解析欄位
    test_item = "Unknown"
    sn = "Unknown"
    test_time = "Unknown"
    
    # 嘗試提取 SN (通常是 WE 開頭或類似格式，這裡假設是中間那段)
    # 這裡做一個簡單的分割處理
    parts = log_info.split('-')
    if len(parts) >= 4:
        # 嘗試找出 SN (通常在第二個位置，但也可能變動)
        # 這裡用一個啟發式方法：找看起來像 SN 的
        for part in parts:
            if part.startswith('WE') and len(part) > 8:
                sn = part
                break
        
        # 測試項目通常在最前面
        if '+' in parts[0]:
            test_item = parts[0].split('+')[1]
        else:
            test_item = parts[0]
            
        # 時間通常在 SN 後面
        # 這裡簡化處理，直接把原始字串當作參考
    
    # 清理錯誤訊息
    clean_error = str(error_msg)
    clean_error = clean_error.replace("===============錯誤原因====================\n\n", "")
    clean_error = clean_error.replace("==================================================\n\n", "")
    clean_error = clean_error.strip()
    
    # 提取簡短錯誤原因 (取第一行或特定關鍵字)
    short_error = clean_error.split('\n')[0] if clean_error else "Unknown Error"
    if "Segmentation fault" in clean_error:
        short_error = "Segmentation Fault"
    elif "executes fail" in clean_error:
        short_error = "Execution Failed"
    
    data.append({
        "Status": "FAIL",
        "Test Item": test_item,
        "SN": sn,
        "Log Info": log_info, # 保留原始資訊備查
        "Error Type": short_error,
        "Detailed Error": clean_error
    })

if not data:
    print("未提取到有效資料，請檢查解析邏輯。")
    sys.exit(1)

# 建立新的 DataFrame
df_new = pd.DataFrame(data)

# 重新排列欄位
cols = ["Status", "Test Item", "SN", "Error Type", "Detailed Error", "Log Info"]
df_new = df_new[cols]

print(f"提取到 {len(df_new)} 筆資料，正在生成優化版 Excel...")

# 寫入新的 Excel
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_new.to_excel(writer, index=False, sheet_name='FAIL_Analysis')
    
    # 獲取 workbook 和 worksheet 物件進行格式化
    workbook = writer.book
    worksheet = writer.sheets['FAIL_Analysis']
    
    # 定義樣式
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(name='Calibri', size=11, color="FFFFFF", bold=True)
    
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fail_font = Font(color="9C0006")
    
    border = Border(left=Side(style='thin'), 
                   right=Side(style='thin'), 
                   top=Side(style='thin'), 
                   bottom=Side(style='thin'))
    
    # 格式化標題列
    for col in range(1, len(cols) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 格式化資料列
    for row in range(2, len(df_new) + 2):
        # 設定整行邊框
        for col in range(1, len(cols) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True) # 自動換行
            
            # Status 欄位特殊格式
            if col == 1: # Status
                cell.fill = fail_fill
                cell.font = fail_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Detailed Error 欄位 (第 5 欄)
            if col == 5:
                # 限制高度，避免太佔空間 (Excel 無法直接設定最大高度，但可以透過不自動調整行高來控制)
                pass

    # 設定欄寬
    worksheet.column_dimensions['A'].width = 10  # Status
    worksheet.column_dimensions['B'].width = 25  # Test Item
    worksheet.column_dimensions['C'].width = 20  # SN
    worksheet.column_dimensions['D'].width = 30  # Error Type
    worksheet.column_dimensions['E'].width = 60  # Detailed Error
    worksheet.column_dimensions['F'].width = 40  # Log Info
    
    # 凍結首列
    worksheet.freeze_panes = 'A2'
    
    # 開啟篩選
    worksheet.auto_filter.ref = worksheet.dimensions

print(f"已生成優化檔案: {output_file}")
