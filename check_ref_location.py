# -*- coding: utf-8 -*-
import openpyxl
import sys
import re

# 設定輸出編碼為 UTF-8
sys.stdout.reconfigure(encoding='utf-8')

# 讀取測試腳本 Excel 檔案
file_path = 'MINE/VALO360_TestFlow_AOCI_4Cam_Stitching_test1_FATP_250924B.xlsx'
wb = openpyxl.load_workbook(file_path)

# 分析測試腳本工作表
sheet_name = '4cam stitching1 test'
ws = wb[sheet_name]

# 讀取標題列並顯示所有欄位
print("=" * 100)
print("檢查所有欄位標題")
print("=" * 100)

headers = []
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    headers.append(header if header else f"Col_{col}")
    print(f"  列 {col:2d}: {header}")

# 尋找 Ref 相關的欄位
print(f"\n\n尋找包含 'Ref' 或 'ref' 的欄位:")
ref_related = []
for i, h in enumerate(headers, 1):
    if h and ('ref' in str(h).lower() or 'Ref' in str(h)):
        ref_related.append((i, h))
        print(f"  列 {i}: {h}")

# 檢查 Reply 欄位中是否包含 Ref 公式
print(f"\n\n檢查 Reply 欄位中的 Ref 公式:")
reply_col = None
for i, h in enumerate(headers, 1):
    if h == 'Reply':
        reply_col = i
        break

if reply_col:
    print(f"  Reply 欄位在第 {reply_col} 列")
    
    # 檢查前 50 行的 Reply 內容
    ref_pattern = re.compile(r'\$\w+:')
    rows_with_ref = []
    
    for row in range(2, min(52, ws.max_row + 1)):
        reply_value = ws.cell(row, reply_col).value
        if reply_value and ref_pattern.search(str(reply_value)):
            rows_with_ref.append((row, str(reply_value)[:150]))
    
    print(f"\n  前 50 行中,有 {len(rows_with_ref)} 行的 Reply 包含 Ref 公式 ($變數名:)")
    
    if rows_with_ref:
        print(f"\n  範例 (前 10 個):")
        for row_num, content in rows_with_ref[:10]:
            print(f"    Row {row_num}: {content}...")

# 統計整個工作表中 Reply 欄位包含 Ref 的行數
if reply_col:
    total_ref_in_reply = 0
    total_validation = 0
    
    val_col = None
    for i, h in enumerate(headers, 1):
        if h == 'Validation':
            val_col = i
            break
    
    for row in range(2, ws.max_row + 1):
        reply_value = ws.cell(row, reply_col).value
        if reply_value and ref_pattern.search(str(reply_value)):
            total_ref_in_reply += 1
        
        if val_col:
            val_value = ws.cell(row, val_col).value
            if val_value and str(val_value).strip():
                total_validation += 1
    
    print(f"\n\n{'=' * 100}")
    print("完整統計")
    print("=" * 100)
    print(f"  總測試步驟: {ws.max_row - 1:,} 行")
    print(f"  Reply 中包含 Ref 公式: {total_ref_in_reply:,} 行 ({total_ref_in_reply/(ws.max_row-1)*100:.1f}%)")
    print(f"  有 Validation: {total_validation:,} 行 ({total_validation/(ws.max_row-1)*100:.1f}%)")
    print(f"  只有 Ref 沒有 Validation: {total_ref_in_reply - total_validation:,} 行")

print("\n" + "=" * 100)
