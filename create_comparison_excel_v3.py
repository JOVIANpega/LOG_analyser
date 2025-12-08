# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import sys
import os

# 設定輸出編碼為 UTF-8
sys.stdout.reconfigure(encoding='utf-8')

def sanitize_excel_value(val):
    """防止 Excel 注入攻擊或格式錯誤 (例如以 = 開頭)"""
    if not val: return ""
    s = str(val)
    if s.startswith(('=', '+', '-', '@')):
        return "'" + s
    return s

input_file = r'MINE\result\FAIL匯總.xlsx'
output_file = r'MINE\result\FAIL匯總_NewFormat_v3.xlsx' # 改用 v3 避免鎖定

print(f"正在讀取原始檔案: {input_file}")

try:
    # 讀取原始 Excel
    wb = openpyxl.load_workbook(input_file)
    
    if 'Summary' not in wb.sheetnames:
        print("錯誤: 找不到 Summary 工作表")
        sys.exit(1)
        
    ws_orig = wb['Summary']
    
    extracted_data = []
    processed_logs = set() # 用於去重
    
    # 遍歷 Summary 的每一行提取資料
    for row in range(1, ws_orig.max_row + 1):
        col1_val = ws_orig.cell(row=row, column=1).value
        col2_val = ws_orig.cell(row=row, column=2).value # 詳細錯誤原因
        
        if not col1_val or not isinstance(col1_val, str):
            continue
            
        # 嚴格過濾無關行
        # 1. 過濾標題和分隔線
        if "錯誤原因統計" in col1_val or "──" in col1_val or "回到 Summary" in col1_val or "工作表快速連結" in col1_val:
            continue
        # 2. 過濾統計行 (例如 "TRY_TEXT is Fail = 2 筆")
        if " = " in col1_val and "筆" in col1_val:
            continue
        # 3. 過濾只有檔名但沒有錯誤內容的行 (通常是連結區塊)
        if not col2_val and "test" in col1_val:
            continue
            
        # 解析 Log 資訊
        log_info = col1_val
        
        # 去重檢查
        if log_info in processed_logs:
            continue
        processed_logs.add(log_info)
        
        # 預設值
        isn = ""
        station = ""
        fail_item = ""
        fail_reason = ""
        suggestion = "提供圖片給PEGA 看"
        
        # 解析 ISN 和 Station
        parts = log_info.split('-')
        
        # Station
        if '+' in parts[0]:
            station = parts[0].split('+')[1].strip()
        else:
            station = parts[0].strip()
            
        # ISN
        for part in parts:
            if part.startswith('WE') and len(part) > 8:
                isn = part
                break
        if not isn and len(parts) > 1:
            isn = parts[1]
            
        # 解析錯誤訊息
        full_error = str(col2_val) if col2_val else ""
        
        # 清理分隔線，但保留內容
        full_error = full_error.replace("===============錯誤原因====================\n\n", "")
        full_error = full_error.replace("==================================================\n\n", "")
        # 移除可能殘留的分隔線
        full_error = re.sub(r'={10,}', '', full_error)
        
        lines = full_error.split('\n')
        clean_lines = [l.strip() for l in lines if l.strip()]
        
        if clean_lines:
            # 預設行為
            first_line = clean_lines[0]
            
            # --- 增強的解析邏輯 ---
            
            # 1. 檢查 "is Fail"
            # 範例: "48:12 [1] B7PL011-202:TRY_TEXT is Fail" -> Item: "TRY_TEXT is Fail", Reason: 後面
            if "is Fail" in first_line:
                if ':' in first_line:
                    # 取最後一個冒號後面，如果長度合理
                    possible_item = first_line.split(':')[-1].strip()
                    fail_item = possible_item if possible_item else first_line
                else:
                    fail_item = first_line
                fail_reason = "\n".join(clean_lines[1:])
                
            # 2. 檢查 "doesn't match"
            # 範例: "doesn't match @SPEC_PASS" -> Item: "doesn't match @SPEC_PASS"
            elif "doesn't match" in first_line:
                fail_item = first_line # 保留完整描述，如 doesn't match @SPEC_PASS
                fail_reason = "\n".join(clean_lines[1:])
                
            # 3. 檢查副檔名 (Dump file issues)
            # 範例: dump_4cam_bs_snap_hdr_dre_0.yuv
            elif any(first_line.lower().endswith(ext) for ext in ['.yuv', '.bmp', '.jpg', '.png', '.log', '.txt']):
                fail_item = first_line
                fail_reason = "\n".join(clean_lines[1:])
                # 特殊處理：如果有 "拍照issue" 等關鍵字在內容中，可以提取出來放在前面 (Optional)
                
            # 4. 其他情況
            else:
                fail_item = first_line
                fail_reason = "\n".join(clean_lines[1:])
                
        # 清理 Reason，移除重複的 Item 資訊
        # 有時候 Error Code 包含 Item，這裡做簡單去重
        if fail_reason and fail_item in fail_reason:
             fail_reason = fail_reason.replace(fail_item, "").strip()

        
        # 如果 FAIL Reason 為空，嘗試用完整內容填補
        if not fail_reason and fail_item:
             # 如果只有一行，且被當作 Item，Reason 就留空或複製
             pass

        extracted_data.append({
            "ISN": isn,
            "Station": station,
            "FAIL Item": fail_item,
            "FAIL Reason": fail_reason,
            "suggestion": suggestion,
            "Original_Log_Name": log_info
        })

    # 建立新的工作表
    if 'Dashboard' in wb.sheetnames:
        wb.remove(wb['Dashboard'])
    ws_new = wb.create_sheet('Dashboard', 0)
    
    headers = ['ISN', 'Station', 'FAIL Item', 'FAIL Reason', 'suggestion', 'Full Log']
    ws_new.append(headers)
    
    # 樣式設定
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(name='Microsoft JhengHei', size=12, color="FFFFFF", bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True) # 靠上對齊
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_new.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    for row_idx, data in enumerate(extracted_data, 2):
        # 寫入資料
        ws_new.cell(row=row_idx, column=1, value=sanitize_excel_value(data['ISN'])).alignment = center_align
        ws_new.cell(row=row_idx, column=2, value=sanitize_excel_value(data['Station'])).alignment = center_align
        ws_new.cell(row=row_idx, column=3, value=sanitize_excel_value(data['FAIL Item'])).alignment = left_align
        ws_new.cell(row=row_idx, column=4, value=sanitize_excel_value(data['FAIL Reason'])).alignment = left_align
        ws_new.cell(row=row_idx, column=5, value=sanitize_excel_value(data['suggestion'])).alignment = center_align
        
        c6 = ws_new.cell(row=row_idx, column=6, value="查看 Log")
        c6.font = Font(color="0000FF", underline="single")
        c6.alignment = center_align
        
        # 連結
        target_sheet = None
        for sheet_name in wb.sheetnames:
            if sheet_name not in ['Summary', 'Dashboard']:
                clean_sheet = sheet_name.split('(')[0].strip()
                if clean_sheet in data['Original_Log_Name']:
                    target_sheet = sheet_name
                    break
        if target_sheet:
            c6.hyperlink = f"#'{target_sheet}'!A1"
        else:
            c6.value = "無連結"
            c6.font = Font(color="000000")

        # 邊框與背景
        fill_color = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") if row_idx % 2 == 1 else None
        for col in range(1, 7):
            cell = ws_new.cell(row=row_idx, column=col)
            cell.border = border
            if fill_color:
                cell.fill = fill_color

    # 欄寬
    ws_new.column_dimensions['A'].width = 20
    ws_new.column_dimensions['B'].width = 20
    ws_new.column_dimensions['C'].width = 40
    ws_new.column_dimensions['D'].width = 80 # 加寬 Reason 欄位
    ws_new.column_dimensions['E'].width = 25
    ws_new.column_dimensions['F'].width = 15

    ws_new.freeze_panes = 'A2'
    wb.save(output_file)
    print(f"已生成優化檔案: {output_file}")
    print(f"共提取 {len(extracted_data)} 筆資料")

except Exception as e:
    print(f"錯誤: {e}")
