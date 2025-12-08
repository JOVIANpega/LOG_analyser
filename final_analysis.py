# -*- coding: utf-8 -*-
import openpyxl
import sys
import re

# 設定輸出編碼為 UTF-8
sys.stdout.reconfigure(encoding='utf-8')

# 讀取測試腳本 Excel 檔案
file_path = 'MINE/VALO360_TestFlow_AOCI_4Cam_Stitching_test1_FATP_250924B.xlsx'
wb = openpyxl.load_workbook(file_path)

# 分析測試腳本工作表
sheet_name = '4cam stitching1 test'
ws = wb[sheet_name]

print("=" * 120)
print("Excel 測試腳本完整分析報告")
print("=" * 120)

# 統計 Ref 和 Validation
ref_col = 10  # Ref 欄位
val_col = 11  # Validation 欄位

total_rows = ws.max_row - 1  # 扣除標題行
ref_count = 0
val_count = 0
both_count = 0
only_ref_count = 0
only_val_count = 0
neither_count = 0

ref_examples = []
val_examples = []
only_ref_examples = []

for row in range(2, ws.max_row + 1):
    ref_value = ws.cell(row, ref_col).value
    val_value = ws.cell(row, val_col).value
    
    has_ref = ref_value and str(ref_value).strip()
    has_val = val_value and str(val_value).strip()
    
    if has_ref:
        ref_count += 1
        if len(ref_examples) < 10:
            ref_examples.append((row, str(ref_value)[:120]))
    
    if has_val:
        val_count += 1
        if len(val_examples) < 10:
            val_examples.append((row, str(val_value)[:120]))
    
    if has_ref and has_val:
        both_count += 1
    elif has_ref and not has_val:
        only_ref_count += 1
        if len(only_ref_examples) < 20:
            only_ref_examples.append((row, str(ref_value)[:120]))
    elif not has_ref and has_val:
        only_val_count += 1
    else:
        neither_count += 1

print(f"\n📊 基本統計:")
print(f"  總測試步驟: {total_rows:,} 行")
print(f"  總欄位數: {ws.max_column} 列")

print(f"\n📈 Ref 欄位統計 (第 {ref_col} 列):")
print(f"  有使用 Ref: {ref_count:,} 行 ({ref_count/total_rows*100:.1f}%)")
print(f"  未使用 Ref: {total_rows - ref_count:,} 行 ({(total_rows - ref_count)/total_rows*100:.1f}%)")

print(f"\n📉 Validation 欄位統計 (第 {val_col} 列):")
print(f"  有使用 Validation: {val_count:,} 行 ({val_count/total_rows*100:.1f}%)")
print(f"  未使用 Validation: {total_rows - val_count:,} 行 ({(total_rows - val_count)/total_rows*100:.1f}%)")

print(f"\n🔍 交叉分析:")
print(f"  同時有 Ref 和 Validation: {both_count:,} 行 ({both_count/total_rows*100:.1f}%)")
print(f"  只有 Ref 沒有 Validation: {only_ref_count:,} 行 ({only_ref_count/total_rows*100:.1f}%) ⚠️")
print(f"  只有 Validation 沒有 Ref: {only_val_count:,} 行 ({only_val_count/total_rows*100:.1f}%)")
print(f"  兩者都沒有: {neither_count:,} 行 ({neither_count/total_rows*100:.1f}%)")

if ref_count > 0:
    print(f"\n⚠️  關鍵發現:")
    print(f"  在 {ref_count:,} 個有 Ref 的測試步驟中:")
    print(f"    - 有 Validation: {both_count:,} 行 ({both_count/ref_count*100:.1f}%)")
    print(f"    - 沒有 Validation: {only_ref_count:,} 行 ({only_ref_count/ref_count*100:.1f}%) ⚠️")
    print(f"\n  這代表 {only_ref_count/ref_count*100:.1f}% 的數據提取缺少品質檢查!")

print(f"\n💡 改進潛力:")
if only_ref_count > 0:
    new_val_count = val_count + only_ref_count
    print(f"  如果為這 {only_ref_count:,} 行添加 Validation:")
    print(f"    - Validation 使用率將從 {val_count/total_rows*100:.1f}% 提升到 {new_val_count/total_rows*100:.1f}%")
    print(f"    - 提升幅度: {(new_val_count - val_count)/total_rows*100:.1f} 個百分點")
    print(f"    - 測試覆蓋率提升: {(new_val_count - val_count)/val_count*100:.0f}%")

# 視覺化統計
print(f"\n\n📊 視覺化統計:")
print("=" * 120)

def print_bar(label, count, total, width=60):
    percentage = count / total * 100
    filled = int(width * count / total)
    bar = "█" * filled + "░" * (width - filled)
    print(f"  {label:35s} [{bar}] {count:5,} ({percentage:5.1f}%)")

print("\nRef 欄位使用情況:")
print_bar("✓ 有使用 Ref", ref_count, total_rows)
print_bar("✗ 未使用 Ref", total_rows - ref_count, total_rows)

print("\nValidation 欄位使用情況:")
print_bar("✓ 有使用 Validation", val_count, total_rows)
print_bar("✗ 未使用 Validation", total_rows - val_count, total_rows)

print("\n交叉分析:")
print_bar("✓ Ref + Validation (完整)", both_count, total_rows)
print_bar("⚠ 只有 Ref (缺少驗證)", only_ref_count, total_rows)
print_bar("? 只有 Validation (異常)", only_val_count, total_rows)
print_bar("✗ 兩者都沒有", neither_count, total_rows)

# 顯示範例
print(f"\n\n📝 Ref 範例 (前 10 個):")
print("-" * 120)
for row_num, content in ref_examples:
    print(f"  Row {row_num:4d}: {content}")

print(f"\n\n📝 Validation 範例 (前 10 個):")
print("-" * 120)
for row_num, content in val_examples:
    print(f"  Row {row_num:4d}: {content}")

if only_ref_examples:
    print(f"\n\n⚠️  只有 Ref 沒有 Validation 的範例 (前 20 個):")
    print("-" * 120)
    for row_num, content in only_ref_examples:
        print(f"  Row {row_num:4d}: {content}")

print("\n" + "=" * 120)
print("報告完成")
print("=" * 120)
