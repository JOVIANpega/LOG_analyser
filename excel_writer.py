# excel_writer.py
# 用途：將log分析結果匯出為Excel檔案，支援PASS/FAIL分頁，欄位依規格
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles.borders import Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
import re
import os

class ExcelWriter:
    def __init__(self):
        pass

    def _sanitize_cell_text(self, value: object) -> str:
        """清理欲寫入儲存格的文字：
        - 轉為字串
        - 移除非法控制字元 (openpyxl 限制)
        - 去除 ANSI/ESC 序列 (更廣泛)
        - 截斷過長文字 (3 萬字元)
        """
        try:
            text = '' if value is None else str(value)
            # 移除 Excel 禁用控制字元 (0x00-0x1F 無法接受，openpyxl 內建 regex)
            text = ILLEGAL_CHARACTERS_RE.sub('', text)
            # 移除常見 ANSI CSI 序列: ESC [ ... letter
            text = re.sub(r'\x1b\[[0-9;?]*[A-Za-z]', '', text)
            # 移除其他 ESC 開頭的短序列: ESC ... 單個字母終止
            text = re.sub(r'\x1b[^A-Za-z]{0,20}[A-Za-z]', '', text)
            # 移除顏色碼樣式殘餘 (保險)
            text = re.sub(r'\x1b', '', text)
            # 若首字為會被當作公式的危險字元，前綴 '\''
            if text and text[0] in ('=', '+', '-', '@', '>', '<'):
                text = "'" + text
            # 長度截斷 (32767為Excel上限，保留安全餘量)
            if len(text) > 30000:
                text = text[:30000]
            return text
        except Exception:
            return '' if value is None else str(value)[:30000]

    def export(self, pass_items, fail_items, output_path):
        """
        匯出分析結果到Excel，分為PASS/FAIL兩個sheet
        pass_items: List[dict]
        fail_items: List[dict]
        """
        wb = Workbook()
        # 移除預設工作表
        wb.remove(wb.active)
        
        # PASS工作表
        ws_pass = wb.create_sheet('PASS')
        pass_headers = ['測項名稱', '指令', '收到指令', 'PASS/FAIL', '執行時間']
        ws_pass.append(pass_headers)
        
        for item in pass_items:
            row = [
                self._sanitize_cell_text(item.get('step_name', '')),
                self._sanitize_cell_text(item.get('command', '')),
                self._sanitize_cell_text(item.get('response', '')),
                'PASS',
                self._sanitize_cell_text(item.get('execution_time', ''))
            ]
            ws_pass.append(row)
        
        # FAIL工作表
        ws_fail = wb.create_sheet('FAIL')
        fail_headers = ['測項名稱', '指令', '錯誤回應', 'Retry次數', 'FAIL原因', '錯誤碼']
        ws_fail.append(fail_headers)
        
        for item in fail_items:
            row = [
                self._sanitize_cell_text(item.get('step_name', '')),
                self._sanitize_cell_text(item.get('command', '')),
                self._sanitize_cell_text(item.get('response', '')),
                item.get('retry', 0),
                self._sanitize_cell_text(item.get('error', '')),
                self._sanitize_cell_text(item.get('error_code', ''))
            ]
            ws_fail.append(row)
        
        # 設定欄寬
        for ws in [ws_pass, ws_fail]:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
        
        # 設定標題樣式
        for ws in [ws_pass, ws_fail]:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 儲存檔案
        wb.save(output_path)
        return output_path

    def _extract_total_secs(self, raw_lines: list) -> float | None:
        """從 raw_lines 嘗試提取測試總時間（秒數）"""
        try:
            for line in (raw_lines[-50:] if len(raw_lines) > 50 else raw_lines):
                if 'testtime' in line.lower() or 'total time' in line.lower():
                    # 嘗試提取數字
                    nums = re.findall(r'(\d+\.?\d*)', line)
                    if nums:
                        val = float(nums[-1])
                        if val > 0:
                            return val
            return None
        except Exception:
            return None

    def _unique_sheet_name(self, wb, base_name: str) -> str:
        """確保工作表名稱不重複"""
        if len(base_name) > 31:
            base_name = base_name[:28] + '...'
        existing = [ws.title for ws in wb.worksheets]
        if base_name not in existing:
            return base_name
        counter = 1
        while f"{base_name}({counter})" in existing:
            counter += 1
        suffix = f"({counter})"
        max_base = 31 - len(suffix)
        return f"{base_name[:max_base]}{suffix}"

    def export_pass_fail_workbooks(self, folder_path: str, pass_logs: list, fail_logs: list):
        """
        輸出兩個活頁簿：
        - PASS匯總.xlsx：Summary + 每個 PASS LOG 的工作表
        - FAIL匯總.xlsx：Summary + 每個 FAIL LOG 的工作表
        pass_logs/fail_logs 需提供：
          [{
             'file_path': 絕對或相對路徑,
             'file_name': 檔名,
             'raw_lines': [原始行...],
             'ui_annotations': [{line_idx, line_content, color, ...}],
             'pass_items': [...],
             'fail_items': [...],
             'summary': { '測試日期時間': str, 'SFIS': 'ON'|'OFF', '測試總時間': str, 'FAIL原因': str可選 }
          }, ...]
        """
        pass_path = os.path.join(folder_path, 'PASS匯總.xlsx')
        fail_path = os.path.join(folder_path, 'FAIL匯總.xlsx')
        self._build_pass_workbook(pass_path, pass_logs)
        self._build_fail_workbook(fail_path, fail_logs)
        return pass_path, fail_path

    def _format_filename_with_timestamp(self, base_name: str) -> str:
        """將檔名中的連續14位時間戳 YYYYMMDDHHMMSS 轉為 YYYY-MMDD-HHMMSS 格式。找不到則原樣回傳。"""
        try:
            m = re.search(r'(20\d{12})', base_name)
            if not m:
                return base_name
            s = m.group(1)
            y, md, hms = s[:4], s[4:8], s[8:]
            return base_name.replace(s, f"{y}-{md}-{hms}")
        except Exception:
            return base_name

    def _build_preview_comment(self, entry: dict) -> str:
        """產生懸停預覽內容：顯示對應工作表名稱與原始LOG前幾行（加上簡易標記）。"""
        try:
            sheet_or_name = entry.get('file_name') or 'LOG'
            raw = entry.get('raw_lines') or []
            preview_lines = []
            header = f"對應工作表: {sheet_or_name}\n******** 預覽 ********"
            preview_lines.append(header)
            for i, line in enumerate(raw[:15], 1):
                s = str(line)
                # 簡易高亮標記
                if 'Do @STEP' in s:
                    s = f"[STEP] {s}"
                if 'FAIL' in s.upper() or 'ERROR' in s.upper():
                    s = f"[FAIL] {s}"
                s = re.sub(r'^\s*>\s*', '▶ ', s)
                s = re.sub(r'^\s*<\s*', '◀ ', s)
                preview_lines.append(s)
            preview_lines.append('************************')
            return '\n'.join(preview_lines)
        except Exception:
            return '對應工作表預覽不可用'

    def _add_input_prompt(self, ws, cell, title: str, message: str):
        """在儲存格上加一個資料驗證提示（白底有框），當選取時顯示。"""
        try:
            # Excel 對於訊息有長度限制，做適當截斷
            msg = (message or '')
            if len(msg) > 250:
                msg = msg[:250] + '…'
            dv = DataValidation(type="custom", formula1="TRUE", allow_blank=True, showInputMessage=True)
            dv.promptTitle = title[:30] if title else ''
            dv.prompt = msg
            ws.add_data_validation(dv)
            dv.add(cell.coordinate)
        except Exception:
            pass

    # 內部：PASS 匯總活頁簿
    def _build_pass_workbook(self, output_path: str, logs: list):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Summary'
        # 設定標籤顏色（紅色）
        try:
            ws.sheet_properties.tabColor = 'FFFF0000'
        except Exception:
            pass
        header_font = Font(name='Microsoft JhengHei', size=16, bold=True, color='FFFFFFFF')
        normal_font = Font(name='Microsoft JhengHei', size=10)
        center = Alignment(horizontal='center', vertical='center')
        deep_green = PatternFill('solid', fgColor='FF1B5E20')
        # 先建立各 LOG 原始工作表，並記錄 sheet 名稱
        sheet_map = {}
        for entry in logs:
            sheet_name_base = entry.get('file_name', 'LOG')
            sheet_name = self._unique_sheet_name(wb, sheet_name_base)
            sheet_map[entry.get('file_name')] = sheet_name
            ws2 = wb.create_sheet(title=sheet_name)
            cell = ws2.cell(row=1, column=1, value=self._sanitize_cell_text(entry.get('file_name')))
            cell.font = Font(name='Microsoft JhengHei', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='FF2ECC71')
            cell.number_format = '@'
            step_labels = [str(i) for i, _ in enumerate(entry.get('pass_items') or [], 1)]
            ws2.cell(row=2, column=1, value=self._sanitize_cell_text('，'.join(step_labels) if step_labels else ''))
            ws2.cell(row=2, column=1).number_format = '@'
            self._write_raw_log_with_annotations(ws2, start_row=3, raw_lines=entry.get('raw_lines') or [], annotations=entry.get('ui_annotations') or [], font=Font(name='Microsoft JhengHei', size=10), step_marks=entry.get('step_marks'))
            self._auto_fit_columns(ws2)
        # Summary 表格：
        headers = ['檔名'] + [f'步驟 {i}' for i in range(1, 11)]
        for c, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=header)
            cell.font = header_font
            cell.alignment = center
            cell.fill = deep_green
        ws.freeze_panes = 'A2'
        max_chunks = 10
        thin = Side(border_style='thin', color='FF888888')
        thick = Side(border_style='thick', color='FF000000')
        start_data_row = ws.max_row + 1
        for entry in logs:
            # 分解steps至10欄
            base = self._sanitize_cell_text(entry.get('file_name') or '')
            base_fmt = self._format_filename_with_timestamp(base)
            sfis = (entry.get('summary') or {}).get('SFIS','')
            sfis = (sfis or '').upper()
            secs = self._extract_total_secs(entry.get('raw_lines') or [])
            sec_txt = f"測試總時間:{secs:.1f} Sec." if secs is not None else ''
            suffix = f"_SFIS_{sfis}" if sfis else ''
            display_name = f"{base_fmt}{suffix} {sec_txt}".strip()
            r = ws.max_row + 1
            # 主檔名
            cell_name = ws.cell(row=r, column=1, value=self._sanitize_cell_text(display_name))
            cell_name.number_format='@'
            cell_name.font = Font(name='Microsoft JhengHei', size=10, color='FF000000')
            cell_name.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top', shrink_to_fit=True)
            # 備註與超連結
            sheet = sheet_map.get(entry.get('file_name'))
            try:
                cell_name.comment = Comment(self._build_preview_comment(entry), "LOG Analyzer")
                cell_name.comment.width = 400
                cell_name.comment.height = 500
            except Exception:
                pass
            if sheet:
                cell_name.hyperlink = f"#'{sheet}'!A1"
                self._add_input_prompt(ws, cell_name, '對應工作表', sheet)
            # 步驟欄位（最多10欄）
            pass_steps = entry.get('pass_items') or []
            for col_idx in range(2, max_chunks + 2):
                step_idx = col_idx - 2
                if step_idx < len(pass_steps):
                    step = pass_steps[step_idx]
                    step_text = f"{step.get('step_name','')}\n{step.get('command','')}\n{step.get('response','')}"
                    cell = ws.cell(row=r, column=col_idx, value=self._sanitize_cell_text(step_text))
                    cell.number_format='@'
                    cell.font = normal_font
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left', shrink_to_fit=True)
                else:
                    cell = ws.cell(row=r, column=col_idx, value='')
                    cell = ws.cell(row=r, column=col_idx, value='')
                    cell.number_format='@'
                    cell.font = normal_font
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left', shrink_to_fit=True)
            last_col = 1 + max_chunks
            for c in range(1, last_col+1):
                ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # 外框粗線
        end_r = ws.max_row
        last_col = 1 + max_chunks
        if end_r >= start_data_row:
            for c in range(1, last_col+1):
                ws.cell(row=start_data_row, column=c).border = ws.cell(row=start_data_row, column=c).border.copy(top=thick)
                ws.cell(row=end_r, column=c).border = ws.cell(row=end_r, column=c).border.copy(bottom=thick)
            for r in range(start_data_row, end_r+1):
                ws.cell(row=r, column=1).border = ws.cell(row=r, column=1).border.copy(left=thick)
                ws.cell(row=r, column=last_col).border = ws.cell(row=r, column=last_col).border.copy(right=thick)
        # 表格底部：SHEET 快速連結（點擊跳轉）
        link_title_row = ws.max_row + 2
        ws.cell(row=link_title_row, column=1, value='工作表快速連結（點擊跳轉）').font = Font(name='Microsoft JhengHei', size=11, bold=True)
        ws.cell(row=link_title_row, column=1).alignment = Alignment(horizontal='left')
        cur = link_title_row + 1
        for entry in logs:
            base = self._sanitize_cell_text(entry.get('file_name') or '')
            base_fmt = self._format_filename_with_timestamp(base)
            sfis = (entry.get('summary') or {}).get('SFIS', '')
            sfis = (sfis or '').upper()
            secs = self._extract_total_secs(entry.get('raw_lines') or [])
            sec_txt = f"測試總時間:{secs:.1f} Sec." if secs is not None else ''
            suffix = f"_SFIS_{sfis}" if sfis else ''
            display_name = f"{base_fmt}{suffix} {sec_txt}".strip()
            c = ws.cell(row=cur, column=1, value=self._sanitize_cell_text(display_name))
            c.number_format='@'
            c.font = Font(name='Microsoft JhengHei', size=11, color='FF0000FF', underline='single')
            c.alignment = Alignment(horizontal='left')
            sheet = sheet_map.get(entry.get('file_name'))
            if sheet:
                c.hyperlink = f"#'{sheet}'!A1"
                try:
                    c.comment = Comment(self._build_preview_comment(entry), 'LOG Analyzer')
                    c.comment.width = 400
                    c.comment.height = 500
                except Exception:
                    pass
                # 白底提示（提示箭頭在左上方，靠近視窗；Excel控制箭頭顯示位置有限）
                self._add_input_prompt(ws, c, '對應工作表', sheet)
            cur += 1
        # 更緊湊的欄寬
        min_widths = {1: 30}
        for i in range(2, last_col+1):
            min_widths[i] = 22
        self._auto_fit_columns(ws, min_widths=min_widths)
        wb.save(output_path)

    def _build_fail_workbook(self, output_path: str, logs: list):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Summary'
        # 設定標籤顏色（紅色）
        try:
            ws.sheet_properties.tabColor = 'FFFF0000'
        except Exception:
            pass
        header_font = Font(name='Microsoft JhengHei', size=16, bold=True, color='FFFFFFFF')
        normal_font = Font(name='Microsoft JhengHei', size=10)
        center = Alignment(horizontal='center', vertical='center')
        deep_green = PatternFill('solid', fgColor='FF1B5E20')
        headers = ['檔名', '詳細錯誤原因']
        ws.append(headers)
        for c in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.alignment = center
            cell.fill = deep_green
        ws.freeze_panes = 'A2'
        # 先建立各 LOG 原始工作表
        sheet_map = {}
        for entry in logs:
            sheet_name_base = entry.get('file_name', 'LOG')
            sheet_name = self._unique_sheet_name(wb, sheet_name_base)
            sheet_map[entry.get('file_name')] = sheet_name
            ws2 = wb.create_sheet(title=sheet_name)
            cell = ws2.cell(row=1, column=1, value=self._sanitize_cell_text(entry.get('file_name')))
            cell.font = Font(name='Microsoft JhengHei', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='FFE74C3C')
            cell.number_format = '@'
            # 在最上面添加回到Summary的快速連結
            back_to_summary_top = ws2.cell(row=2, column=1, value='🔙 回到 Summary 頁面')
            back_to_summary_top.number_format='@'
            back_to_summary_top.font = Font(name='Microsoft JhengHei', size=12, bold=True, color='FF008000', underline='single')
            back_to_summary_top.alignment = Alignment(horizontal='left')
            back_to_summary_top.hyperlink = f"#'Summary'!A1"
            back_to_summary_top.fill = PatternFill('solid', fgColor='FFE6FFE6')  # 淺綠色背景
            
            # 添加分隔線
            ws2.cell(row=3, column=1, value='─' * 50).font = Font(name='Microsoft JhengHei', size=10, color='FF808080')
            
            # 顯示完整錯誤原因區塊
            detailed_error = self._build_detailed_error_summary(entry)
            error_lines = detailed_error.split('\n')
            current_row = 4
            
            for line in error_lines:
                if line.strip():
                    cell = ws2.cell(row=current_row, column=1, value=self._sanitize_cell_text(line))
                    cell.number_format = '@'
                    
                    # 設定不同行的樣式
                    if "===============錯誤原因====================" in line:
                        cell.font = Font(name='Microsoft JhengHei', size=12, bold=True, color='FF0000')
                    elif "🔴 突出錯誤" in line:
                        cell.font = Font(name='Microsoft JhengHei', size=11, bold=True, color='FF0000')
                    elif "=" * 50 in line:
                        cell.font = Font(name='Microsoft JhengHei', size=10)
                    elif "執行指令:" in line:
                        cell.font = Font(name='Microsoft JhengHei', size=11, color='FF0000FF')
                    elif any(keyword in line.lower() for keyword in ['is fail', 'executes fail', "doesn't match", 'all test aborted']):
                        cell.font = Font(name='Microsoft JhengHei', size=11, bold=True, color='FF0000')
                        cell.fill = PatternFill('solid', fgColor='FFFFFF99')  # 淺黃色背景
                    else:
                        cell.font = Font(name='Microsoft JhengHei', size=11)
                    
                    current_row += 1
                else:
                    current_row += 1
            
            # 添加分隔線
            ws2.cell(row=current_row, column=1, value=self._sanitize_cell_text("=" * 60)).number_format='@'
            ws2.cell(row=current_row, column=1).font = Font(name='Microsoft JhengHei', size=10)
            current_row += 1
            
            # PASS步驟資訊
            pass_steps = [str(i) for i, _ in enumerate(entry.get('pass_items') or [], 1)]
            ws2.cell(row=current_row, column=1, value=self._sanitize_cell_text('PASS步驟: ' + '，'.join(pass_steps) if pass_steps else 'PASS步驟: 無'))
            ws2.cell(row=current_row, column=1).number_format='@'
            ws2.cell(row=current_row, column=1).font = Font(name='Microsoft JhengHei', size=11, color='FF008000')
            start_row = current_row + 1
            
            # 寫入原始LOG內容，並標記錯誤行
            self._write_raw_log_with_annotations(ws2, start_row=start_row, raw_lines=entry.get('raw_lines') or [], annotations=entry.get('ui_annotations') or [], font=Font(name='Microsoft JhengHei', size=11), step_marks=entry.get('step_marks'))
            self._auto_fit_columns(ws2)
        # 在Summary頁面最上面添加錯誤統計
        self._add_error_statistics(ws, logs)
        
        # Summary：每個LOG一行，顯示檔名和詳細錯誤原因
        thin = Side(border_style='thin', color='FF888888')
        thick = Side(border_style='thick', color='FF000000')
        start_data_row = ws.max_row + 1
        for entry in logs:
            # 檔名欄
            base = self._sanitize_cell_text(entry.get('file_name') or '')
            base_fmt = self._format_filename_with_timestamp(base)
            sfis = (entry.get('summary') or {}).get('SFIS','')
            sfis = (sfis or '').upper()
            secs = self._extract_total_secs(entry.get('raw_lines') or [])
            sec_txt = f"測試總時間:{secs:.1f} Sec." if secs is not None else ''
            suffix = f"_SFIS_{sfis}" if sfis else ''
            display_name = f"{base_fmt}{suffix} {sec_txt}".strip()
            r = ws.max_row + 1
            cell_name = ws.cell(row=r, column=1, value=self._sanitize_cell_text(display_name))
            cell_name.number_format='@'
            cell_name.font = Font(name='Microsoft JhengHei', size=10, color='FF000000')
            cell_name.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top', shrink_to_fit=True)
            # 備註與超連結
            sheet = sheet_map.get(entry.get('file_name'))
            try:
                cell_name.comment = Comment(self._build_preview_comment(entry), "LOG Analyzer")
                cell_name.comment.width = 400
                cell_name.comment.height = 500
            except Exception:
                pass
            if sheet:
                cell_name.hyperlink = f"#'{sheet}'!A1"
                # 白底提示
                self._add_input_prompt(ws, cell_name, '對應工作表', entry.get('file_name') or '')
            
            # 詳細錯誤原因欄 - 包含主要錯誤和執行指令
            detailed_error = self._build_detailed_error_summary(entry)
            cell_reason = ws.cell(row=r, column=2, value=self._sanitize_cell_text(detailed_error))
            cell_reason.number_format='@'
            cell_reason.font = Font(name='Microsoft JhengHei', size=11)
            cell_reason.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left', shrink_to_fit=False)
            
            # 如果錯誤原因包含 "doesn't match"，用特殊格式突出顯示
            if "doesn't match" in detailed_error.lower():
                cell_reason.font = Font(name='Microsoft JhengHei', size=11, bold=True, color='FFFF0000')
                cell_reason.fill = PatternFill('solid', fgColor='FFFFFF99')  # 淺黃色背景
            
            for c in range(1, 2+1):
                ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        end_r = ws.max_row
        if end_r >= start_data_row:
            for c in range(1, 2+1):
                ws.cell(row=start_data_row, column=c).border = ws.cell(row=start_data_row, column=c).border.copy(top=thick)
                ws.cell(row=end_r, column=c).border = ws.cell(row=end_r, column=c).border.copy(bottom=thick)
            for r in range(start_data_row, end_r+1):
                ws.cell(row=r, column=1).border = ws.cell(row=r, column=1).border.copy(left=thick)
                ws.cell(row=r, column=2).border = ws.cell(row=r, column=2).border.copy(right=thick)
        # 表格底部：SHEET 快速連結
        link_title_row = ws.max_row + 2
        ws.cell(row=link_title_row, column=1, value='工作表快速連結（點擊跳轉）').font = Font(name='Microsoft JhengHei', size=10, bold=True)
        ws.cell(row=link_title_row, column=1).alignment = Alignment(horizontal='left')
        cur = link_title_row + 1
        
        # 添加回到Summary的快速連結
        back_to_summary = ws.cell(row=cur, column=1, value='🔙 回到 Summary 頁面')
        back_to_summary.number_format='@'
        back_to_summary.font = Font(name='Microsoft JhengHei', size=12, bold=True, color='FF008000', underline='single')
        back_to_summary.alignment = Alignment(horizontal='left')
        back_to_summary.hyperlink = f"#'Summary'!A1"
        back_to_summary.fill = PatternFill('solid', fgColor='FFE6FFE6')  # 淺綠色背景
        cur += 1
        
        # 添加分隔線
        ws.cell(row=cur, column=1, value='─' * 50).font = Font(name='Microsoft JhengHei', size=10, color='FF808080')
        cur += 1
        
        for entry in logs:
            base = self._sanitize_cell_text(entry.get('file_name') or '')
            base_fmt = self._format_filename_with_timestamp(base)
            sfis = (entry.get('summary') or {}).get('SFIS','')
            sfis = (sfis or '').upper()
            secs = self._extract_total_secs(entry.get('raw_lines') or [])
            sec_txt = f"測試總時間:{secs:.1f} Sec." if secs is not None else ''
            suffix = f"_SFIS_{sfis}" if sfis else ''
            display_name = f"{base_fmt}{suffix} {sec_txt}".strip()
            c = ws.cell(row=cur, column=1, value=self._sanitize_cell_text(display_name))
            c.number_format='@'
            c.font = Font(name='Microsoft JhengHei', size=10, color='FF0000FF', underline='single')
            c.alignment = Alignment(horizontal='left')
            sheet = sheet_map.get(entry.get('file_name'))
            if sheet:
                c.hyperlink = f"#'{sheet}'!A1"
                try:
                    c.comment = Comment(self._build_preview_comment(entry), 'LOG Analyzer')
                    c.comment.width = 400
                    c.comment.height = 500
                except Exception:
                    pass
                self._add_input_prompt(ws, c, '對應工作表', sheet)
            cur += 1
        self._auto_fit_columns(ws, min_widths={1: 30, 2: 80})
        wb.save(output_path)

    def _write_raw_log_with_annotations(self, ws, start_row: int, raw_lines: list, annotations: list, font: Font, step_marks: dict | None = None):
        color_map = {
            'black': 'FF000000',
            'red': 'FFE74C3C',
            'green': 'FF2ECC71',
            'blue': 'FF3498DB',
            'purple': 'FF9B59B6',
        }
        marks = step_marks or {}
        error_line_found = False
        first_error_row = None
        
        # 以 annotations 的 color 欄位對應文字顏色；在步驟起始行前加上 1. 2. ... 標號
        for i, raw in enumerate(raw_lines, start=start_row):
            src_idx = i - start_row  # 對應原始 raw_lines 索引
            # 使用統一的清理函數
            line = str(raw)
            
            # 檢查是否為錯誤行
            is_error_line = False
            line_lower = line.lower()
            if any(keyword in line_lower for keyword in [
                'is fail', 'segmentation fault', 'core dumped', 'executes fail', 
                "doesn't match", 'timeout', 'exception', 'error', 'fail'
            ]):
                is_error_line = True
                if not error_line_found:
                    first_error_row = i
                    error_line_found = True
            
            # 先檢查是否有步驟標號
            if src_idx in marks:
                line = f"{marks[src_idx]}. {line}"
            
            # 清理行內容
            line = self._sanitize_cell_text(line)
            
            # 套用樣式：找到對應的 annotation
            cell = ws.cell(row=i, column=1, value=line)
            cell.font = font
            cell.number_format = '@'
            
            # 突出顯示錯誤行
            if is_error_line:
                cell.font = Font(name=font.name, size=font.size, color='FF0000', bold=True)
                cell.fill = PatternFill('solid', fgColor='FFFFFF99')  # 淺黃色背景
            
            # 預設
            found_anno = None
            for anno in annotations:
                if anno.get('line_idx') == src_idx:
                    found_anno = anno
                    break
            if found_anno and 'color' in found_anno:
                color_name = found_anno['color']
                if color_name in color_map:
                    color_hex = color_map[color_name]
                    cell.font = Font(name=font.name, size=font.size, color=color_hex, bold=font.bold)
        
        # 如果有找到錯誤行，設定超連結到第一個錯誤行
        if first_error_row:
            # 在A1儲存格添加超連結到第一個錯誤行
            ws.cell(row=1, column=1).hyperlink = f"#{ws.title}!A{first_error_row}"
        
        # 在每個LOG工作表底部添加回到Summary的快速連結
        last_row = ws.max_row + 2
        back_to_summary = ws.cell(row=last_row, column=1, value='🔙 回到 Summary 頁面')
        back_to_summary.number_format='@'
        back_to_summary.font = Font(name='Microsoft JhengHei', size=12, bold=True, color='FF008000', underline='single')
        back_to_summary.alignment = Alignment(horizontal='left')
        back_to_summary.hyperlink = f"#'Summary'!A1"
        back_to_summary.fill = PatternFill('solid', fgColor='FFE6FFE6')  # 淺綠色背景

    def _build_detailed_error_summary(self, entry: dict) -> str:
        """建立詳細的錯誤摘要，包含主要錯誤和執行指令"""
        try:
            fail_items = entry.get('fail_items', [])
            raw_lines = entry.get('raw_lines', [])
            
            if not fail_items and not raw_lines:
                return "未知錯誤"
            
            # 收集所有錯誤資訊
            error_details = []
            commands = []
            critical_errors = []
            herr_errors = []
            
            # 從 raw_lines 中提取錯誤資訊
            for line in raw_lines:
                line_str = str(line).strip()
                if not line_str:
                    continue
                    
                line_lower = line_str.lower()
                
                # 收集 HERR 錯誤
                if 'herr:' in line_lower:
                    herr_errors.append(line_str)
                
                # 收集嚴重錯誤
                if any(keyword in line_lower for keyword in [
                    'segmentation fault', 'core dumped', 'executes fail', 
                    "doesn't match", 'timeout', 'exception', 'all test aborted'
                ]):
                    critical_errors.append(line_str)
                
                # 收集一般錯誤
                if any(keyword in line_lower for keyword in [
                    'is fail', 'error', 'fail'
                ]):
                    if line_str not in error_details:
                        error_details.append(line_str)
            
            # 從 fail_items 中提取資訊
            for item in fail_items:
                # 主要錯誤原因
                error = item.get('error', '')
                if error and error not in error_details:
                    error_details.append(error)
                
                # 執行的指令
                command = item.get('command', '')
                if command and command not in commands:
                    commands.append(command)
                
                # 完整回應中的錯誤資訊
                full_response = item.get('full_response', '')
                if full_response:
                    lines = full_response.split('\n')
                    for line in lines:
                        line = line.strip()
                        if not line:
                            continue
                        line_lower = line.lower()
                        # 尋找關鍵錯誤
                        if any(keyword in line_lower for keyword in [
                            'is fail', 'segmentation fault', 'core dumped', 
                            'executes fail', "doesn't match", 'timeout', 'exception'
                        ]):
                            if line not in error_details:
                                error_details.append(line)
            
            # 組合詳細摘要
            summary_parts = []
            
            # 主要錯誤原因
            main_error = (entry.get('summary') or {}).get('FAIL原因', '')
            if main_error:
                summary_parts.append(f"===============錯誤原因====================")
                summary_parts.append("")
                
                # 突出顯示 "doesn't match" 錯誤
                if "doesn't match" in main_error.lower():
                    summary_parts.append("🔴 突出錯誤 (doesn't match):")
                    summary_parts.append(main_error)
                else:
                    summary_parts.append(main_error)
                
                summary_parts.append("")
                summary_parts.append("=" * 50)
                summary_parts.append("")
            
            # 嚴重錯誤
            if critical_errors:
                for error in critical_errors[:5]:  # 最多顯示5個嚴重錯誤
                    summary_parts.append(error)
                summary_parts.append("")
            
            # HERR 錯誤
            if herr_errors:
                for error in herr_errors[:10]:  # 最多顯示10個HERR錯誤
                    summary_parts.append(error)
                summary_parts.append("")
            
            # 其他錯誤詳情
            if error_details:
                for detail in error_details[:5]:  # 最多顯示5個錯誤詳情
                    if detail != main_error and detail not in summary_parts:
                        summary_parts.append(detail)
                summary_parts.append("")
            
            # 執行指令
            if commands:
                unique_commands = list(dict.fromkeys(commands))  # 去重
                for cmd in unique_commands[:3]:  # 最多顯示3個指令
                    summary_parts.append(f"執行指令: {cmd}")
            
            return '\n'.join(summary_parts) if summary_parts else "未知錯誤"
            
        except Exception as e:
            return f"錯誤摘要生成失敗: {str(e)}"
    
    def _add_error_statistics(self, ws, logs: list):
        """在Summary頁面最上面添加錯誤統計"""
        try:
            # 收集所有錯誤原因
            error_counts = {}
            
            for entry in logs:
                main_error = (entry.get('summary') or {}).get('FAIL原因', '')
                if main_error:
                    # 清理錯誤原因，提取主要部分
                    clean_error = self._extract_main_error_type(main_error)
                    if clean_error:
                        error_counts[clean_error] = error_counts.get(clean_error, 0) + 1
            
            if not error_counts:
                return
            
            # 在現有內容上方插入錯誤統計
            # 先將現有內容向下移動
            max_row = ws.max_row
            for row in range(max_row, 0, -1):
                for col in range(1, 3):  # 假設有2欄
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        ws.cell(row=row + len(error_counts) + 3, column=col, value=cell.value)
                        # 複製樣式
                        if cell.font:
                            ws.cell(row=row + len(error_counts) + 3, column=col).font = cell.font
                        if cell.fill:
                            ws.cell(row=row + len(error_counts) + 3, column=col).fill = cell.fill
                        if cell.alignment:
                            ws.cell(row=row + len(error_counts) + 3, column=col).alignment = cell.alignment
                        if cell.border:
                            ws.cell(row=row + len(error_counts) + 3, column=col).border = cell.border
                        # 清除原位置
                        ws.cell(row=row, column=col).value = None
            
            # 添加錯誤統計標題
            title_row = 1
            ws.cell(row=title_row, column=1, value='📊 錯誤原因統計').font = Font(name='Microsoft JhengHei', size=14, bold=True, color='FF0000')
            ws.cell(row=title_row, column=1).fill = PatternFill('solid', fgColor='FFFFE6E6')
            
            # 添加統計內容
            current_row = 2
            for error_type, count in sorted(error_counts.items(), key=lambda x: x[1], reverse=True):
                ws.cell(row=current_row, column=1, value=f'{error_type} = {count} 筆')
                ws.cell(row=current_row, column=1).font = Font(name='Microsoft JhengHei', size=11, bold=True)
                ws.cell(row=current_row, column=1).fill = PatternFill('solid', fgColor='FFFFFF99')  # 淺黃色背景
                current_row += 1
            
            # 添加分隔線
            ws.cell(row=current_row, column=1, value='─' * 50).font = Font(name='Microsoft JhengHei', size=10, color='FF808080')
            
        except Exception as e:
            print(f"添加錯誤統計時發生錯誤: {e}")
    
    def _extract_main_error_type(self, error_text: str) -> str:
        """從錯誤文字中提取主要錯誤類型"""
        try:
            if not error_text:
                return ""
            
            # 移除時間戳記和錯誤代碼
            clean_text = error_text
            
            # 移除時間戳記格式 (如 "2025/09/23 10:48:12 [1]")
            import re
            clean_text = re.sub(r'\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2} \[\d+\]', '', clean_text)
            
            # 移除錯誤代碼 (如 "<ErrorCode: AFFW57>")
            clean_text = re.sub(r'<ErrorCode: [^>]+>', '', clean_text)
            
            # 移除測試時間 (如 "----- 61.019937 Sec.")
            clean_text = re.sub(r'----- \d+\.\d+ Sec\.', '', clean_text)
            
            # 清理多餘空格
            clean_text = clean_text.strip()
            
            # 提取主要錯誤類型
            if ':' in clean_text:
                # 格式如 "B7PL011-202:Chec Frimware version is Fail"
                parts = clean_text.split(':', 1)
                if len(parts) > 1:
                    main_part = parts[1].strip()
                    # 移除測試編號前綴
                    main_part = re.sub(r'^[A-Z0-9]+-\d+:', '', main_part)
                    return main_part.strip()
            
            # 如果沒有冒號，直接返回清理後的文字
            return clean_text
            
        except Exception:
            return error_text

    def _auto_fit_columns(self, ws, min_widths: dict | None = None):
        """自動調整欄寬"""
        min_w = min_widths or {}
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            col_num = col[0].column
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max(max_length + 2, min_w.get(col_num, 10))
            adjusted_width = min(adjusted_width, 100)  # 最大寬度限制
            ws.column_dimensions[column].width = adjusted_width