# -*- coding: utf-8 -*-
import openpyxl
import sys

# 設定輸出編碼為 UTF-8
sys.stdout.reconfigure(encoding='utf-8')

# 讀取 Excel 檔案
wb = openpyxl.load_workbook('Analysis_CSV_FILE/VALO360_AOCI_4Cam_Stitching_test1_log_Analysis_CSV.xlsx')

print("=" * 80)
print("Excel 檔案分析報告")
print("=" * 80)

print(f"\n工作表名稱: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n\n{'=' * 80}")
    print(f"工作表: {sheet_name}")
    print(f"{'=' * 80}")
    print(f"總行數: {ws.max_row}")
    print(f"總列數: {ws.max_column}")
    
    # 讀取標題列
    headers = []
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        headers.append(header if header else f"Column_{col}")
    
    print(f"\n欄位標題 ({len(headers)} 個):")
    for i, h in enumerate(headers, 1):
        print(f"  {i}. {h}")
    
    # 顯示前 10 行資料
    print(f"\n前 10 行資料預覽:")
    print("-" * 80)
    
    for row_idx in range(1, min(11, ws.max_row + 1)):
        print(f"\n第 {row_idx} 行:")
        for col_idx in range(1, min(ws.max_column + 1, 15)):  # 最多顯示前 15 列
            cell_value = ws.cell(row_idx, col_idx).value
            if cell_value is not None:
                # 截斷過長的內容
                str_value = str(cell_value)
                if len(str_value) > 50:
                    str_value = str_value[:50] + "..."
                print(f"  {headers[col_idx-1]}: {str_value}")

    # 統計資料
    print(f"\n\n資料統計:")
    print("-" * 80)
    
    # 檢查 Ref 欄位
    ref_col_idx = None
    for i, h in enumerate(headers, 1):
        if h == "Ref":
            ref_col_idx = i
            break
    
    if ref_col_idx:
        ref_count = 0
        ref_examples = []
        for row in range(2, ws.max_row + 1):
            ref_value = ws.cell(row, ref_col_idx).value
            if ref_value and str(ref_value).strip():
                ref_count += 1
                if len(ref_examples) < 5:
                    ref_examples.append(str(ref_value))
        
        print(f"Ref 欄位有值的行數: {ref_count}")
        if ref_examples:
            print(f"Ref 範例:")
            for ex in ref_examples:
                print(f"  - {ex}")
    
    # 檢查 Validation 欄位
    val_col_idx = None
    for i, h in enumerate(headers, 1):
        if h == "Validation":
            val_col_idx = i
            break
    
    if val_col_idx:
        val_count = 0
        val_examples = []
        for row in range(2, ws.max_row + 1):
            val_value = ws.cell(row, val_col_idx).value
            if val_value and str(val_value).strip():
                val_count += 1
                if len(val_examples) < 5:
                    val_examples.append(str(val_value))
        
        print(f"\nValidation 欄位有值的行數: {val_count}")
        if val_examples:
            print(f"Validation 範例:")
            for ex in val_examples:
                print(f"  - {ex}")

print("\n" + "=" * 80)
print("分析完成")
print("=" * 80)
