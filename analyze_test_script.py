# -*- coding: utf-8 -*-
import openpyxl
import sys

# 設定輸出編碼為 UTF-8
sys.stdout.reconfigure(encoding='utf-8')

# 讀取測試腳本 Excel 檔案
file_path = 'MINE/VALO360_TestFlow_AOCI_4Cam_Stitching_test1_FATP_250924B.xlsx'
wb = openpyxl.load_workbook(file_path)

print("=" * 100)
print("測試腳本 Excel 檔案分析")
print("=" * 100)

print(f"\n檔案: {file_path}")
print(f"工作表名稱: {wb.sheetnames}")

# 分析每個工作表
for sheet_name in wb.sheetnames[:3]:  # 只分析前 3 個工作表
    ws = wb[sheet_name]
    print(f"\n\n{'=' * 100}")
    print(f"工作表: {sheet_name}")
    print(f"{'=' * 100}")
    print(f"總行數: {ws.max_row}")
    print(f"總列數: {ws.max_column}")
    
    # 讀取標題列 (假設在第一行)
    headers = []
    for col in range(1, min(ws.max_column + 1, 30)):  # 最多讀取 30 列
        header = ws.cell(1, col).value
        headers.append(header if header else f"Col_{col}")
    
    print(f"\n欄位標題 (前 30 個):")
    for i, h in enumerate(headers, 1):
        print(f"  {i}. {h}")
    
    # 尋找關鍵欄位的位置
    key_columns = {}
    for i, h in enumerate(headers, 1):
        if h in ['Send', 'Reply', 'Ref', 'Validation', 'send', 'reply', 'ref', 'validation']:
            key_columns[h] = i
    
    if key_columns:
        print(f"\n找到的關鍵欄位:")
        for col_name, col_idx in key_columns.items():
            print(f"  {col_name}: 第 {col_idx} 列")
    
    # 顯示前 20 行資料
    print(f"\n前 20 行資料預覽:")
    print("-" * 100)
    
    for row_idx in range(1, min(21, ws.max_row + 1)):
        print(f"\n第 {row_idx} 行:")
        
        # 只顯示有內容的欄位
        has_content = False
        for col_idx in range(1, min(len(headers) + 1, 30)):
            cell_value = ws.cell(row_idx, col_idx).value
            if cell_value is not None and str(cell_value).strip():
                has_content = True
                # 截斷過長的內容
                str_value = str(cell_value)
                if len(str_value) > 80:
                    str_value = str_value[:80] + "..."
                print(f"  [{headers[col_idx-1]}]: {str_value}")
        
        if not has_content:
            print("  (空白行)")
    
    # 統計 Ref 和 Validation 欄位
    if 'Ref' in key_columns:
        ref_col = key_columns['Ref']
        ref_count = 0
        ref_examples = []
        
        for row in range(2, min(ws.max_row + 1, 100)):
            ref_value = ws.cell(row, ref_col).value
            if ref_value and str(ref_value).strip():
                ref_count += 1
                if len(ref_examples) < 10:
                    ref_examples.append(str(ref_value))
        
        print(f"\n\nRef 欄位統計 (前 100 行):")
        print(f"  有值的行數: {ref_count}")
        if ref_examples:
            print(f"  範例 (前 10 個):")
            for i, ex in enumerate(ref_examples, 1):
                print(f"    {i}. {ex}")
    
    if 'Validation' in key_columns:
        val_col = key_columns['Validation']
        val_count = 0
        val_examples = []
        
        for row in range(2, min(ws.max_row + 1, 100)):
            val_value = ws.cell(row, val_col).value
            if val_value and str(val_value).strip():
                val_count += 1
                if len(val_examples) < 10:
                    val_examples.append(str(val_value))
        
        print(f"\n\nValidation 欄位統計 (前 100 行):")
        print(f"  有值的行數: {val_count}")
        if val_examples:
            print(f"  範例 (前 10 個):")
            for i, ex in enumerate(val_examples, 1):
                print(f"    {i}. {ex}")

print("\n" + "=" * 100)
print("分析完成")
print("=" * 100)
