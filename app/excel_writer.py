# excel_writer.py
# 用途：將log分析結果匯出為Excel檔案，支援PASS/FAIL分頁，欄位依規格
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles.borders import Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
import re
import os

class ExcelWriter:
    def __init__(self):
        pass

    def _insert_header_info(self, ws, header_info, start_row=4):
        """插入置頂 Header 資訊 (綠底黑字)"""
        if not header_info:
            return start_row
            
        try:
            lines = header_info.split('\n')
            current_row = start_row
            
            for line in lines:
                if not line.strip(): continue
                
                cell = ws.cell(row=current_row, column=1, value=self._sanitize_cell_text(line))
                cell.number_format = '@'
                cell.font = Font(name='Consolas', size=12, bold=True, color='FF000000') # 黑字
                cell.fill = PatternFill('solid', fgColor='FF90EE90') # 淺綠色背景
                cell.alignment = Alignment(horizontal='left', vertical='center')
                current_row += 1
                
            # 空一行
            return current_row + 1
        except Exception as e:
            print(f"插入 Header 失敗: {e}")
            return start_row

    def _sanitize_cell_text(self, value: object) -> str:
        """清理欲寫入儲存格的文字：
        - 轉為字串
        - 移除非法控制字元 (openpyxl 限制)
        - 去除 ANSI/ESC 序列 (更廣泛)
        - 截斷過長文字 (3 萬字元)
        """
        try:
            text = '' if value is None else str(value)
            # 移除 Excel 禁用控制字元 (0x00-0x1F 無法接受，openpyxl 內建 regex)
            text = ILLEGAL_CHARACTERS_RE.sub('', text)
            # 移除常見 ANSI CSI 序列: ESC [ ... letter
            text = re.sub(r'\x1b\[[0-9;?]*[A-Za-z]', '', text)
            # 移除其他 ESC 開頭的短序列: ESC ... 單個字母終止
            text = re.sub(r'\x1b[^A-Za-z]{0,20}[A-Za-z]', '', text)
            # 移除顏色碼樣式殘餘 (保險)
            text = re.sub(r'\x1b', '', text)
            # 若首字為會被當作公式的危險字元，前綴 '\''
            if text and text[0] in ('=', '+', '-', '@', '>', '<'):
                text = "'" + text
            # 長度截斷 (32767為Excel上限，保留安全餘量)
            if len(text) > 30000:
                text = text[:30000]
            return text
        except Exception:
            return '' if value is None else str(value)[:30000]

    def export(self, pass_items, fail_items, output_path):
        """
        匯出分析結果到Excel，分為PASS/FAIL兩個sheet
        pass_items: List[dict]
        fail_items: List[dict]
        """
        wb = Workbook()
        # 移除預設工作表
        wb.remove(wb.active)
        
        # PASS工作表
        ws_pass = wb.create_sheet('PASS')
        pass_headers = ['測項名稱', '指令', '收到指令', 'PASS/FAIL', '執行時間']
        ws_pass.append(pass_headers)
        
        for item in pass_items:
            row = [
                self._sanitize_cell_text(item.get('step_name', '')),
                self._sanitize_cell_text(item.get('command', '')),
                self._sanitize_cell_text(item.get('response', '')),
                'PASS',
                self._sanitize_cell_text(item.get('execution_time', ''))
            ]
            ws_pass.append(row)
        
        # FAIL工作表
        ws_fail = wb.create_sheet('FAIL')
        fail_headers = ['測項名稱', '指令', '錯誤回應', 'Retry次數', 'FAIL原因', '錯誤碼']
        ws_fail.append(fail_headers)
        
        for item in fail_items:
            row = [
                self._sanitize_cell_text(item.get('step_name', '')),
                self._sanitize_cell_text(item.get('command', '')),
                self._sanitize_cell_text(item.get('response', '')),
                item.get('retry', 0),
                self._sanitize_cell_text(item.get('error', '')),
                self._sanitize_cell_text(item.get('error_code', ''))
            ]
            ws_fail.append(row)
        
        # 設定欄寬
        for ws in [ws_pass, ws_fail]:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
        
        # 設定標題樣式
        for ws in [ws_pass, ws_fail]:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 儲存檔案
        try:
            wb.save(output_path)
        finally:
            try:
                wb.close()
            except Exception:
                pass
        return output_path

    def _extract_system_info(self, raw_lines: list) -> dict:
        """從 raw_lines 擷取系統資訊"""
        try:
            system_info = {}
            for line in raw_lines[:100]:  # 只檢查前100行，系統資訊通常在LOG開頭
                line_str = str(line).strip()
                
                # 擷取各種系統資訊
                if 'System Version is' in line_str:
                    match = re.search(r'System Version is ([^.]+)', line_str)
                    if match:
                        system_info['System Version'] = match.group(1).strip()
                
                elif 'Script File is' in line_str:
                    match = re.search(r'Script File is ([^.]+)', line_str)
                    if match:
                        system_info['Script File'] = match.group(1).strip()
                
                elif 'Script Version is' in line_str:
                    match = re.search(r'Script Version is ([^.]+)', line_str)
                    if match:
                        system_info['Script Version'] = match.group(1).strip()
                
                elif 'Script Creator is' in line_str:
                    match = re.search(r'Script Creator is ([^.]+)', line_str)
                    if match:
                        system_info['Script Creator'] = match.group(1).strip()
                
                elif 'Utility Version is' in line_str:
                    match = re.search(r'Utility Version is ([^.]+)', line_str)
                    if match:
                        system_info['Utility Version'] = match.group(1).strip()
                
                elif 'DeviceID is' in line_str:
                    match = re.search(r'DeviceID is ([^.]+)', line_str)
                    if match:
                        system_info['DeviceID'] = match.group(1).strip()
                
                elif 'SFIS is' in line_str:
                    match = re.search(r'SFIS is ([^.]+)', line_str)
                    if match:
                        system_info['SFIS'] = match.group(1).strip()
                
                elif 'MiniCloud2 is' in line_str:
                    match = re.search(r'MiniCloud2 is ([^.]+)', line_str)
                    if match:
                        system_info['MiniCloud2'] = match.group(1).strip()
                
                elif 'FTP is' in line_str:
                    match = re.search(r'FTP is ([^.]+)', line_str)
                    if match:
                        system_info['FTP'] = match.group(1).strip()
                
                elif 'Active IPs:' in line_str:
                    # 擷取下一行的IP資訊
                    idx = raw_lines.index(line)
                    if idx + 1 < len(raw_lines):
                        next_line = str(raw_lines[idx + 1]).strip()
                        if next_line and not next_line.endswith('.'):
                            system_info['Active IPs'] = next_line
            
            return system_info
        except Exception:
            return {}

    def _extract_total_secs(self, raw_lines: list) -> float | None:
        """從 raw_lines 嘗試提取測試總時間（秒數）"""
        try:
            for line in (raw_lines[-50:] if len(raw_lines) > 50 else raw_lines):
                line_str = str(line).strip()
                line_lower = line_str.lower()
                
                # 優先尋找 "All phase Total Test Time" 格式
                if 'all phase total test time' in line_lower:
                    # 尋找 "----- XXX.XXXXXXX Sec." 格式
                    time_match = re.search(r'----- (\d+\.?\d*) Sec\.', line_str)
                    if time_match:
                        val = float(time_match.group(1))
                        if val > 0:
                            return val
                
                # 備用：尋找其他時間格式
                elif 'testtime' in line_lower or 'total time' in line_lower:
                    # 嘗試提取數字
                    nums = re.findall(r'(\d+\.?\d*)', line_str)
                    if nums:
                        val = float(nums[-1])
                        if val > 0:
                            return val
            return None
        except Exception:
            return None

    def _unique_sheet_name(self, wb, base_name: str) -> str:
        """確保工作表名稱不重複"""
        # 先清理名稱中的非法字元與空白
        base_name = self._sanitize_sheet_title(base_name)
        if len(base_name) > 31:
            base_name = base_name[:28] + '...'
        existing = [ws.title for ws in wb.worksheets]
        if base_name not in existing:
            return base_name
        counter = 1
        while f"{base_name}({counter})" in existing:
            counter += 1
        suffix = f"({counter})"
        max_base = 31 - len(suffix)
        return f"{base_name[:max_base]}{suffix}"

    def _sanitize_sheet_title(self, title: str) -> str:
        """移除Excel工作表名稱不允許的字元並修剪長度。
        禁用字元: : \\ / ? * [ ]，且長度<=31，不可為空。
        """
        try:
            s = str(title) if title is not None else 'Sheet'
            # 移除路徑與副檔名殘留
            s = s.replace('\\', ' ').replace('/', ' ')
            # 移除禁止字元
            import re
            s = re.sub(r'[:\\/\?\*\[\]]', ' ', s)
            # 去除前後單引號
            s = s.strip().strip("'")
            # 轉為可見字串
            s = s if s else 'Sheet'
            # Excel 限制 31 字
            return s[:31]
        except Exception:
            return 'Sheet'

    def _extract_isn_from_filename(self, filename: str) -> str:
        """從檔名嘗試提取 ISN (WE開頭 或 純數字10碼以上)"""
        try:
            if not filename: return ""
            # 常見格式 1+4cam...-WE2536...-... 或 ...-1110250497-...
            parts = filename.split('-')
            for part in parts:
                # 匹配 WE 開頭
                if part.startswith('WE') and len(part) > 8:
                    return part
                # 匹配純數字 (長度 9-15)
                if part.isdigit() and 9 <= len(part) <= 15:
                    return part
            return ""
        except:
            return ""

    def export_pass_fail_workbooks(self, folder_path: str, pass_logs: list, fail_logs: list):
        """
        輸出兩個活頁簿：
        - PASS匯總.xlsx：Summary + 每個 PASS LOG 的工作表
        - FAIL匯總.xlsx：Summary + 每個 FAIL LOG 的工作表
        pass_logs/fail_logs 需提供：
          [{
             'file_path': 絕對或相對路徑,
             'file_name': 檔名,
             'raw_lines': [原始行...],
             'ui_annotations': [{line_idx, line_content, color, ...}],
             'pass_items': [...],
             'fail_items': [...],
             'summary': { '測試日期時間': str, 'SFIS': 'ON'|'OFF', '測試總時間': str, 'FAIL原因': str可選 }
          }, ...]
        """
        pass_path = os.path.join(folder_path, 'PASS匯總.xlsx')
        fail_path = os.path.join(folder_path, 'FAIL匯總.xlsx')
        self._build_pass_workbook(pass_path, pass_logs)
        self._build_fail_workbook(fail_path, fail_logs)
        return pass_path, fail_path

    def _extract_station_from_filename_content(self, filename: str) -> str:
        """從檔名提取 Station 名稱
        規則範例: 1+4cam stitching1 test-1110... -> 4cam stitching1
        去除前面的數字和+號，去除 test- 及其後面所有內容
        """
        try:
            if not filename: return "Unknown Station"
            
            # 先去除副檔名
            base = os.path.splitext(filename)[0]
            
            # 分割字串，尋找 "test" 或 "-"
            # 策略：找到第一個 "test" (不分大小寫) 並截斷
            match = re.search(r'(?i)test', base)
            if match:
                station_part = base[:match.start()]
            else:
                # 如果沒有 test，嘗試用第一個 "-" 分割，如果 "-" 後面是數字（雖然這可能切錯）
                # 簡單起見，如果沒有 test，用第一個 "-"
                parts = base.split('-')
                station_part = parts[0]
            
            # 清理：去除前面的 "1+", "2+" 等數字加號組合，以及前後空白
            # 例如 "1+4cam" -> "4cam", "12+Station" -> "Station"
            station_part = re.sub(r'^\d+\+', '', station_part)
            
            clean_station = station_part.strip()
            return clean_station if clean_station else "Unknown Station"
            
        except Exception:
            return "Unknown Station"


    def _build_fail_list_sheet(self, wb, logs):
        """建立 FAIL_LIST 工作表 (依使用者要求的 CSV 格式，並統計相同錯誤)"""
        ws = wb.create_sheet("FAIL_LIST", 0) # 放在第一頁
        try:
            ws.sheet_properties.tabColor = 'FFFF0000' # 紅色標籤
        except Exception:
            pass
        
        # 設定標題 (新增 Count 欄位)
        headers = ['ISN', 'Station', 'FAIL Item', 'FAIL Reason', 'suggestion', 'Count']
        ws.append(headers)
        
        # 樣式設定
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        content_font = Font(name='Calibri', size=11)
        fill_blue = PatternFill('solid', fgColor='FF4472C4')
        center_align = Alignment(horizontal='center', vertical='center')
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = fill_blue
            cell.alignment = center_align
        
        # 1. 收集所有資料行與統計錯誤
        pending_rows = []
        error_counts = {}
        
        for entry in logs:
            fail_items = entry.get('fail_items', [])
            if not fail_items:
                continue
                
            fname = entry.get('file_name', '')
            isn = self._extract_isn_from_filename(fname)
            station = self._extract_station_from_filename_content(fname)
            suggestion = "請參閱 PEGA SOP"
            
            for item in fail_items:
                step_name = item.get('step_name', '')
                result = item.get('result', '')
                fail_item_str = f"{step_name}"
                if result and result not in step_name:
                     fail_item_str += f" {result}"
                
                reason = item.get('error', '')
                if not reason or reason == 'FAIL':
                    reason = item.get('response', '')
                
                # 標準化錯誤原因以進行統計 (去除時間戳、錯誤碼等變動資訊)
                # 使用 _extract_main_error_type (它已經包含在 _normalize_error_group 中，但直接調用更單純)
                # 這裡使用 _normalize_error_group 可以得到更乾淨的分類
                norm_key = self._normalize_error_group(reason)
                error_counts[norm_key] = error_counts.get(norm_key, 0) + 1
                
                pending_rows.append({
                    'isn': isn,
                    'station': station,
                    'item': fail_item_str,
                    'reason': reason,
                    'suggestion': suggestion,
                    'norm_key': norm_key
                })
        
        # 2. 寫入資料
        for row_data in pending_rows:
            count = error_counts.get(row_data['norm_key'], 0)
            
            row_values = [
                self._sanitize_cell_text(row_data['isn']),
                self._sanitize_cell_text(row_data['station']),
                self._sanitize_cell_text(row_data['item']),
                self._sanitize_cell_text(row_data['reason']),
                self._sanitize_cell_text(row_data['suggestion']),
                count
            ]
            ws.append(row_values)
            
            # Apply content font
            current_row = ws.max_row
            for col_idx, value in enumerate(row_values, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = content_font
                if col_idx == 6: # Count column
                    cell.alignment = center_align
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                
        # 強制自動調整欄寬
        min_widths = {
            1: 15, # ISN
            2: 20, # Station
            3: 30, # FAIL Item
            4: 50, # FAIL Reason
            5: 20, # suggestion
            6: 10  # Count
        }
        
        for col_idx, min_w in min_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min_w

        # 根據內文調整 (如果有更長的)
        for col in ws.columns:
            max_len = 0
            col_idx = col[0].column
            col_letter = get_column_letter(col_idx)
            
            # 跳過 Reason 欄位因為它可能很長且已設定 wrap_text，太寬不好看
            if col_idx == 4: 
                continue
                
            for cell in col:
                try:
                    val = str(cell.value) if cell.value else ""
                    # 考慮中文字寬度
                    length = len(val.encode('utf-8')) * 0.5 
                    if length > max_len:
                        max_len = length
                except:
                    pass
            
            current_w = ws.column_dimensions[col_letter].width
            new_w = min(max(current_w, max_len + 2), 80)
            ws.column_dimensions[col_letter].width = new_w

    def _format_filename_with_timestamp(self, base_name: str) -> str:
        """將檔名中的連續14位時間戳 YYYYMMDDHHMMSS 轉為 YYYY-MMDD-HHMMSS 格式。找不到則原樣回傳。"""
        try:
            m = re.search(r'(20\d{12})', base_name)
            if not m:
                return base_name
            s = m.group(1)
            y, md, hms = s[:4], s[4:8], s[8:]
            return base_name.replace(s, f"{y}-{md}-{hms}")
        except Exception:
            return base_name

    def _build_preview_comment(self, entry: dict) -> str:
        """產生懸停預覽內容：顯示對應工作表名稱與原始LOG前幾行（加上簡易標記）。"""
        try:
            sheet_or_name = entry.get('file_name') or 'LOG'
            raw = entry.get('raw_lines') or []
            preview_lines = []
            header = f"對應工作表: {sheet_or_name}\n******** 預覽 ********"
            preview_lines.append(header)
            for i, line in enumerate(raw[:15], 1):
                s = str(line)
                # 簡易高亮標記
                if 'Do @STEP' in s:
                    s = f"[STEP] {s}"
                if 'FAIL' in s.upper() or 'ERROR' in s.upper():
                    s = f"[FAIL] {s}"
                s = re.sub(r'^\s*>\s*', '▶ ', s)
                s = re.sub(r'^\s*<\s*', '◀ ', s)
                preview_lines.append(s)
            preview_lines.append('************************')
            return '\n'.join(preview_lines)
        except Exception:
            return '對應工作表預覽不可用'

    def _add_input_prompt(self, ws, cell, title: str, message: str):
        """在儲存格上加一個資料驗證提示（白底有框），當選取時顯示。"""
        try:
            # Excel 對於訊息有長度限制，做適當截斷
            msg = (message or '')
            if len(msg) > 250:
                msg = msg[:250] + '…'
            dv = DataValidation(type="custom", formula1="TRUE", allow_blank=True, showInputMessage=True)
            dv.promptTitle = title[:30] if title else ''
            dv.prompt = msg
            ws.add_data_validation(dv)
            dv.add(cell.coordinate)
        except Exception:
            pass

    # 內部：PASS 匯總活頁簿
    def _safe_save_workbook(self, wb, output_path):
        """安全保存工作簿，避免Excel警告"""
        try:
            # 設定工作簿屬性以避免Excel警告
            wb.properties.creator = "PEGA Log Analyzer"
            wb.properties.title = "Log Analysis Results"
            wb.properties.description = "Log analysis results generated by PEGA Log Analyzer"
            wb.properties.subject = "Log Analysis"
            wb.properties.keywords = "LOG,Analysis,PEGA"
            
            # 設定安全屬性
            wb.security.lockStructure = False
            wb.security.lockWindows = False
            
            # 清理可能導致問題的內容
            for ws in wb.worksheets:
                # 確保所有單元格都有正確的格式
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            # 確保文字格式正確
                            if isinstance(cell.value, str):
                                # 移除可能導致問題的控制字符
                                cell.value = ''.join(char for char in cell.value if ord(char) >= 32 or char in '\t\n\r')
                                cell.number_format = '@'  # 文字格式
            
            self._safe_save_workbook(wb, output_path)
            
        except Exception as e:
            print(f"保存Excel檔案時發生錯誤: {e}")
            # 嘗試基本保存
            try:
                self._safe_save_workbook(wb, output_path)
            except Exception as e2:
                print(f"基本保存也失敗: {e2}")
                raise e2

    def _build_pass_workbook(self, output_path: str, logs: list):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Summary'
        # 設定標籤顏色（紅色）
        try:
            ws.sheet_properties.tabColor = 'FFFF0000'
        except Exception:
            pass
        header_font = Font(name='Calibri', size=16, bold=True, color='FFFFFFFF')
        normal_font = Font(name='Calibri', size=10)
        center = Alignment(horizontal='center', vertical='center')
        deep_green = PatternFill('solid', fgColor='FF1B5E20')
        # 先建立各 LOG 原始工作表，並記錄 sheet 名稱
        sheet_map = {}
        for entry in logs:
            fname = entry.get('file_name', 'LOG')
            isn = self._extract_isn_from_filename(fname)
            # 優先使用 ISN 作為 Sheet 名稱，若無則使用檔名
            sheet_name_base = self._sanitize_sheet_title(isn if isn else fname)
            sheet_name = self._unique_sheet_name(wb, sheet_name_base)
            sheet_map[entry.get('file_name')] = sheet_name
            ws2 = wb.create_sheet(title=sheet_name)
            
            # 在最上面添加回到 Summary 的連結
            summary_link_cell = ws2.cell(row=1, column=1, value='🔙 回到 Summary 頁面')
            summary_link_cell.number_format = '@'
            summary_link_cell.font = Font(name='Calibri', size=11, bold=True, color='FF008000', underline='single')
            summary_link_cell.alignment = Alignment(horizontal='left', vertical='center')
            summary_link_cell.hyperlink = f"#'Summary'!A1"
            summary_link_cell.fill = PatternFill('solid', fgColor='FFE6FFE6')  # 淺綠色背景
            ws2.row_dimensions[1].height = 20
            
            # 添加分隔線
            ws2.cell(row=2, column=1, value='─' * 50).font = Font(name='Calibri', size=10, color='FF808080')
            ws2.row_dimensions[2].height = 15
            
            # 檔名標題
            cell = ws2.cell(row=3, column=1, value=self._sanitize_cell_text(entry.get('file_name')))
            cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='FF2ECC71')
            cell.number_format = '@'
            
            # 插入 Header Info (綠底黑字)
            current_row = 4
            header_info = entry.get('header_info', '')
            if header_info:
                current_row = self._insert_header_info(ws2, header_info, start_row=current_row)
            
            # 在最上面添加 PASS 步驟詳情
            pass_steps = entry.get('pass_items', [])
            if pass_steps:
                pass_steps_text = self._build_pass_steps_summary(pass_steps)
                pass_lines = pass_steps_text.split('\n')
                # current_row 已經由 _insert_header_info 更新
                
                for line in pass_lines:
                    if line.strip():
                        cell = ws2.cell(row=current_row, column=1, value=self._sanitize_cell_text(line))
                        cell.number_format = '@'
                        
                        # 設定不同行的樣式
                        if "===============PASS步驟詳情====================" in line:
                            cell.font = Font(name='Calibri', size=12, bold=True, color='FF008000')
                        elif "✅ 步驟" in line:
                            cell.font = Font(name='Calibri', size=11, bold=True, color='FF008000')
                        elif "   執行指令:" in line:
                            cell.font = Font(name='Calibri', size=11, color='FF0000FF')
                        elif "   回應:" in line:
                            cell.font = Font(name='Calibri', size=11, color='FF000000')
                        elif "   執行時間:" in line:
                            cell.font = Font(name='Calibri', size=11, color='FF666666')
                        elif "=" * 50 in line:
                            cell.font = Font(name='Calibri', size=10)
                        else:
                            cell.font = Font(name='Calibri', size=11)
                        
                        # 設定行高以顯示更多文字
                        ws2.row_dimensions[current_row].height = 25
                        current_row += 1
                    else:
                        current_row += 1
                
                # 添加分隔線
                ws2.cell(row=current_row, column=1, value=self._sanitize_cell_text("=" * 60)).number_format='@'
                ws2.cell(row=current_row, column=1).font = Font(name='Calibri', size=10)
                current_row += 1
                
                # 寫入原始LOG內容
                self._write_raw_log_with_annotations(ws2, start_row=current_row, raw_lines=entry.get('raw_lines') or [], annotations=entry.get('ui_annotations') or [], font=Font(name='Calibri', size=11), step_marks=entry.get('step_marks'))
            else:
                # 沒有 PASS 步驟時，直接寫入原始LOG
                self._write_raw_log_with_annotations(ws2, start_row=current_row, raw_lines=entry.get('raw_lines') or [], annotations=entry.get('ui_annotations') or [], font=Font(name='Calibri', size=11), step_marks=entry.get('step_marks'))
            
            # 在最下面添加回到 Summary 的連結（只添加一個）
            bottom_link_row = ws2.max_row + 2
            bottom_summary_link_cell = ws2.cell(row=bottom_link_row, column=1, value='🔙 回到 Summary 頁面')
            bottom_summary_link_cell.number_format = '@'
            bottom_summary_link_cell.font = Font(name='Calibri', size=11, bold=True, color='FF008000', underline='single')
            bottom_summary_link_cell.alignment = Alignment(horizontal='left', vertical='center')
            bottom_summary_link_cell.hyperlink = f"#'Summary'!A1"
            bottom_summary_link_cell.fill = PatternFill('solid', fgColor='FFE6FFE6')  # 淺綠色背景
            ws2.row_dimensions[bottom_link_row].height = 20
            
            # 設定所有行的行高以顯示更多文字
            for row_num in range(1, ws2.max_row + 1):
                if ws2.row_dimensions[row_num].height == 15:  # 預設行高
                    ws2.row_dimensions[row_num].height = 20
            # 調整欄寬以顯示所有文字
            self._auto_fit_columns(ws2, min_widths={1: 150})
        # Summary 表格：
        headers = ['檔名', 'PASS步驟數', '測試總時間', 'SFIS']
        for c, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=header)
            cell.font = header_font
            cell.alignment = center
            cell.fill = deep_green
        ws.freeze_panes = 'A2'
        thin = Side(border_style='thin', color='FF888888')
        thick = Side(border_style='thick', color='FF000000')
        start_data_row = ws.max_row + 1
        for entry in logs:
            # 檔名欄
            base = self._sanitize_cell_text(entry.get('file_name') or '')
            base_fmt = self._format_filename_with_timestamp(base)
            
            # 擷取系統資訊
            system_info = self._extract_system_info(entry.get('raw_lines') or [])
            
            # 組合顯示名稱，使用多行格式讓使用者容易看懂
            display_lines = [base_fmt]
            
            # 添加系統資訊，分行顯示
            if system_info.get('System Version'):
                display_lines.append(f"系統版本: {system_info['System Version']}")
            if system_info.get('Script File'):
                display_lines.append(f"腳本檔案: {system_info['Script File']}")
            if system_info.get('Script Version'):
                display_lines.append(f"腳本版本: {system_info['Script Version']}")
            if system_info.get('Script Creator'):
                display_lines.append(f"建立者: {system_info['Script Creator']}")
            if system_info.get('Utility Version'):
                display_lines.append(f"工具版本: {system_info['Utility Version']}")
            if system_info.get('DeviceID'):
                display_lines.append(f"設備ID: {system_info['DeviceID']}")
            if system_info.get('SFIS'):
                display_lines.append(f"SFIS: {system_info['SFIS']}")
            if system_info.get('MiniCloud2'):
                display_lines.append(f"MiniCloud2: {system_info['MiniCloud2']}")
            if system_info.get('FTP'):
                display_lines.append(f"FTP: {system_info['FTP']}")
            if system_info.get('Active IPs'):
                display_lines.append(f"IP位址: {system_info['Active IPs']}")
            
            # 添加測試時間
            secs = self._extract_total_secs(entry.get('raw_lines') or [])
            if secs is not None:
                display_lines.append(f"測試時間: {secs:.1f} 秒")
            
            display_name = "\n".join(display_lines)
            
            r = ws.max_row + 1
            cell_name = ws.cell(row=r, column=1, value=self._sanitize_cell_text(display_name))
            cell_name.number_format='@'
            cell_name.font = Font(name='Calibri', size=10, color='FF000000')
            cell_name.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top', shrink_to_fit=True)
            # 備註與超連結
            sheet = sheet_map.get(entry.get('file_name'))
            try:
                cell_name.comment = Comment(self._build_preview_comment(entry), "LOG Analyzer")
                cell_name.comment.width = 400
                cell_name.comment.height = 500
            except Exception:
                pass
            # 移除超連結，只保留提示
            if sheet:
                self._add_input_prompt(ws, cell_name, '對應工作表', entry.get('file_name') or '')
            # PASS步驟數欄
            pass_count = len(entry.get('pass_items') or [])
            cell_count = ws.cell(row=r, column=2, value=pass_count)
            cell_count.font = normal_font
            cell_count.alignment = center
            
            # 測試總時間欄
            sec_txt = f"{secs:.1f} Sec." if secs is not None else ''
            cell_time = ws.cell(row=r, column=3, value=sec_txt)
            cell_time.font = normal_font
            cell_time.alignment = center
            # SFIS欄
            sfis_value = system_info.get('SFIS', '')
            cell_sfis = ws.cell(row=r, column=4, value=sfis_value)
            cell_sfis.font = normal_font
            cell_sfis.alignment = center
            for c in range(1, len(headers)+1):
                ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # 外框粗線
        end_r = ws.max_row
        if end_r >= start_data_row:
            for c in range(1, len(headers)+1):
                ws.cell(row=start_data_row, column=c).border = ws.cell(row=start_data_row, column=c).border.copy(top=thick)
                ws.cell(row=end_r, column=c).border = ws.cell(row=end_r, column=c).border.copy(bottom=thick)
            for r in range(start_data_row, end_r+1):
                ws.cell(row=r, column=1).border = ws.cell(row=r, column=1).border.copy(left=thick)
                ws.cell(row=r, column=len(headers)).border = ws.cell(row=r, column=len(headers)).border.copy(right=thick)
        # 表格底部：SHEET 快速連結（點擊跳轉）
        link_title_row = ws.max_row + 2
        ws.cell(row=link_title_row, column=1, value='工作表快速連結（點擊跳轉）').font = Font(name='Microsoft JhengHei', size=11, bold=True)
        ws.cell(row=link_title_row, column=1).alignment = Alignment(horizontal='left')
        cur = link_title_row + 1
        for entry in logs:
            base = self._sanitize_cell_text(entry.get('file_name') or '')
            base_fmt = self._format_filename_with_timestamp(base)
            sfis = (entry.get('summary') or {}).get('SFIS', '')
            sfis = (sfis or '').upper()
            secs = self._extract_total_secs(entry.get('raw_lines') or [])
            sec_txt = f"測試總時間:{secs:.1f} Sec." if secs is not None else ''
            suffix = f"_SFIS_{sfis}" if sfis else ''
            display_name = f"{base_fmt}{suffix} {sec_txt}".strip()
            c = ws.cell(row=cur, column=1, value=self._sanitize_cell_text(display_name))
            c.number_format='@'
            c.font = Font(name='Microsoft JhengHei', size=11, color='FF0000FF', underline='single')
            c.alignment = Alignment(horizontal='left')
            sheet = sheet_map.get(entry.get('file_name'))
            if sheet:
                c.hyperlink = f"#'{sheet}'!A1"
                try:
                    c.comment = Comment(self._build_preview_comment(entry), 'LOG Analyzer')
                    c.comment.width = 400
                    c.comment.height = 500
                except Exception:
                    pass
                # 白底提示（提示箭頭在左上方，靠近視窗；Excel控制箭頭顯示位置有限）
                self._add_input_prompt(ws, c, '對應工作表', sheet)
            cur += 1
        # 更緊湊的欄寬
        self._auto_fit_columns(ws, min_widths={1: 30, 2: 12, 3: 15, 4: 8})
        
        # 設定 Summary 頁面所有文字字體為 11 (Calibri)
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # 保留原有的字體屬性，只修改大小和名稱
                    current_font = cell.font
                    cell.font = Font(
                        name='Calibri',
                        size=11,
                        bold=current_font.bold,
                        italic=current_font.italic,
                        color=current_font.color,
                        underline=current_font.underline
                    )
        
        try:
            wb.save(output_path)
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def _build_fail_workbook(self, output_path: str, logs: list):
        wb = Workbook()
        # 移除預設工作表
        if wb.sheetnames:
            wb.remove(wb.active)
            
        # 加入 FAIL_LIST sheet
        self._build_fail_list_sheet(wb, logs)
        
        ws = wb.active
        if ws.title != 'FAIL_LIST':
             ws = wb.create_sheet('Summary')
        else:
             ws = wb.create_sheet('Summary')

        ws.title = 'Summary'
        # 設定標籤顏色（紅色）
        try:
            ws.sheet_properties.tabColor = 'FFFF0000'
        except Exception:
            pass
        header_font = Font(name='Calibri', size=16, bold=True, color='FFFFFFFF')
        normal_font = Font(name='Calibri', size=10)
        center = Alignment(horizontal='center', vertical='center')
        deep_green = PatternFill('solid', fgColor='FF1B5E20')
        headers = ['檔名', '詳細錯誤原因']
        ws.append(headers)
        for c in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.alignment = center
            cell.fill = deep_green
        ws.freeze_panes = 'A2'
        # 先建立各 LOG 原始工作表
        sheet_map = {}
        for entry in logs:
            fname = entry.get('file_name', 'LOG')
            isn = self._extract_isn_from_filename(fname)
            # 優先使用 ISN 作為 Sheet 名稱，若無則使用檔名
            sheet_name_base = self._sanitize_sheet_title(isn if isn else fname)
            sheet_name = self._unique_sheet_name(wb, sheet_name_base)
            sheet_map[entry.get('file_name')] = sheet_name
            ws2 = wb.create_sheet(title=sheet_name)
            cell = ws2.cell(row=1, column=1, value=self._sanitize_cell_text(entry.get('file_name')))
            cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='FFE74C3C')
            cell.number_format = '@'
            # 在最上面添加回到Summary的快速連結
            back_to_summary_top = ws2.cell(row=2, column=1, value='🔙 回到 Summary 頁面')
            back_to_summary_top.number_format='@'
            back_to_summary_top.font = Font(name='Calibri', size=12, bold=True, color='FF008000', underline='single')
            back_to_summary_top.alignment = Alignment(horizontal='left')
            back_to_summary_top.hyperlink = f"#'Summary'!A1"
            back_to_summary_top.fill = PatternFill('solid', fgColor='FFE6FFE6')  # 淺綠色背景
            
            # 添加分隔線
            ws2.cell(row=3, column=1, value='─' * 50).font = Font(name='Calibri', size=10, color='FF808080')
            
            # 插入 Header Info (綠底黑字)
            current_row = 4
            header_info = entry.get('header_info', '')
            if header_info:
                current_row = self._insert_header_info(ws2, header_info, start_row=current_row)
            
            # 顯示完整錯誤原因區塊
            detailed_error = self._build_detailed_error_summary(entry)
            error_lines = detailed_error.split('\n')
            # current_row 已經由 _insert_header_info 更新
            
            for line in error_lines:
                if line.strip():
                    cell = ws2.cell(row=current_row, column=1, value=self._sanitize_cell_text(line))
                    cell.number_format = '@'
                    
                    # 設定不同行的樣式
                    if "===============錯誤原因====================" in line:
                        cell.font = Font(name='Calibri', size=12, bold=True, color='FF0000')
                    elif "🔴 突出錯誤" in line:
                        cell.font = Font(name='Calibri', size=11, bold=True, color='FF0000')
                    elif "=" * 50 in line:
                        cell.font = Font(name='Calibri', size=10)
                    elif "執行指令:" in line:
                        cell.font = Font(name='Calibri', size=11, color='FF0000FF')
                    elif any(keyword in line.lower() for keyword in ['is fail', 'executes fail', "doesn't match", 'all test aborted']):
                        cell.font = Font(name='Calibri', size=11, bold=True, color='FF0000')
                        cell.fill = PatternFill('solid', fgColor='FFFFFF99')  # 淺黃色背景
                    else:
                        cell.font = Font(name='Calibri', size=11)
                    
                    # 設定行高以顯示更多文字
                    ws2.row_dimensions[current_row].height = 25
                    current_row += 1
                else:
                    current_row += 1
            
            # 添加分隔線
            ws2.cell(row=current_row, column=1, value=self._sanitize_cell_text("=" * 60)).number_format='@'
            ws2.cell(row=current_row, column=1).font = Font(name='Microsoft JhengHei', size=10)
            current_row += 1
            
            # PASS步驟資訊
            pass_steps = [str(i) for i, _ in enumerate(entry.get('pass_items') or [], 1)]
            ws2.cell(row=current_row, column=1, value=self._sanitize_cell_text('PASS步驟: ' + '，'.join(pass_steps) if pass_steps else 'PASS步驟: 無'))
            ws2.cell(row=current_row, column=1).number_format='@'
            ws2.cell(row=current_row, column=1).font = Font(name='Calibri', size=11, color='FF008000')
            start_row = current_row + 1
            
            # 寫入原始LOG內容，並標記錯誤行
            self._write_raw_log_with_annotations(ws2, start_row=start_row, raw_lines=entry.get('raw_lines') or [], annotations=entry.get('ui_annotations') or [], font=Font(name='Calibri', size=11), step_marks=entry.get('step_marks'))
            
            # 設定所有行的行高以顯示更多文字
            for row_num in range(1, ws2.max_row + 1):
                if ws2.row_dimensions[row_num].height == 15:  # 預設行高
                    ws2.row_dimensions[row_num].height = 20
            self._auto_fit_columns(ws2)
        # 在Summary頁面最上面添加錯誤統計
        try:
            self._add_error_statistics(ws, logs)
        except Exception as e:
            print(f"添加錯誤統計時發生錯誤: {e}")
        
        # Summary：每個LOG一行，顯示檔名和詳細錯誤原因
        thin = Side(border_style='thin', color='FF888888')
        thick = Side(border_style='thick', color='FF000000')
        start_data_row = ws.max_row + 1
        for entry in logs:
            # 檔名欄
            base = self._sanitize_cell_text(entry.get('file_name') or '')
            base_fmt = self._format_filename_with_timestamp(base)
            sfis = (entry.get('summary') or {}).get('SFIS','')
            sfis = (sfis or '').upper()
            secs = self._extract_total_secs(entry.get('raw_lines') or [])
            sec_txt = f"測試總時間:{secs:.1f} Sec." if secs is not None else ''
            suffix = f"_SFIS_{sfis}" if sfis else ''
            display_name = f"{base_fmt}{suffix} {sec_txt}".strip()
            r = ws.max_row + 1
            cell_name = ws.cell(row=r, column=1, value=self._sanitize_cell_text(display_name))
            cell_name.number_format='@'
            cell_name.font = Font(name='Calibri', size=10, color='FF000000')
            cell_name.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top', shrink_to_fit=True)
            # 備註與超連結
            sheet = sheet_map.get(entry.get('file_name'))
            try:
                cell_name.comment = Comment(self._build_preview_comment(entry), "LOG Analyzer")
                cell_name.comment.width = 400
                cell_name.comment.height = 500
            except Exception:
                pass
            if sheet:
                cell_name.hyperlink = f"#'{sheet}'!A1"
            # 白底提示
            self._add_input_prompt(ws, cell_name, '對應工作表', entry.get('file_name') or '')
            
            # 詳細錯誤原因欄 - 包含主要錯誤和執行指令
            detailed_error = self._build_detailed_error_summary(entry)
            cell_reason = ws.cell(row=r, column=2, value=self._sanitize_cell_text(detailed_error))
            cell_reason.number_format='@'
            cell_reason.font = Font(name='Calibri', size=11)
            cell_reason.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left', shrink_to_fit=False)
            # 設定行高以顯示更多文字
            ws.row_dimensions[r].height = 120
            
            # 如果錯誤原因包含 "doesn't match"，用特殊格式突出顯示
            if "doesn't match" in detailed_error.lower():
                cell_reason.font = Font(name='Calibri', size=11, bold=True, color='FFFF0000')
                cell_reason.fill = PatternFill('solid', fgColor='FFFFFF99')  # 淺黃色背景
            
            for c in range(1, 2+1):
                ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
            end_r = ws.max_row
            if end_r >= start_data_row:
                for c in range(1, 2+1):
                    ws.cell(row=start_data_row, column=c).border = ws.cell(row=start_data_row, column=c).border.copy(top=thick)
                    ws.cell(row=end_r, column=c).border = ws.cell(row=end_r, column=c).border.copy(bottom=thick)
                for r in range(start_data_row, end_r+1):
                    ws.cell(row=r, column=1).border = ws.cell(row=r, column=1).border.copy(left=thick)
                    ws.cell(row=r, column=2).border = ws.cell(row=r, column=2).border.copy(right=thick)
            # 表格底部：SHEET 快速連結
            link_title_row = ws.max_row + 2
            ws.cell(row=link_title_row, column=1, value='工作表快速連結（點擊跳轉）').font = Font(name='Calibri', size=10, bold=True)
            ws.cell(row=link_title_row, column=1).alignment = Alignment(horizontal='left')
            cur = link_title_row + 1
        
            # 添加回到Summary的快速連結
            back_to_summary = ws.cell(row=cur, column=1, value='🔙 回到 Summary 頁面')
            back_to_summary.number_format='@'
            back_to_summary.font = Font(name='Calibri', size=12, bold=True, color='FF008000', underline='single')
            back_to_summary.alignment = Alignment(horizontal='left')
            back_to_summary.hyperlink = f"#'Summary'!A1"
            back_to_summary.fill = PatternFill('solid', fgColor='FFE6FFE6')  # 淺綠色背景
            cur += 1
        
            # 添加分隔線
            ws.cell(row=cur, column=1, value='─' * 50).font = Font(name='Microsoft JhengHei', size=10, color='FF808080')
            cur += 1
        
        for entry in logs:
            base = self._sanitize_cell_text(entry.get('file_name') or '')
            base_fmt = self._format_filename_with_timestamp(base)
            sfis = (entry.get('summary') or {}).get('SFIS','')
            sfis = (sfis or '').upper()
            secs = self._extract_total_secs(entry.get('raw_lines') or [])
            sec_txt = f"測試總時間:{secs:.1f} Sec." if secs is not None else ''
            suffix = f"_SFIS_{sfis}" if sfis else ''
            display_name = f"{base_fmt}{suffix} {sec_txt}".strip()
            c = ws.cell(row=cur, column=1, value=self._sanitize_cell_text(display_name))
            c.number_format='@'
            c.font = Font(name='Microsoft JhengHei', size=10, color='FF0000FF', underline='single')
            c.alignment = Alignment(horizontal='left')
            sheet = sheet_map.get(entry.get('file_name'))
            if sheet:
                c.hyperlink = f"#'{sheet}'!A1"
                try:
                    c.comment = Comment(self._build_preview_comment(entry), 'LOG Analyzer')
                    c.comment.width = 400
                    c.comment.height = 500
                except Exception:
                    pass
            self._add_input_prompt(ws, c, '對應工作表', sheet)
            cur += 1
            self._auto_fit_columns(ws, min_widths={1: 30, 2: 120})
            
            # 設定 Summary 頁面所有文字字體為 11
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        # 保留原有的字體屬性，只修改大小
                        current_font = cell.font
                        cell.font = Font(
                            name='Microsoft JhengHei',
                            size=11,
                            bold=current_font.bold,
                            italic=current_font.italic,
                            color=current_font.color,
                            underline=current_font.underline
                        )
            
            try:
                wb.save(output_path)
            finally:
                try:
                    wb.close()
                except Exception:
                    pass

    def _write_raw_log_with_annotations(self, ws, start_row: int, raw_lines: list, annotations: list, font: Font, step_marks: dict | None = None):
        color_map = {
            'black': 'FF000000',
            'red': 'FFE74C3C',
            'green': 'FF2ECC71',
            'blue': 'FF3498DB',
            'purple': 'FF9B59B6',
        }
        marks = step_marks or {}
        error_line_found = False
        first_error_row = None
        
        # 以 annotations 的 color 欄位對應文字顏色；在步驟起始行前加上 1. 2. ... 標號
        for i, raw in enumerate(raw_lines, start=start_row):
            src_idx = i - start_row  # 對應原始 raw_lines 索引
            # 使用統一的清理函數
            line = str(raw)
            
            # 檢查是否為錯誤行
            is_error_line = False
            line_lower = line.lower()
            if any(keyword in line_lower for keyword in [
                'is fail', 'segmentation fault', 'core dumped', 'executes fail', 
                "doesn't match", 'timeout', 'exception', 'error', 'fail'
            ]):
                is_error_line = True
                if not error_line_found:
                    first_error_row = i
                    error_line_found = True
            
            # 先檢查是否有步驟標號
            if src_idx in marks:
                line = f"{marks[src_idx]}. {line}"
            
            # 清理行內容
            line = self._sanitize_cell_text(line)
            
            # 套用樣式：找到對應的 annotation
            cell = ws.cell(row=i, column=1, value=line)
            cell.font = font
            cell.number_format = '@'
            
            # 突出顯示錯誤行
            if is_error_line:
                cell.font = Font(name=font.name, size=font.size, color='FF0000', bold=True)
                cell.fill = PatternFill('solid', fgColor='FFFFFF99')  # 淺黃色背景
            
            # 預設
            found_anno = None
            for anno in annotations:
                if anno.get('line_idx') == src_idx:
                    found_anno = anno
                    break
            if found_anno and 'color' in found_anno:
                color_name = found_anno['color']
                if color_name in color_map:
                    color_hex = color_map[color_name]
                    cell.font = Font(name=font.name, size=font.size, color=color_hex, bold=font.bold)
        
        # 如果有找到錯誤行，設定超連結到第一個錯誤行
        if first_error_row:
            # 在A1儲存格添加超連結到第一個錯誤行
            ws.cell(row=1, column=1).hyperlink = f"#{ws.title}!A{first_error_row}"
        
        # 移除重複的回到Summary連結，因為已經在_build_pass_workbook中添加了

    def _build_detailed_error_summary(self, entry: dict) -> str:
        """建立詳細的錯誤摘要，包含主要錯誤和執行指令"""
        try:
            fail_items = entry.get('fail_items', [])
            raw_lines = entry.get('raw_lines', [])
            
            if not fail_items and not raw_lines:
                return "未知錯誤"
            
            # 收集所有錯誤資訊
            error_details = []
            commands = []
            critical_errors = []
            herr_errors = []
            
            # 從 raw_lines 中提取錯誤資訊
            for line in raw_lines:
                line_str = str(line).strip()
                if not line_str:
                    continue
                    
                line_lower = line_str.lower()
                
                # 收集 HERR 錯誤
                if 'herr:' in line_lower:
                    herr_errors.append(line_str)
                
                # 收集嚴重錯誤
                if any(keyword in line_lower for keyword in [
                    'segmentation fault', 'core dumped', 'executes fail', 
                    "doesn't match", 'timeout', 'exception', 'all test aborted'
                ]):
                    critical_errors.append(line_str)
                
                # 收集一般錯誤
                if any(keyword in line_lower for keyword in [
                    'is fail', 'error', 'fail'
                ]):
                    if line_str not in error_details:
                        error_details.append(line_str)
            
            # 從 fail_items 中提取資訊
            for item in fail_items:
                # 主要錯誤原因
                error = item.get('error', '')
                if error and error not in error_details:
                    error_details.append(error)
                
                # 執行的指令
                command = item.get('command', '')
                if command and command not in commands:
                    commands.append(command)
                
                # 完整回應中的錯誤資訊
                full_response = item.get('full_response', '')
                if full_response:
                    lines = full_response.split('\n')
                    for line in lines:
                        line = line.strip()
                        if not line:
                            continue
                        # 不再過濾關鍵字，因為回應中可能包含關鍵數值 (如 AVE_SNR = 28.8...)
                        if line not in error_details:
                            error_details.append(line)
            
            # 組合詳細摘要
            summary_parts = []
            
            # 主要錯誤原因 (來自 main_enhanced.py 的 prioritized selection)
            main_error = (entry.get('summary') or {}).get('FAIL原因', '')
            
            # === STRICT PRIORITY FILTERING ===
            # 如果主要錯誤是 "doesn't match" 或 "is Fail"，僅顯示該錯誤，忽略其他雜訊
            is_priority_error = main_error and (
                "doesn't match" in main_error.lower() or 
                "is fail" in main_error.lower()
            )
            
            if is_priority_error:
                summary_parts.append(f"===============錯誤原因====================")
                summary_parts.append("")
                
                if "doesn't match" in main_error.lower():
                    summary_parts.append("🔴 突出錯誤 (doesn't match):")
                
                summary_parts.append(main_error)
                
                # 僅附加執行指令 (為了方便對照)
                if commands:
                    # 嘗試找到與該錯誤相關的指令
                    related_cmd = None
                    for item in fail_items:
                        if item.get('error') == main_error:
                            related_cmd = item.get('command')
                            break
                    if related_cmd:
                        summary_parts.append(f"執行指令: {related_cmd}")
                
                return '\n'.join(summary_parts)
                
            # === Fallback: Generic Reporting (for generic FAIL/ERROR) ===
            if main_error:
                summary_parts.append(f"===============錯誤原因====================")
                summary_parts.append("")
                summary_parts.append(main_error)
                
                summary_parts.append("")
                summary_parts.append("=" * 50)
                summary_parts.append("")
            
            # 嚴重錯誤 (Filtered)
            if critical_errors:
                unique_critical = []
                for error in critical_errors[:5]:
                    if error != main_error and error not in unique_critical:
                        unique_critical.append(error)
                if unique_critical:
                    for error in unique_critical:
                        summary_parts.append(error)
                    summary_parts.append("")
            
            # HERR 錯誤
            if herr_errors:
                for error in herr_errors[:5]:
                    if error != main_error:
                        summary_parts.append(error)
                summary_parts.append("")
            
            # 其他錯誤詳情
            if error_details:
                for detail in error_details[:3]:
                    if detail != main_error and detail not in summary_parts:
                        summary_parts.append(detail)
                summary_parts.append("")
            
            # 執行指令
            if commands:
                unique_commands = list(dict.fromkeys(commands))
                for cmd in unique_commands[:3]:
                    summary_parts.append(f"執行指令: {cmd}")
            
            return '\n'.join(summary_parts) if summary_parts else "未知錯誤"
            
        except Exception as e:
            return f"錯誤摘要生成失敗: {str(e)}"
    
    def _add_error_statistics(self, ws, logs: list):
        """在Summary頁面最上面添加錯誤統計"""
        try:
            # 收集所有錯誤原因
            error_counts = {}
            
            for entry in logs:
                main_error = (entry.get('summary') or {}).get('FAIL原因', '')
                if main_error:
                    # 清理錯誤原因，提取主要部分（標準化群組鍵）
                    group_key = self._normalize_error_group(main_error)
                    if group_key:
                        error_counts[group_key] = error_counts.get(group_key, 0) + 1
            
            if not error_counts:
                print("沒有找到錯誤原因，跳過統計")
                return
            
            print(f"找到 {len(error_counts)} 種錯誤類型: {list(error_counts.keys())}")
            
            # 直接在現有內容上方添加錯誤統計，不移動現有內容
            # 添加錯誤統計標題
            title_row = 1
            ws.cell(row=title_row, column=1, value='📊 錯誤原因統計').font = Font(name='Microsoft JhengHei', size=14, bold=True, color='FF0000')
            ws.cell(row=title_row, column=1).fill = PatternFill('solid', fgColor='FFFFE6E6')
            
            # 添加統計內容
            current_row = 2
            for error_type, count in sorted(error_counts.items(), key=lambda x: x[1], reverse=True):
                ws.cell(row=current_row, column=1, value=f'{error_type} = {count} 筆')
                ws.cell(row=current_row, column=1).font = Font(name='Microsoft JhengHei', size=11, bold=True)
                ws.cell(row=current_row, column=1).fill = PatternFill('solid', fgColor='FFFFFF99')  # 淺黃色背景
                current_row += 1
            
            # 添加分隔線
            ws.cell(row=current_row, column=1, value='─' * 50).font = Font(name='Microsoft JhengHei', size=10, color='FF808080')
            
            print(f"錯誤統計已添加到第 {title_row} 到 {current_row} 行")
            
        except Exception as e:
            print(f"添加錯誤統計時發生錯誤: {e}")
            import traceback
            traceback.print_exc()

    def _normalize_error_group(self, error_text: str) -> str:
        """將錯誤字串標準化為群組鍵，用於彙總相似錯誤。
        規則：
        - 先抽出主要錯誤片段（沿用 _extract_main_error_type）
        - 若包含冒號（如 B7PL011-202:TRY_TEXT is Fail），去除編號僅保留冒號後文本
        - 若包含 "is Fail"，統一裁切為『... is Fail』
        - 全部忽略多餘空白與大小寫（但回傳時維持原始大小寫格式）
        """
        try:
            base = self._extract_main_error_type(error_text) or error_text or ''
            s = str(base).strip()
            # 去除前置代碼（XXXX-123: 或 XXXX:）
            import re
            if ':' in s:
                parts = s.split(':', 1)
                # 若冒號後為主要內容，採用它
                s = parts[1].strip() or s
            # 對 "is Fail" 做標準化，只取到 is Fail 結尾
            m = re.search(r'(.+?\bis\s*Fail)\b', s, flags=re.IGNORECASE)
            if m:
                s = m.group(1)
            # 移除多餘空白
            s = re.sub(r'\s+', ' ', s).strip()
            # 標準化大小寫（回傳人類可讀格式）
            # 盡量保留原樣，僅確保 Fail 大寫
            s = re.sub(r'\bis\s*fail\b', 'is Fail', s, flags=re.IGNORECASE)
            return s
        except Exception:
            return (error_text or '').strip()
    
    def _build_pass_steps_summary(self, pass_items: list) -> str:
        """建立 PASS 步驟摘要，以簡潔格式列出所有完成的步驟"""
        try:
            if not pass_items:
                return "無 PASS 步驟"
            
            summary_parts = []
            summary_parts.append("===============PASS步驟詳情====================")
            summary_parts.append("")
            
            # 只顯示步驟名稱，讓使用者一眼看清楚測了多少項目
            for i, item in enumerate(pass_items, 1):
                step_name = item.get('step_name', f'步驟 {i}')
                summary_parts.append(f"✅ 步驟 {i}: {step_name}")
            
            summary_parts.append("")
            summary_parts.append(f"總共完成 {len(pass_items)} 個測試步驟")
            summary_parts.append("=" * 50)
            return '\n'.join(summary_parts)
            
        except Exception as e:
            return f"PASS 步驟摘要生成失敗: {str(e)}"

    def _extract_main_error_type(self, error_text: str) -> str:
        """從錯誤文字中提取主要錯誤類型"""
        try:
            if not error_text:
                return ""
            
            # 移除時間戳記和錯誤代碼
            clean_text = error_text
            
            # 移除時間戳記格式 (如 "2025/09/23 10:48:12 [1]")
            import re
            clean_text = re.sub(r'\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2} \[\d+\]', '', clean_text)
            
            # 移除錯誤代碼 (如 "<ErrorCode: AFFW57>")
            clean_text = re.sub(r'<ErrorCode: [^>]+>', '', clean_text)
            
            # 移除測試時間 (如 "----- 61.019937 Sec.")
            clean_text = re.sub(r'----- \d+\.\d+ Sec\.', '', clean_text)
            
            # 清理多餘空格
            clean_text = clean_text.strip()
            
            # 優先提取 "doesn't match" 相關錯誤
            if "doesn't match" in clean_text.lower():
                # 尋找 "doesn't match" 後面的內容
                match_pattern = r"doesn't match\s+([^@\n]+)"
                match_result = re.search(match_pattern, clean_text, re.IGNORECASE)
                if match_result:
                    return f"doesn't match {match_result.group(1).strip()}"
                else:
                    return "doesn't match"
            
            # 提取主要錯誤類型
            if ':' in clean_text:
                # 格式如 "B7PL011-202:Chec Frimware version is Fail"
                parts = clean_text.split(':', 1)
                if len(parts) > 1:
                    main_part = parts[1].strip()
                    # 移除測試編號前綴
                    main_part = re.sub(r'^[A-Z0-9]+-\d+:', '', main_part)
                    # 移除行號前綴 (如 "45 [1]" 或 "27 [1]")
                    main_part = re.sub(r'^\d+\s*\[\d+\]', '', main_part)
                    return main_part.strip()
            
            # 如果沒有冒號，直接返回清理後的文字
            return clean_text
            
        except Exception:
            return error_text

    def _auto_fit_columns(self, ws, min_widths: dict | None = None):
        """自動調整欄寬"""
        min_w = min_widths or {}
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            col_num = col[0].column
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max(max_length + 2, min_w.get(col_num, 10))
            adjusted_width = min(adjusted_width, 150)  # 最大寬度限制
            ws.column_dimensions[column].width = adjusted_width