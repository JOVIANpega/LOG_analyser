# -*- coding: utf-8 -*-
"""
Summary Builder Module
Handles creation of Summary and Data sheets for both PASS and FAIL workbooks.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .excel_utils import (
    sanitize_cell_text,
    extract_isn_from_filename,
    extract_station_from_filename,
    format_filename_with_timestamp,
    generate_header_info_text,
    auto_fit_columns
)

class SummaryBuilder:
    """構建 Excel 匯總頁面 (Summary/Data)"""

    def __init__(self):
        self.header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        self.content_font = Font(name='Calibri', size=11)
        self.fill_blue = PatternFill('solid', fgColor='4472C4')
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_top_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    def create_summary_sheet(self, wb, logs, title="Summary", prefix=""):
        """建立匯總頁面 (包含置頂統計與時間曲線)"""
        ws = wb.create_sheet(title, 0)
        
        # 1. 預留頂部空間並插入匯總資訊 (A1:B6 區域)
        # 這裡需要先計算出 logs 的時間資料
        time_labels_data = []
        for entry in logs:
            fname = entry.get('file_name', entry.get('filename', ''))
            isn = extract_isn_from_filename(fname)
            from .excel_utils import extract_total_secs
            test_secs, _ = extract_total_secs(entry.get('raw_lines', []))
            if test_secs:
                label = isn if isn else fname[:20]
                time_labels_data.append((label, test_secs))
        
        self._add_summary_header(ws, len(logs), time_labels_data, prefix=prefix)
        
        # 2. 寫入清單標題 (從第 9 行開始)
        list_start_row = 9
        headers = ["ISN (點擊跳轉)", "檔案分析", "站別", "時間 (Sec)", "狀態"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=list_start_row, column=col, value=h)
            cell.font = self.header_font
            cell.fill = self.fill_blue
            cell.alignment = self.center_align
            
        # 3. 填寫清單資料
        for row_idx, entry in enumerate(logs, 0):
            curr_row = list_start_row + 1 + row_idx
            fname = entry.get('file_name', entry.get('filename', ''))
            isn = extract_isn_from_filename(fname)
            station = extract_station_from_filename(fname)
            
            # 提取時間
            from .excel_utils import extract_total_secs
            test_secs, _ = extract_total_secs(entry.get('raw_lines', []))
            test_time_str = f"{test_secs:.2f}" if test_secs else "Unknown"
            
            # 工作表名稱
            sheet_name = entry.get('sheet_name', isn or fname[:25])
            
            # A欄: ISN (綠色 + 連結)
            # cell_isn = ws.cell(row=curr_row, column=1, value=isn if isn else "Unknown ISN")
            
            isn_text = isn if isn else "Unknown ISN"
            safe_isn = isn_text.replace('"', '""')
            cell_isn = ws.cell(row=curr_row, column=1)
            cell_isn.value = f'=HYPERLINK("#\'{sheet_name}\'!A1", "{safe_isn}")'
            
            cell_isn.font = Font(name='Calibri', size=11, color='2E7D32', bold=True, underline='single')
            # cell_isn.hyperlink = f"#'{sheet_name}'!A1"
            cell_isn.alignment = self.center_align
            
            # B欄: 檔案名稱
            cell_file = ws.cell(row=curr_row, column=2, value=fname)
            cell_file.font = self.content_font
            cell_file.alignment = self.left_top_align
            
            # C欄: 站別
            cell_station = ws.cell(row=curr_row, column=3, value=station)
            cell_station.font = self.content_font
            cell_station.alignment = self.center_align
            
            # D欄: 時間
            cell_time = ws.cell(row=curr_row, column=4, value=test_time_str)
            cell_time.font = self.content_font
            cell_time.alignment = self.center_align
            
            # E欄: 狀態
            status = "PASS" if not entry.get('fail_items') else "FAIL"
            cell_status = ws.cell(row=curr_row, column=5, value=status)
            cell_status.alignment = self.center_align
            if status == "FAIL":
                cell_status.font = Font(name='Calibri', size=11, color="FF0000", bold=True)
            else:
                cell_status.font = self.content_font
            
        # 優化欄寬：ISN寬度適中，檔案分析欄位最寬
        auto_fit_columns(ws, {1: 22, 2: 55, 3: 30, 4: 15, 5: 10})
        
        # 4. 新增圖表
        if time_labels_data:
            self._add_time_distribution_chart(ws, time_labels_data)
            
        return ws

    def _add_summary_header(self, ws, total_count, time_data, prefix=""):
        """插入頂部置頂資訊與統計 (同步 FAIL 風格)"""
        # 標題列
        ws.merge_cells('A1:F1')
        clean_prefix = prefix.strip('_ ')
        display_title = f" 📋 {clean_prefix} PASS 分析匯總報告 " if clean_prefix else " 📋 PASS 分析匯總報告 (Summary & List) "
        title_cell = ws.cell(row=1, column=1, value=display_title)
        title_cell.font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill('solid', fgColor='2E7D32') # 綠色
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 統計數據 (A3:B6)
        avg_time = sum(d[1] for d in time_data) / len(time_data) if time_data else 0
        max_time = max(d[1] for d in time_data) if time_data else 0
        min_time = min(d[1] for d in time_data) if time_data else 0

        stats = [
            ("項目數量", f"{total_count} 筆"),
            ("平均時間", f"{avg_time:.2f} Sec"),
            ("最長時間", f"{max_time:.2f} Sec"),
            ("最短時間", f"{min_time:.2f} Sec")
        ]

        for i, (label, val) in enumerate(stats):
            row = 3 + i
            # Label
            lbl_cell = ws.cell(row=row, column=1, value=label)
            lbl_cell.font = Font(name='Calibri', size=11, bold=True)
            # Value
            v_cell = ws.cell(row=row, column=2, value=val)
            v_cell.font = Font(name='Calibri', size=11)
            v_cell.alignment = Alignment(horizontal='center')

    def _add_time_distribution_chart(self, ws, time_data):
        """新增測試時間分佈圖 (採用最穩定的資料引用方式)"""
        try:
            # 1. 寫入數據到 Z 欄之後 (Column 26+)
            data_col = 26 # Z
            ws.cell(row=1, column=data_col, value="ID").font = self.content_font
            ws.cell(row=1, column=data_col+1, value="Time").font = self.content_font
            
            for i, (label, val) in enumerate(time_data):
                # ID (Z欄) - 加上 # 符號強力強制 Excel 視為字串標籤，避免大數值座標軸錯亂
                c_id = ws.cell(row=i+2, column=data_col, value=f"#{label}")
                c_id.font = self.content_font
                # Time (AA欄)
                c_val = ws.cell(row=i+2, column=data_col+1, value=float(val))
                c_val.font = self.content_font
                
            # 2. 建立圖表
            from openpyxl.chart import LineChart, Reference
            from openpyxl.drawing.fill import SolidFillProperties
            
            chart = LineChart()
            chart.title = "測試時間分佈圖 (Time Curve / Sec)"
            chart.y_axis.title = "Time (Sec)"
            chart.x_axis.title = "Logs"
            chart.y_axis.majorUnit = 100
            
            # 使用當前頁面的 Reference (不跨頁，最穩定)
            # data_ref 指向數值欄位 (AA)
            data_ref = Reference(ws, min_col=data_col+1, min_row=1, max_row=len(time_data)+1)
            # cats_ref 指向標籤欄位 (Z)
            cats_ref = Reference(ws, min_col=data_col, min_row=2, max_row=len(time_data)+1)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            # 3. 配置系列樣式 (綠線紅點)
            if chart.series:
                s = chart.series[0]
                s.marker.symbol = "circle"
                s.marker.size = 7
                # 線條與點的顏色設定 (確保在所有 Excel 版本可見)
                s.graphicalProperties.line.solidFill = SolidFillProperties(srgbClr="2E7D32")
                s.marker.graphicalProperties.solidFill = SolidFillProperties(srgbClr="FF0000")
                s.marker.graphicalProperties.line.solidFill = SolidFillProperties(srgbClr="FF0000")
                
            chart.legend = None
            chart.width = 32
            chart.height = 11
            
            # 4. 錨定在 G3
            ws.add_chart(chart, "G3")
            
        except Exception as e:
            import traceback
            print(f"[ERROR] PASS 圖表生成失敗: {str(e)}")
            traceback.print_exc()

    def _add_time_statistics_table(self, ws, times):
        """(過時) 舊的時間統計表，目前由 _add_time_distribution_chart 取代"""
        pass

