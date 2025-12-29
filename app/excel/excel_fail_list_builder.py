# -*- coding: utf-8 -*-
"""
FAIL_LIST Builder Module
Handles the construction of the FAIL_LIST sheet and error statistics.
"""
import re
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from .excel_utils import (
    sanitize_cell_text, 
    extract_isn_from_filename, 
    extract_station_from_filename,
    auto_fit_columns
)

class FailListBuilder:
    """構建 FAIL_LIST 工作表與錯誤統計"""
    
    def __init__(self):
        self.header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        self.content_font = Font(name='Calibri', size=11)
        self.fill_blue = PatternFill('solid', fgColor='4472C4')
        self.fill_red = PatternFill('solid', fgColor='FF0000')
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.top_left_align = Alignment(wrap_text=True, vertical='top', horizontal='left')

    def build_fail_list_sheet(self, wb, logs):
        """建立 FAIL_LIST 工作表 (依使用者要求的 CSV 格式，並統計相同錯誤)"""
        ws = wb.create_sheet("FAIL_LIST", 0) # 放在第一頁
        try:
            ws.sheet_properties.tabColor = 'FFFF0000' # 紅色標籤
        except Exception:
            pass
        
        # 設定標題 (新增 Count 欄位)
        headers = ['ISN', 'Station', 'FAIL Item', 'FAIL Reason', 'suggestion', 'Count']
        ws.append(headers)
        
        # 標題樣式
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_font
            # FAIL Reason (Column D) 使用紅色背景
            if col == 4:
                cell.fill = self.fill_red
            else:
                cell.fill = self.fill_blue
            cell.alignment = self.center_align
        
        # 1. 收集所有資料行與統計錯誤
        pending_rows = []
        error_counts = {}
        
        for entry in logs:
            # --- 核心改進：每份 Log 僅在 FAIL_LIST 記錄一個「最主要」的錯誤 ---
            # 優先使用 parser 判定出的 last_fail (內含優先級邏輯)
            primary_fail = entry.get('last_fail')
            
            # 如果沒有判定出 Last Fail，則從 fail_items 中抓取最後一項
            if not primary_fail and entry.get('fail_items'):
                primary_fail = entry['fail_items'][-1]
            
            if not primary_fail:
                continue
                
            fname = entry.get('file_name', entry.get('filename', ''))
            isn = extract_isn_from_filename(fname)
            station = extract_station_from_filename(fname)
            suggestion = "請參閱 PEGA SOP"

            # 提取 Item 名稱
            step_name = primary_fail.get('step_name', '')
            result_val = primary_fail.get('result', '')
            fail_item_str = f"{step_name}"
            if result_val and result_val not in step_name:
                 fail_item_str += f" {result_val}"
            
            # 清理 Item 名稱
            fail_item_str = re.sub(r'\s+\b(is\s+)?FAIL\b.*$', '', fail_item_str, flags=re.IGNORECASE).strip()
            fail_item_str = fail_item_str.strip() or "Unknown Item"
            
            # 從該失敗項目的 Log 中尋找最真實的 Reason (優先搜尋 DOESN'T MATCH)
            reason = ""
            full_log = primary_fail.get('full_log', [])
            for log_line in reversed(full_log):
                if "doesn't match" in log_line.lower():
                    reason = log_line.strip()
                    break
            
            if not reason:
                reason = primary_fail.get('error', '')
            if not reason or reason == 'FAIL':
                reason = primary_fail.get('response', '')
            
            # === 新增：計算錯誤行在 Excel 中的實際位置 ===
            error_excel_row = None
            sheet_name = entry.get('sheet_name', '')
            raw_lines = entry.get('raw_lines', [])
            
            if raw_lines and sheet_name:
                # 使用與 FAIL Reason 相同的 bottom-up 優先級查找錯誤行
                error_line_idx = None
                dm_pattern = re.compile(r"doesn't match", re.IGNORECASE)
                is_fail_pattern = re.compile(r"is Fail", re.IGNORECASE)
                fail_pattern = re.compile(r"FAIL", re.IGNORECASE)
                
                # 優先級 1: doesn't match
                for i in range(len(raw_lines)-1, -1, -1):
                    if dm_pattern.search(raw_lines[i]):
                        error_line_idx = i
                        break
                
                # 優先級 2: is Fail
                if error_line_idx is None:
                    for i in range(len(raw_lines)-1, -1, -1):
                        if is_fail_pattern.search(raw_lines[i]):
                            error_line_idx = i
                            break
                
                # 優先級 3: FAIL
                if error_line_idx is None:
                    for i in range(len(raw_lines)-1, -1, -1):
                        if fail_pattern.search(raw_lines[i]):
                            error_line_idx = i
                            break
                
                # 計算 Excel 行號 (考慮 header_info 和其他偏移)
                if error_line_idx is not None:
                    # Row 1: [回到 Summary]
                    # Row 3+: Header Info (假設 ~5 行)
                    # 之後是錯誤預覽框 (如果有的話，動態計算)
                    # 再之後是實際 Log
                    # 簡化計算：假設 actual_log_start 約在第 10-15 行
                    # 更精確的做法是從 entry 中取得，但目前先用估算
                    header_info_lines = len(entry.get('header_info', '').split('\n')) if entry.get('header_info') else 3
                    error_preview_lines = 10  # 錯誤預覽框估計佔用行數
                    actual_log_start = 1 + 1 + header_info_lines + error_preview_lines
                    error_excel_row = actual_log_start + error_line_idx
            
            # 統計
            error_counts[fail_item_str] = error_counts.get(fail_item_str, 0) + 1
            
            pending_rows.append({
                'isn': isn,
                'station': station,
                'item': fail_item_str,
                'reason': reason,
                'suggestion': suggestion,
                'norm_key': fail_item_str,
                'sheet_name': sheet_name,  # ⚠️ 新增
                'error_row': error_excel_row  # ⚠️ 新增
            })
        
        # 2. 排序資料 (依內容分類，讓同類型錯誤擺在一起)
        pending_rows.sort(key=lambda x: (x['item'], x['isn']))
        
        # 3. 寫入資料
        for row_data in pending_rows:
            count = error_counts.get(row_data['norm_key'], 0)
            
            row_values = [
                sanitize_cell_text(row_data['isn']),
                sanitize_cell_text(row_data['station']),
                sanitize_cell_text(row_data['item']),
                sanitize_cell_text(row_data['reason']),
                sanitize_cell_text(row_data['suggestion']),
                count
            ]
            ws.append(row_values)
            
            current_row = ws.max_row
            for col_idx, value in enumerate(row_values, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = self.content_font
                
                # === ISN 列添加超鏈接 (跳轉到錯誤行) ===
                if col_idx == 1 and row_data.get('sheet_name') and row_data.get('error_row'):
                    target_sheet = row_data['sheet_name']
                    target_row = row_data['error_row']
                    cell.hyperlink = f"#'{target_sheet}'!A{target_row}"
                    cell.font = Font(name='Consolas', size=12, color='0563C1', underline='single', bold=True)
                    cell.alignment = self.center_align
                elif col_idx in (1, 2, 3, 6): # ISN, Station, FAIL Item, Count 不換行且置中
                    cell.alignment = self.center_align
                elif col_idx == 4: # FAIL Reason 換行且靠上
                    cell.alignment = self.top_left_align
                else: # suggestion 置中靠上
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                
        # 自動調整欄寬
        min_widths = {
            1: 25, # ISN
            2: 30, # Station
            3: 45, # FAIL Item
            4: 65, # FAIL Reason
            5: 20, # suggestion
            6: 12  # Count
        }
        auto_fit_columns(ws, min_widths=min_widths)
        
        # 恢復錯誤統計表格 (使用者要求保留表格，移除圓餅圖)
        if error_counts:
            self._add_error_statistics_table(ws, error_counts)
            
        return ws

    def _add_error_statistics_table(self, ws, error_counts):
        """在 FAIL_LIST 頁面添加錯誤統計表 (依使用者要求恢復)"""
        try:
            table_start_row = ws.max_row + 3
            
            ws.cell(row=table_start_row, column=1, value="FAIL Item 統計 (記數)").font = Font(name='Calibri', size=12, bold=True)
            ws.cell(row=table_start_row + 1, column=1, value="FAIL Item 名稱").font = self.header_font
            ws.cell(row=table_start_row + 1, column=2, value="數量").font = self.header_font
            ws.cell(row=table_start_row + 1, column=3, value="佔比").font = self.header_font
            
            # 套用標題樣式
            for col in range(1, 4):
                cell = ws.cell(row=table_start_row + 1, column=col)
                cell.fill = self.fill_blue
                cell.alignment = self.center_align
            
            total_errors = sum(error_counts.values())
            summary_start_row = table_start_row + 2
            
            sorted_items = sorted(error_counts.items(), key=lambda x: x[1], reverse=True)
            yellow_fill = PatternFill('solid', fgColor='FFFF00')
            from openpyxl.styles import Border, Side
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
            
            for idx, (error_type, count) in enumerate(sorted_items):
                row = summary_start_row + idx
                percentage = (count / total_errors * 100) if total_errors > 0 else 0
                
                c1 = ws.cell(row=row, column=1, value=sanitize_cell_text(error_type))
                c2 = ws.cell(row=row, column=2, value=count)
                c3 = ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
                
                # 套用樣式與黃色背景
                for c in [c1, c2, c3]:
                    c.font = Font(name='Calibri', size=11)
                    c.fill = yellow_fill
                    c.border = thin_border
                    
                c2.alignment = self.center_align
                c3.alignment = self.center_align
        except Exception as e:
            print(f"[WARNING] 錯誤統計表生成失敗: {e}")

    def _add_pie_chart(self, ws, error_counts):
        """原有的圓餅圖函數 (目前不再調用，保留供未來參考)"""
        try:
            # 原本的完整統計邏輯在上面已經抽離為 _add_error_statistics_table
            pass
        except Exception:
            pass

