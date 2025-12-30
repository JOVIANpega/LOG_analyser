# -*- coding: utf-8 -*-
"""
Sheet Builder Module
Handles detailed log writing and annotation.
"""
import re
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from .excel_utils import sanitize_cell_text

def write_raw_log_with_annotations(ws, start_row: int, raw_lines: list, annotations: list, font, fail_items=None, log_type='UNKNOWN'):
    """將原始 LOG 與顏色標籤寫入 Excel (同步 GUI 的 Premium 外觀)"""
    color_map = {
        'black': 'FF000000',
        'red': 'FFFF0000',
        'green': 'FF28a745',
        'blue': 'FF007bff',
        'purple': 'FF6f42c1',
        'pink': 'FFe83e8c'
    }
    
    # 建立一個索引查表
    anno_lookup = {a['line_idx']: a for a in annotations}
    
    # ⚠️ 僅對 FAIL 日誌執行錯誤檢測與預覽生成
    error_line_idx = None
    error_block_preview = []
    
    if log_type == 'FAIL':
        # 優先級偵測錯誤點 (從最下面往上找)
        dm_pattern = re.compile(r"doesn't match", re.IGNORECASE)
        is_fail_pattern = re.compile(r"is Fail", re.IGNORECASE)
        abort_pattern = re.compile(r"All Test Aborted", re.IGNORECASE)
        status_pattern = re.compile(r"Status:False", re.IGNORECASE)
        fail_pattern = re.compile(r"FAIL", re.IGNORECASE)
        error_pattern = re.compile(r"ERROR", re.IGNORECASE)
    
        # 1. 最高優先級：DOESN'T MATCH
        for i in range(len(raw_lines)-1, -1, -1):
            if dm_pattern.search(raw_lines[i]):
                error_line_idx = i
                break
        
        # 2. 次優先級：is Fail / All Test Aborted
        if error_line_idx is None:
            for i in range(len(raw_lines)-1, -1, -1):
                cur_line = raw_lines[i]
                if is_fail_pattern.search(cur_line) or abort_pattern.search(cur_line):
                    error_line_idx = i
                    break
        
        # 3. 第三優先級：Status:False
        if error_line_idx is None:
            for i in range(len(raw_lines)-1, -1, -1):
                if status_pattern.search(raw_lines[i]):
                    error_line_idx = i
                    break
        
        # 3. 第三優先級：FAIL
        if error_line_idx is None:
            for i in range(len(raw_lines)-1, -1, -1):
                if fail_pattern.search(raw_lines[i]):
                    error_line_idx = i
                    break
        
        # 4. 最後備選：ERROR
        if error_line_idx is None:
            for i in range(len(raw_lines)-1, -1, -1):
                if error_pattern.search(raw_lines[i]):
                    error_line_idx = i
                    break
        
        # 擷取錯誤區塊文字 (用於置頂顯示預覽)
        if error_line_idx is not None:
            b_start = error_line_idx
            for j in range(error_line_idx, max(-1, error_line_idx - 20), -1):
                if '>' in raw_lines[j] or 'Do @STEP' in raw_lines[j]:
                    b_start = j
                    break
            b_end = error_line_idx
            for j in range(error_line_idx, min(len(raw_lines), error_line_idx + 10)):
                b_end = j
                if j > error_line_idx and ('>' in raw_lines[j] or 'Do @STEP' in raw_lines[j]):
                    b_end = j - 1
                    break
                if 'Test Completed' in raw_lines[j] or 'executes fail' in raw_lines[j]:
                    break
            error_block_preview = raw_lines[b_start : b_end + 1]

    # --- 步驟 A: 預先計算預覽區塊高度以利超連結定位 ---
    preview_box_height = 0
    
    # FAIL 日誌：錯誤預覽框
    if error_line_idx is not None and log_type == 'FAIL':
        preview_box_height = 1 + len(error_block_preview) + 1 + 1
    
    # PASS/FAIL 日誌：測試項目預覽框 (按 Phase 分組)
    pass_phases_preview = {} # phase_name -> [validations]
    if fail_items:
        # 將所有項按照 Phase 進行分組
        for item in fail_items:
            p_name = item.get('phase', 'Unknown Phase')
            if p_name not in pass_phases_preview:
                pass_phases_preview[p_name] = []
            
            # 收集此測項內的所有比對結果
            validations = item.get('validations', [])
            for v in validations:
                pass_phases_preview[p_name].append({
                    'content': v.get('content', ''),
                    'status': v.get('status', 'PASS'),
                    'line_idx': v.get('line_idx')
                })
        
        # 預算顯示高度：每個 Phase 標題(1) + 每個 Validation(1)
        # 限制總行數避免預覽框過長 (增加限制到 100 行，確保能看到後面的項)
        temp_row_count = 0
        for p_name, v_list in pass_phases_preview.items():
            temp_row_count += 1 # 標題
            temp_row_count += len(v_list) # 比對項
            if temp_row_count > 100: break # 硬限制
        
        if temp_row_count > 0:
            preview_box_height = 1 + temp_row_count + 1
    
    actual_log_start = start_row + preview_box_height
    curr_h_row = start_row

    # --- 步驟 B1: FAIL 日誌寫入錯誤預覽區塊 ---
    if error_line_idx is not None and log_type == 'FAIL':
        # 1. 標題行
        title_font = Font(name='Microsoft JhengHei', size=12, bold=True, color='FFFFFF')
        title_fill = PatternFill('solid', fgColor='FFFF0000') # 純紅標題
        title_cell = ws.cell(row=curr_h_row, column=1, value="  [ 發現錯誤點 (預覽) ]  ")
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        curr_h_row += 1
        
        # 2. 錯誤內容 (每行皆可點擊跳轉)
        # 使用剛才擷取預覽塊時的 b_start 作為索引基底
        pink_fill = PatternFill('solid', fgColor='FFFFE1E1')
        red_font = Font(name='Consolas', size=11, color='FFC00000', bold=True)
        
        if error_block_preview:
            for i, line_p in enumerate(error_block_preview):
                target_log_idx = b_start + i
                target_excel_row = actual_log_start + target_log_idx
                
                cp = ws.cell(row=curr_h_row, column=1, value="  >> " + sanitize_cell_text(line_p))
                cp.font = red_font
                cp.fill = pink_fill
                
                # 建立超連結跳轉到下方對應位置
                cp.hyperlink = f"#'{ws.title}'!A{target_excel_row}"
                curr_h_row += 1
        
        # 3. 底部跳轉按鈕 (整塊 Box 的結尾)
        target_err_row = actual_log_start + error_line_idx
        jump_cell = ws.cell(row=curr_h_row, column=1, value=f" [ 🚀 直接跳轉至錯誤行 Row {target_err_row} ] ")
        jump_cell.font = Font(name='Microsoft JhengHei', size=11, bold=True, color="FF0000BB", underline="single")
        jump_cell.fill = pink_fill
        jump_cell.alignment = Alignment(horizontal='center')
        jump_cell.hyperlink = f"#'{ws.title}'!A{target_err_row}"
        
        curr_h_row += 2 # 留空行
    
    # --- 步驟 B2: 寫入測試項目預覽區塊 (分組層次顯示) ---
    detail_preview_row = 1 # 預設第一行
    if pass_phases_preview:
        # 1. 標題行 (藍底白字)
        detail_preview_row = curr_h_row
        title_font = Font(name='Microsoft JhengHei', size=12, bold=True, color='FFFFFF')
        title_fill = PatternFill('solid', fgColor='FF4472C4')  # 藍色標題
        title_cell = ws.cell(row=curr_h_row, column=1, value="  [ 比對項目細節 ]  ")
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        curr_h_row += 1
        
        # 2. 內容渲染 (標題 + 比對細項)
        # Phase 標題：採用清晰的黑字
        phase_fill = PatternFill('solid', fgColor='FFF2F7FF') # 極淺藍(接近白)
        phase_font = Font(name='Consolas', size=11, color='FF000000', bold=True)
        
        # 比對細節：採用 GUI 風格的綠色系 (PASS 代表色)
        pass_detail_fill = PatternFill('solid', fgColor='FFE2EFDA') # 淺綠背景
        pass_detail_font = Font(name='Consolas', size=10, color='FF375623') # 深綠文字
        
        total_rendered = 0
        limit = 100 # 總量上限
        
        for phase_name, validations in pass_phases_preview.items():
            if total_rendered >= limit: break
            
            # --- 顯示 Phase 標題 (黑字) ---
            p_cell = ws.cell(row=curr_h_row, column=1, value=f"  ■ {phase_name}")
            p_cell.font = phase_font
            p_cell.fill = phase_fill
            curr_h_row += 1
            total_rendered += 1
            
            # --- 顯示該 Phase 下的所有比對項 (綠色系) ---
            for v in validations:
                if total_rendered >= limit: break
                
                v_content = v.get('content', '')
                v_status = v.get('status', 'PASS')
                mark = '✓' if v_status == 'PASS' else '✗'
                
                # 仿照 GUI 的層次感：縮排並加上圖示
                detail_text = f"     └ {mark} {v_content}"
                d_cell = ws.cell(row=curr_h_row, column=1, value=sanitize_cell_text(detail_text))
                
                # --- 新增：超連結跳轉到原始 Log 位置 ---
                target_log_idx = v.get('line_idx')
                if target_log_idx is not None:
                    target_excel_row = actual_log_start + target_log_idx
                    d_cell.hyperlink = f"#'{ws.title}'!A{target_excel_row}"
                    # 為可點擊項增加底線
                    # d_cell.font = Font(name='Consolas', size=10, color='FF375623', underline='single') 
                    # 考慮到表格美觀，暫不強制底線，滑鼠移上去會有手勢即可
                
                
                # 根據狀態決定顏色 (通常 PASS 為主)
                if v_status == 'PASS':
                    d_cell.font = pass_detail_font
                    d_cell.fill = pass_detail_fill
                else:
                    # 如果有異常，顯示紅粉色
                    d_cell.font = Font(name='Consolas', size=10, color='FFC00000')
                    d_cell.fill = PatternFill('solid', fgColor='FFFFE1E1')
                    
                curr_h_row += 1
                total_rendered += 1
        
        curr_h_row += 1  # 留空行

    # --- 步驟 D: 寫入正式 Log ---
    for i, raw in enumerate(raw_lines):
        row_idx = actual_log_start + i
        line = str(raw)
        anno = anno_lookup.get(i, {})
        
        # 章節標頭 (綠底白字)
        if anno.get('show_separator'):
            title = anno.get('separator_title', 'PHASE')
            sep_cell = ws.cell(row=row_idx, column=1, value=f" --   [ {title} ]")
            sep_cell.font = Font(name='Consolas', size=12, bold=True, color='FFFFFF', underline='single')
            sep_cell.fill = PatternFill('solid', fgColor='FF2E7D32') 
            sep_cell.alignment = Alignment(horizontal='center')
            # PHASE 加上超連結回預覽框 (或 Summary)
            if detail_preview_row > 1:
                sep_cell.hyperlink = f"#'{ws.title}'!A{detail_preview_row}"
            else:
                back_sheet = 'FAIL_LIST' if log_type == 'FAIL' else 'Summary'
                sep_cell.hyperlink = f"#'{back_sheet}'!A1"
            continue 

        # 一般日誌行
        cell = ws.cell(row=row_idx, column=1)
        cell.value = sanitize_cell_text(line)
        
        # 取得標註樣式
        txt_color_name = anno.get('color', 'black')
        txt_color = color_map.get(txt_color_name, color_map['black'])
        bg_hex = anno.get('background', 'white')
        is_bold = anno.get('is_bold', False)
        
        # 應用背景 (Excel 需要 'FFxxxxxx' 格式)
        if bg_hex and bg_hex.startswith('#'):
            cell.fill = PatternFill('solid', fgColor='FF' + bg_hex[1:])
        elif bg_hex == 'white':
            pass # 預設白色底
        
        # 應用字體
        cell.font = Font(name='Consolas', size=10, color=txt_color, bold=is_bold)

    error_excel_row = None
    if log_type == 'FAIL' and error_line_idx is not None:
        error_excel_row = actual_log_start + error_line_idx

    # 返回最後一行、錯誤行位置以及預覽框起始位置
    last_row = actual_log_start + len(raw_lines)
    return (last_row, error_excel_row, detail_preview_row)

def insert_header_info(ws, header_info, start_row=4):
    """插入置頂 Header 資訊"""
    if not header_info:
        return start_row
        
    lines = header_info.split('\n')
    header_font = Font(name='Consolas', size=14, bold=True, color='000000')
    header_fill = PatternFill('solid', fgColor='90EE90') # 淺綠
    
    for i, line in enumerate(lines):
        cell = ws.cell(row=start_row + i, column=1, value=line)
        cell.font = header_font
        cell.fill = header_fill
        
    return start_row + len(lines) + 1
