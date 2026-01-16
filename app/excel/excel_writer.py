# -*- coding: utf-8 -*-
"""
ExcelWriter Main Module (Modularized)
Coordinates the creation of PASS and FAIL workbooks.
"""
from __future__ import annotations
import os
import time
import traceback
from datetime import datetime
# Local imports
from .excel_utils import (
    auto_fit_columns, 
    unique_sheet_name,
    extract_isn_from_filename,
    extract_system_info,
    generate_header_info_text
)

def safe_save_workbook(wb, output_path):
    """安全保存工作簿，處理權限錯誤 (如檔案已開啟)"""
    if not os.path.exists(output_path):
        try:
            wb.save(output_path)
            return output_path
        except:
            pass

    # 如果存不進去，嘗試加上時間戳記或序號
    name, ext = os.path.splitext(output_path)
    # 第一次嘗試：精簡時間戳
    ts = datetime.now().strftime("%H%M%S")
    new_path = f"{name}_{ts}{ext}"
    
    try:
        wb.save(new_path)
        return new_path
    except:
        # 最終嘗試：不斷嘗試序號直到成功
        for i in range(1, 100):
            final_path = f"{name}_{ts}_{i}{ext}"
            try:
                wb.save(final_path)
                return final_path
            except:
                continue
    return output_path

class ExcelWriter:
    """Excel 匯出核心類別 (座標員)"""
    
    def __init__(self):
        from .excel_fail_list_builder import FailListBuilder
        from .excel_summary_builder import SummaryBuilder
        
        self.fail_builder = FailListBuilder()
        self.summary_builder = SummaryBuilder()

    def _extract_total_secs(self, raw_lines: list):
        """便利方法：調用工具函數提取秒數"""
        from .excel_utils import extract_total_secs
        return extract_total_secs(raw_lines)

    def _extract_isn_from_filename(self, filename: str):
         """便利方法：調用工具函數提取 ISN"""
         from .excel_utils import extract_isn_from_filename
         return extract_isn_from_filename(filename)

    def export_pass_fail_workbooks(self, folder_path: str, pass_logs: list, fail_logs: list, source_path: str | list | None = None):
        """
        輸出活頁簿，並根據來源自動命名
        """
        # 提取前綴關鍵字 (如 4Cam_Calibration)
        prefix = self._determine_prefix(fail_logs + pass_logs, source_path)
        
        fail_saved_path = None
        # 1. 處理 FAIL 活頁簿 (僅當有內容時)
        if fail_logs:
            fail_name = f"{prefix}FAIL匯總.xlsx"
            fail_path = os.path.join(folder_path, fail_name)
            fail_saved_path = self._build_fail_workbook(fail_path, fail_logs)
        
        pass_saved_path = None
        # 2. 處理 PASS 活頁簿 (僅當有內容時)
        if pass_logs:
            pass_name = f"{prefix}PASS匯總.xlsx"
            pass_path = os.path.join(folder_path, pass_name)
            pass_saved_path = self._build_pass_workbook(pass_path, pass_logs, prefix=prefix)
        
        return pass_saved_path, fail_saved_path, None

    def _determine_prefix(self, logs: list, source_path: str | list | None = None) -> str:
        """從 LOG 檔名或來源壓縮檔提取關鍵字 (例如 4Cam_Calibration)"""
        candidate = ""
        from .excel_utils import extract_station_from_filename
        
        # A. 優先檢查來源路徑 (如果是 zip 或 7z)
        if isinstance(source_path, str) and (source_path.lower().endswith('.zip') or source_path.lower().endswith('.7z')):
            candidate = extract_station_from_filename(source_path)
            
        # B. 如果沒結果，檢查 logs 中的第一個
        if (not candidate or candidate == "Unknown") and logs:
            for entry in logs:
                fname = entry.get('file_name', entry.get('filename', ''))
                if fname:
                    candidate = extract_station_from_filename(fname)
                    if candidate and candidate != "Unknown":
                        break
        
        if candidate and candidate != "Unknown":
            # 避免前綴含有不合法字元
            from .excel_utils import sanitize_sheet_title
            safe_prefix = sanitize_sheet_title(candidate)
            return f"{safe_prefix}_"
            
        return ""

    def _build_fail_workbook(self, output_path, logs):
        import openpyxl
        wb = openpyxl.Workbook()
        
        # 0. 先行確定所有工作表名稱，避免 Summary 連結失效
        for entry in logs:
            isn = extract_isn_from_filename(entry.get('file_name', ''))
            sheet_name = unique_sheet_name(wb, isn or "FAIL_Detail")
            entry['sheet_name'] = sheet_name
            # 在 wb 中預佔位置（建立空分頁）
            wb.create_sheet(sheet_name)
        
        # 1. 直接建立統一的 FAIL_LIST (包含 Summary 資訊)
        # 我們將它的建立延後到填寫詳細內容之後，以便取得正確的 error_excel_row
        
        # 2. 建立 FAIL_LIST (在填寫詳細內容之前，先計算錯誤行位置)
        # 註：我們需要先填寫詳細內容，才能得到精確的 error_excel_row
        # 所以這裡先暫存 FAIL_LIST 的建立，等詳細內容填完再建立
        
        # 3. 填寫詳細內容並記錄錯誤行位置
        for entry in logs:
            entry['log_type'] = 'FAIL'  # ⚠️ 確保標註為 FAIL 以觸發預覽框
            sheet_name = entry['sheet_name']
            ws = wb[sheet_name]
            error_row = self._write_detailed_log(ws, entry)
            # 將精確的錯誤行號存回 entry，供 FAIL_LIST 使用
            if error_row:
                entry['error_excel_row'] = error_row
        
        # 現在建立 FAIL_LIST (此時 entry 中已有 error_excel_row)
        self.fail_builder.build_fail_list_sheet(wb, logs)
        
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
            
        return safe_save_workbook(wb, output_path)

    def _build_pass_workbook(self, output_path, logs, prefix=""):
        import openpyxl
        wb = openpyxl.Workbook()
        
        # 0. 先行確定所有工作表名稱
        for entry in logs:
            isn = extract_isn_from_filename(entry.get('file_name', ''))
            sheet_name = unique_sheet_name(wb, isn or "PASS_Detail")
            entry['sheet_name'] = sheet_name
            wb.create_sheet(sheet_name)
            
        # 1. 建立 Summary (傳入前綴關鍵字)
        sws = self.summary_builder.create_summary_sheet(wb, logs, title="Summary", prefix=prefix)
        try:
            sws.sheet_properties.tabColor = '0000FF'
        except:
            pass
        
        # 2. 填寫詳細內容
        for entry in logs:
            entry['log_type'] = 'PASS'  # ⚠️ 標註為 PASS
            sheet_name = entry['sheet_name']
            ws = wb[sheet_name]
            self._write_detailed_log(ws, entry)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return safe_save_workbook(wb, output_path)

    def _write_detailed_log(self, ws, entry):
        """寫入詳細的 LOG 內容與標註 (同步 GUI 外觀並補回超連結)"""
        
    def _write_detailed_log(self, ws, entry):
        """寫入詳細的 LOG 內容與標註 (同步 GUI 外觀並補回超連結)"""
        
        from openpyxl.styles import Font, Alignment, PatternFill
        from .sheet_builder import write_raw_log_with_annotations, insert_header_info

        # 1. 置頂 [回到 Summary/FAIL_LIST] 連結 與 TOP/DOWN - Font 16, 深藍底白字
        link_font = Font(name='Calibri', size=16, bold=True, color="FFFFFF", underline="single")
        deep_blue_fill = PatternFill('solid', fgColor='000080')
        
        # 根據日誌類型決定返回目標
        log_type = entry.get('log_type', 'UNKNOWN')
        back_target = 'FAIL_LIST' if log_type == 'FAIL' else 'Summary'
        
        # 建立整合的導航列
        # A1: [回到 Summary], B1: [TOP], C1: [DOWN]
        c_back = ws.cell(row=1, column=1)
        c_back.value = f'=HYPERLINK("#\'{back_target}\'!A1", "[回到 {back_target}]")'
        c_back.font = link_font
        c_back.fill = deep_blue_fill
        c_back.alignment = Alignment(horizontal='center')
        
        c_top = ws.cell(row=1, column=2)
        c_top.value = f'=HYPERLINK("#\'{ws.title}\'!A1", "[ TOP ]")'
        c_top.font = link_font
        c_top.fill = deep_blue_fill
        c_top.alignment = Alignment(horizontal='center')
        
        # 2. 置頂資訊 (如果沒有預設資訊，則現場生成)
        header_info = entry.get('header_info', '')
        if not header_info:
            header_info = generate_header_info_text(entry)
            
        curr_row = insert_header_info(ws, header_info, start_row=3)
        
        # 3. 原始 LOG (帶有 Premium 背景顏色與文字顏色)
        raw_lines = entry.get('raw_lines', [])
        annotations = entry.get('ui_annotations', [])
        log_type = entry.get('log_type', 'UNKNOWN') 
        
        # 準備匯總項目 (包含 PASS 與 FAIL 的，用於 [比對項目細節] 預覽)
        items_to_display = entry.get('pass_items', []) + entry.get('fail_items', [])
        
        content_font = Font(name='Calibri', size=11)
        
        # ⚠️ 傳入 log_type 和對應的項目列表
        # ⚠️ 接收返回的最後行、錯誤行位置，以及預覽框的起始行
        # write_raw_log_with_annotations 會回傳 (最後行, 錯誤行, 預覽框起始行)
        last_data_row, error_excel_row, detail_preview_row = write_raw_log_with_annotations(ws, curr_row, raw_lines, annotations, content_font, fail_items=items_to_display, log_type=log_type)
        
        # 更新導航列中的 DOWN 連結 (指向最下方)
        c_down = ws.cell(row=1, column=3)
        c_down.value = f'=HYPERLINK("#\'{ws.title}\'!A{last_data_row}", "[ DOWN ]")'
        c_down.font = link_font
        c_down.fill = deep_blue_fill
        c_down.alignment = Alignment(horizontal='center')

        # 4. 置底連動
        bottom_row = last_data_row + 2
        c_bot = ws.cell(row=bottom_row, column=1)
        c_bot.value = f'=HYPERLINK("#\'{back_target}\'!A1", "[回到 {back_target}]")'
        c_bot.font = link_font
        c_bot.fill = deep_blue_fill
        c_bot.alignment = Alignment(horizontal='center')
        
        c_bot_top = ws.cell(row=bottom_row, column=2)
        c_bot_top.value = f'=HYPERLINK("#\'{ws.title}\'!A1", "[ TOP ]")'
        c_bot_top.font = link_font
        c_bot_top.fill = deep_blue_fill
        c_bot_top.alignment = Alignment(horizontal='center')
        
        auto_fit_columns(ws, {1: 130, 2: 20, 3: 20})
        
        # 返回錯誤行位置 (供 FAIL_LIST 超鏈接使用)
        return error_excel_row

    # 向下相容
    def export(self, pass_items, fail_items, output_path):
        pass # 目前主要由 export_pass_fail_workbooks 負責