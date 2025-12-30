import os
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
import glob
from datetime import datetime
import re
import pandas as pd
from openpyxl.utils import get_column_letter
# 延後載入：openpyxl 將在內部方法載入以提升啟動速度
class CSVProcessor:
    def __init__(self, app):
        self.app = app
        self.csv_files = []
        self.analysis_dir = None
        self.max_display_length = 50  # 預設最大顯示長度
        self.selected_files = []  # 使用者選擇的檔案
        
    def select_directory(self):
        """讓使用者選擇包含CSV檔案的目錄"""
        directory = filedialog.askdirectory(title="選擇包含CSV檔案的目錄")
        if directory:
            self.scan_csv_files(directory)
            return True
        return False
    
    def select_files(self):
        """讓使用者直接選擇CSV檔案（支援多選）"""
        file_paths = filedialog.askopenfilenames(
            title="選擇CSV檔案（可多選）",
            filetypes=[
                ("CSV檔案", "*.csv"),
                ("所有檔案", "*.*")
            ]
        )
        if file_paths:
            self.csv_files = list(file_paths)
            self.show_csv_selection_window()
            return True
        return False
    
    def scan_csv_files(self, directory):
        """掃描目錄中的CSV檔案"""
        self.csv_files = []
        patterns = ['*.csv', '*.CSV']
        
        for pattern in patterns:
            files = glob.glob(os.path.join(directory, pattern))
            self.csv_files.extend(files)
        
        # 顯示找到的CSV檔案
        if self.csv_files:
            self.show_csv_selection_window()
        else:
            messagebox.showinfo("提示", "在選擇的目錄中沒有找到CSV檔案")
    
    def show_csv_selection_window(self):
        """顯示CSV檔案選擇視窗（含checkbox）"""
        # 創建新視窗顯示CSV檔案列表
        self.selection_window = tk.Toplevel(self.app.root)
        self.selection_window.title("選擇要處理的CSV檔案")
        self.selection_window.configure(bg='white')
        
        # 🟢 強制置頂與模態設定
        self.selection_window.transient(self.app.root)
        self.selection_window.attributes("-topmost", True)
        self.selection_window.grab_set()
        
        # 居中設定
        win_w, win_h = 750, 550
        try:
            self.selection_window.update_idletasks()
            parent_x = self.app.root.winfo_rootx()
            parent_y = self.app.root.winfo_rooty()
            parent_w = self.app.root.winfo_width()
            parent_h = self.app.root.winfo_height()
            x = parent_x + (parent_w - win_w) // 2
            y = parent_y + (parent_h - win_h) // 2
            self.selection_window.geometry(f"{win_w}x{win_h}+{x}+{y}")
        except:
            self.selection_window.geometry(f"{win_w}x{win_h}")
            
        self.selection_window.resizable(True, True)
        
        # 主框架
        main_frame = tk.Frame(self.selection_window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 標題
        title_label = tk.Label(main_frame, text=f"找到 {len(self.csv_files)} 個CSV檔案", 
                              font=('Microsoft JhengHei', 14, 'bold'), bg='white')
        title_label.pack(pady=(0, 10))
        
        # 說明文字
        info_label = tk.Label(main_frame, 
                             text="請選擇要處理的CSV檔案（至少選擇一個）", 
                             font=('Microsoft JhengHei', 10), fg='#1565C0', bg='white')
        info_label.pack(pady=(0, 10))
        
        # 全選/全不選按鈕框架
        button_frame = tk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Button(button_frame, text=" 全選 ", 
                 command=self.select_all_files,
                 bg='#43A047', fg='white', font=('Microsoft JhengHei', 9)).pack(side=tk.LEFT, padx=5)
        
        tk.Button(button_frame, text=" 全不選 ", 
                 command=self.deselect_all_files,
                 bg='#E53935', fg='white', font=('Microsoft JhengHei', 9)).pack(side=tk.LEFT, padx=5)
        
        # 檔案列表框架
        list_frame = tk.Frame(main_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        # 創建 Treeview with checkboxes (移除檔案大小，加寬檔案名稱)
        columns = ('選擇', '檔案名稱', '修改時間')
        self.tree = ttk.Treeview(list_frame, columns=columns, show='headings')
        
        # 設定欄位寬度
        self.tree.heading('選擇', text='選擇')
        self.tree.heading('檔案名稱', text='檔案名稱')
        self.tree.heading('修改時間', text='修改時間')
        
        self.tree.column('選擇', width=60, anchor='center')
        self.tree.column('檔案名稱', width=450)
        self.tree.column('修改時間', width=150)
        
        # 添加檔案資訊
        self.checkbox_vars = []  # 儲存checkbox變數
        
        for i, file_path in enumerate(self.csv_files):
            file_name = os.path.basename(file_path)
            mod_time = os.path.getmtime(file_path)
            
            # 創建checkbox變數
            var = tk.BooleanVar()
            var.set(i == 0)  # 預設第一個打勾
            self.checkbox_vars.append(var)
            
            # 插入資料 (對應新的欄位數量)
            self.tree.insert('', 'end', values=(
                '☑' if var.get() else '☐',  # checkbox顯示
                file_name,
                datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d %H:%M:%S')
            ))
            
            # 綁定checkbox變更事件
            var.trace('w', lambda *args, idx=i: self.update_checkbox_display(idx))
        
        # 滾動條
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 綁定點擊事件
        self.tree.bind('<Button-1>', self.on_tree_click)
        
        # 底部按鈕框架
        bottom_frame = tk.Frame(main_frame)
        bottom_frame.pack(fill=tk.X, pady=(10, 0))
        
        # 狀態標籤
        self.status_label = tk.Label(bottom_frame, text="已選擇 1 個檔案", 
                                    font=('Microsoft JhengHei', 10), fg='#2E7D32', bg='white')
        self.status_label.pack(side=tk.RIGHT, padx=20)
        
        # 按鈕 (調換順序：開始處理放左邊)
        tk.Button(bottom_frame, text=" 開始處理 ", 
                 command=self.start_processing,
                 bg='#1565C0', fg='white', font=('Microsoft JhengHei', 10, 'bold'),
                 width=12, pady=5, cursor='hand2').pack(side=tk.LEFT, padx=5)
        
        tk.Button(bottom_frame, text=" 取消 ", 
                 command=self.selection_window.destroy,
                 bg='#757575', fg='white', font=('Microsoft JhengHei', 10),
                 width=10, pady=5, cursor='hand2').pack(side=tk.LEFT, padx=5)
        
        # 更新狀態
        self.update_status_label()
    
    def on_tree_click(self, event):
        """處理Treeview點擊事件"""
        item = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        
        if item and column == '#1':  # 點擊選擇欄位
            # 找到對應的檔案索引
            item_index = self.tree.index(item)
            if 0 <= item_index < len(self.checkbox_vars):
                # 切換checkbox狀態
                current_value = self.checkbox_vars[item_index].get()
                self.checkbox_vars[item_index].set(not current_value)
    
    def update_checkbox_display(self, index):
        """更新checkbox顯示"""
        if 0 <= index < len(self.checkbox_vars):
            var = self.checkbox_vars[index]
            items = list(self.tree.get_children())
            if 0 <= index < len(items):
                item = items[index]
                values = list(self.tree.item(item, 'values'))
                values[0] = '☑' if var.get() else '☐'
                self.tree.item(item, values=values)
                self.update_status_label()
    
    def select_all_files(self):
        """全選所有檔案"""
        for var in self.checkbox_vars:
            var.set(True)
    
    def deselect_all_files(self):
        """全不選所有檔案"""
        for var in self.checkbox_vars:
            var.set(False)
    
    def update_status_label(self):
        """更新狀態標籤"""
        selected_count = sum(1 for var in self.checkbox_vars if var.get())
        self.status_label.config(text=f"已選擇 {selected_count} 個檔案")
        
        # 更新開始處理按鈕狀態
        if selected_count == 0:
            self.status_label.config(fg='red')
        else:
            self.status_label.config(fg='green')
    
    def start_processing(self):
        """開始處理選中的CSV檔案"""
        # 檢查是否至少選擇一個檔案
        selected_files = []
        for i, var in enumerate(self.checkbox_vars):
            if var.get():
                selected_files.append(self.csv_files[i])
        
        if not selected_files:
            messagebox.showwarning("警告", "請至少選擇一個CSV檔案進行處理")
            return
        
        self.selected_files = selected_files
        self.selection_window.destroy()
        
        # 創建Analysis_CSV_FILE目錄
        self.create_analysis_directory()
        
        # 複製檔案並處理
        self.copy_and_process_files()
    
    def create_analysis_directory(self):
        """創建Analysis_CSV_FILE目錄"""
        current_dir = os.path.dirname(os.path.abspath(__file__))
        self.analysis_dir = os.path.join(current_dir, "Analysis_CSV_FILE")
        
        if not os.path.exists(self.analysis_dir):
            os.makedirs(self.analysis_dir)
            print(f"已創建目錄: {self.analysis_dir}")
    
    def copy_and_process_files(self):
        """複製檔案並逐一處理"""
        success_count = 0
        fail_count = 0
        
        # 建立進度視窗
        progress_window = self.create_progress_window()
        
        # 同步啟動主畫面的閃爍 LOGO 與狀態列
        if hasattr(self.app, '_show_progress'):
            self.app._show_progress("處理CSV檔案", f"準備處理 {len(self.selected_files)} 個檔案...")
            self.app._progress_set_determinate(len(self.selected_files))
        
        output_files = []
        for i, csv_file in enumerate(self.selected_files):
            try:
                msg = os.path.basename(csv_file)
                # 更新進度視窗
                progress_window.update_progress(i, len(self.selected_files), msg)
                
                # 同步更新主畫面狀態列與進度條
                if hasattr(self.app, '_safe_update_progress'):
                    self.app._safe_update_progress(i, len(self.selected_files), f"正在處理: {msg}")
                
                # 複製檔案
                copied_file = self.copy_file(csv_file)
                
                # 處理CSV檔案
                output_file = self.create_output_filename(csv_file)
                if self.process_csv_file(copied_file, output_file):
                    output_files.append(output_file)
                    success_count += 1
                else:
                    fail_count += 1
                
            except Exception as e:
                print(f"處理檔案 {csv_file} 時發生錯誤: {e}")
                fail_count += 1
        
        # 結束進度
        progress_window.complete()
        if hasattr(self.app, '_close_progress'):
            self.app._close_progress()
        
        # 處理完成後依據結果顯示訊息
        if success_count > 0:
            self.ask_open_files(success_count, fail_count, output_files)
        else:
            messagebox.showerror("失敗", f"所有檔案處理皆失敗 ({fail_count} 個檔案)。請檢查錯誤訊息。")
    
    def ask_open_files(self, success_count, fail_count, output_files):
        """詢問使用者是否要開啟處理後的檔案 - 修正版：強制置頂並美化"""
        win = tk.Toplevel(self.app.root)
        win.title("處理完成")
        win.configure(bg='white')
        
        # 🟢 強制置頂與模態設定
        win.transient(self.app.root)
        win.attributes("-topmost", True)
        win.grab_set()
        
        # 居中視窗
        win_w, win_h = 520, 320
        try:
            win.update_idletasks()
            sw = win.winfo_screenwidth()
            sh = win.winfo_screenheight()
            x = (sw - win_w) // 2
            y = (sh - win_h) // 2
            win.geometry(f"{win_w}x{win_h}+{x}+{y}")
        except:
            win.geometry(f"{win_w}x{win_h}")

        main_frame = tk.Frame(win, bg='white', padx=35, pady=30)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 標題
        tk.Label(main_frame, text="✅ CSV 報表處理完成", 
                 font=('Microsoft JhengHei', 16, 'bold'), 
                 fg='#2E7D32', bg='white').pack(pady=(0, 20))

        # 內容資訊
        msg = f"成功處理: {success_count} 個檔案"
        if fail_count > 0:
            msg += f"\n失敗: {fail_count} 個檔案"
        
        tk.Label(main_frame, text=msg, font=('Microsoft JhengHei', 12), bg='white', justify=tk.LEFT).pack(pady=5)
        tk.Label(main_frame, text="是否要立即開啟生成的 Excel 報表與資料夾？", 
                 font=('Microsoft JhengHei', 11), bg='white', fg='#666').pack(pady=(15, 10))

        # 按鈕區
        btn_frame = tk.Frame(main_frame, bg='white', pady=10)
        btn_frame.pack(side=tk.BOTTOM, fill=tk.X)

        def on_confirm():
            import subprocess
            import platform
            
            # 開啟生成的 Excel 檔案 (最多 5 個)
            for i, f in enumerate(output_files):
                if i >= 5: break 
                try:
                    if platform.system() == "Windows":
                        os.startfile(f)
                    else:
                        subprocess.run(["open", f])
                except:
                    pass

            # 最後開啟目錄
            try:
                if platform.system() == "Windows":
                    subprocess.run(["explorer", self.analysis_dir])
                elif platform.system() == "Darwin":
                    subprocess.run(["open", self.analysis_dir])
                else:
                    subprocess.run(["xdg-open", self.analysis_dir])
            except:
                pass
            win.destroy()

        def on_cancel():
            win.destroy()

        # 使用標準 tk 按鈕並模擬現代外觀
        tk.Button(btn_frame, text=" 否 (僅關閉) ", command=on_cancel, 
                  bg='#E0E0E0', fg='black', font=('Microsoft JhengHei', 10),
                  width=15, pady=8, cursor='hand2', relief=tk.FLAT).pack(side=tk.RIGHT, padx=5)

        tk.Button(btn_frame, text=" 是 (開啟報表) ", command=on_confirm, 
                  bg='#1976D2', fg='white', font=('Microsoft JhengHei', 10, 'bold'),
                  width=18, pady=8, cursor='hand2', relief=tk.FLAT).pack(side=tk.RIGHT, padx=5)
        
        win.protocol("WM_DELETE_WINDOW", on_cancel)
        win.wait_window()
    
    def copy_file(self, source_file):
        """複製檔案到Analysis_CSV_FILE目錄"""
        file_name = os.path.basename(source_file)
        dest_file = os.path.join(self.analysis_dir, file_name)
        shutil.copy2(source_file, dest_file)
        return dest_file
    
    def detect_header_row(self, csv_file, encoding, sep):
        """檢測CSV檔案的列標題行位置和元數據行"""
        try:
            with open(csv_file, 'r', encoding=encoding, errors='ignore') as f:
                lines = []
                for i, line in enumerate(f):
                    if i >= 50:  # 增加檢查範圍到前50行
                        break
                    lines.append(line.strip())
            
            # 常見的列標題關鍵字（增加更多細化關鍵字）
            common_headers = ['serial', 'id', 'time', 'result', 'test', 'project', 'station', 
                            'factory', 'operator', 'cycle', 'diag', 'gui', 'fixture', 'error']
            
            header_row_index = None
            metadata_rows = []
            
            for i, line in enumerate(lines):
                if not line:
                    continue
                
                # 分割行
                parts = line.split(sep)
                line_lower = line.lower()
                
                # 檢查是否包含常見的列標題關鍵字
                # 降低門檻：只要有 2 個關鍵字匹配且欄位數大於 5，或者包含 serial_id 這種強特徵
                matches = sum(1 for header in common_headers if header in line_lower)
                
                if (matches >= 2 and len(parts) >= 5) or ('serial_id' in line_lower):
                    header_row_index = i
                    metadata_rows = lines[:i]  # 前面的行都是元數據
                    print(f"[DEBUG] Found header at row {i}: {line[:50]}...")
                    break
            
            return header_row_index, metadata_rows
        except Exception as e:
            print(f"[DEBUG] detect_header_row error: {e}")
            return None, []
    
    def parse_metadata(self, metadata_rows):
        """解析元數據行，提取有用信息"""
        metadata = {
            'contact': '',
            'error_codes': {},
            'limits': {}
        }
        
        for row in metadata_rows:
            row_lower = row.lower()
            # 提取聯繫信息
            if 'email' in row_lower or 'contact' in row_lower:
                metadata['contact'] = row
            # 提取ErrorCode
            elif 'errorcode' in row_lower:
                metadata['error_codes']['header'] = row
            # 提取Upperlimit/Lowerlimit
            elif 'upperlimit' in row_lower:
                metadata['limits']['upper'] = row
            elif 'lowerlimit' in row_lower:
                metadata['limits']['lower'] = row
        
        return metadata
    
    def process_csv_file(self, csv_file, output_file=None):
        """處理單個CSV檔案 - 改進版，支持元數據行檢測"""
        try:
            df = None
            encoding_used = None
            separator_used = None
            header_row_index = None
            metadata_rows = []
            metadata_info = {}
            original_line_count = 0
            
            # 先計算原始行數
            original_line_count = self.count_csv_lines(csv_file)
            
            # 嘗試多種編碼和分隔符
            encodings = ['utf-8', 'utf-8-sig', 'gbk', 'big5', 'latin-1', 'cp1252', 'iso-8859-1']
            separators = [',', ';', '\t', '|']
            
            for encoding in encodings:
                for sep in separators:
                    try:
                        # 先檢測標題行位置
                        header_row, metadata = self.detect_header_row(csv_file, encoding, sep)
                        if header_row is not None:
                            # 使用檢測到的標題行讀取
                            # 增加 low_memory=False 以處理混合類型，使用 engine='python' 處理異常行
                            df = pd.read_csv(csv_file, encoding=encoding, sep=sep,
                                            skiprows=header_row,
                                            engine='python', on_bad_lines='skip',
                                            keep_default_na=False, na_values=[''],
                                            quoting=0) # 預設引號處理
                            
                            if len(df) > 0:
                                encoding_used = encoding
                                separator_used = sep
                                header_row_index = header_row
                                metadata_rows = metadata
                                metadata_info = self.parse_metadata(metadata)
                                print(f"[OK] 檢測到元數據行（前{header_row}行），使用編碼 {encoding} 和分隔符 '{sep}'")
                                break
                        else:
                            # 如果沒檢測到，嘗試第一行作為標題
                            df = pd.read_csv(csv_file, encoding=encoding, sep=sep, 
                                            engine='python', on_bad_lines='skip',
                                            keep_default_na=False, na_values=[''])
                            if len(df) > 0:
                                encoding_used = encoding
                                separator_used = sep
                                print(f"[OK] 使用編碼 {encoding} 和分隔符 '{sep}' 讀取檔案（無元數據行）")
                                break
                    except Exception as e:
                        continue
                
                if df is not None and len(df) > 0:
                    break
            
            # 如果還是失敗，嘗試自動檢測分隔符 (不指定 sep)
            if df is None or len(df) == 0:
                for encoding in encodings:
                    try:
                        df = pd.read_csv(csv_file, encoding=encoding, sep=None, 
                                        engine='python', on_bad_lines='skip',
                                        keep_default_na=False, na_values=[''])
                        if len(df) > 0:
                            encoding_used = encoding
                            print(f"[OK] 使用編碼 {encoding} (自動檢測分隔符) 讀取檔案")
                            break
                    except:
                        continue
            
            if df is None or len(df) == 0:
                raise Exception("無法讀取CSV檔案，請檢查檔案格式和編碼")
            
            # 驗證資料完整性
            print(f"  原始CSV行數: {original_line_count}, 讀取行數: {len(df)}, 列數: {len(df.columns)}")
            if original_line_count > 0 and len(df) < (original_line_count - 1) * 0.9:
                print(f"  [WARN] 警告：讀取的行數可能少於原始檔案（可能因編碼或格式問題）")
            
            # 將元數據信息附加到DataFrame
            df.attrs['metadata'] = metadata_info
            df.attrs['header_row'] = header_row_index
            df.attrs['metadata_rows'] = metadata_rows
            
            # 數據排序：將 FAIL 放在最上面
            # 建立一個排序參考列
            def get_sort_score(row):
                if self.is_fail_row(row): return 0
                if self.is_pass_row(row): return 1
                return 2
            
            df['_sort_key'] = df.apply(get_sort_score, axis=1)
            df = df.sort_values('_sort_key').drop(columns=['_sort_key'])
            
            # 創建新的Excel檔案
            if output_file is None:
                output_file = self.create_output_filename(csv_file)
            self.create_formatted_excel(df, output_file, csv_file)
            return True
        except PermissionError:
            error_msg = f"無法儲存檔案！請先關閉已開啟的 Excel 報表：\n{os.path.basename(output_file)}"
            print(error_msg)
            messagebox.showerror("檔案被佔用", error_msg)
            return False
        except Exception as e:
            error_msg = f"處理CSV檔案 {os.path.basename(csv_file)} 失敗: {e}"
            print(error_msg)
            messagebox.showerror("錯誤", error_msg)
            return False
    
    def count_csv_lines(self, csv_file):
        """計算CSV檔案的原始行數（用於驗證）"""
        try:
            encodings = ['utf-8', 'utf-8-sig', 'gbk', 'big5', 'latin-1', 'cp1252']
            for encoding in encodings:
                try:
                    with open(csv_file, 'r', encoding=encoding, errors='ignore') as f:
                        return sum(1 for _ in f)
                except:
                    continue
            return 0
        except:
            return 0
    
    def create_output_filename(self, csv_file):
        """創建輸出檔案名稱"""
        base_name = os.path.splitext(os.path.basename(csv_file))[0]
        return os.path.join(self.analysis_dir, f"{base_name}_Analysis_CSV.xlsx")
    
    def create_formatted_excel(self, df, output_file, csv_file):
        """創建格式化的Excel檔案 - 增強版報表"""
        from openpyxl import Workbook
        wb = Workbook()
        
        # 獲取元數據信息
        metadata_info = df.attrs.get('metadata', {})
        metadata_rows = df.attrs.get('metadata_rows', [])
        
        # 創建主要數據工作表（放在第一個）
        ws_data = wb.active
        ws_data.title = "數據明細"
        self.create_data_sheet(ws_data, df, metadata_info)
        
        # 創建統計摘要工作表
        ws_summary = wb.create_sheet("統計摘要", 0)
        self.create_summary_sheet(ws_summary, df, csv_file, metadata_info, metadata_rows)

        # 創建進階分析 (基於建議)
        self.create_analysis_charts_sheet(wb, df, metadata_info)
        self.create_spc_analysis_sheet(wb, df, metadata_info)
        
        # 儲存檔案
        wb.save(output_file)
        print(f"[OK] 已處理並儲存: {os.path.basename(output_file)}")
    
    def create_summary_sheet(self, ws, df, csv_file, metadata_info=None, metadata_rows=None):
        """創建統計摘要工作表 - 改進版，支持元數據和錯誤代碼分類"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        if metadata_info is None:
            metadata_info = {}
        if metadata_rows is None:
            metadata_rows = []
        # 樣式定義 - 統一使用 Calibri (內容 11, 標題 14)
        font_name = 'Calibri'
        title_font = Font(name=font_name, bold=True, size=14, color="FFFFFF")
        header_font = Font(name=font_name, bold=True, size=11, color="FFFFFF")
        normal_font = Font(name=font_name, size=11)
        bold_font = Font(name=font_name, bold=True, size=11)
        
        title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        info_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # --- 新增說明區塊 ---
        ws.merge_cells(f'A{current_row}:D{current_row}')
        desc_cell = ws.cell(row=current_row, column=1, value="【說明】本頁彙總了檔案的整體測試量與良率，並自動分類主要錯誤代碼。")
        desc_cell.font = Font(name=font_name, size=11, italic=True, color="555555")
        current_row += 1
        ws.merge_cells(f'A{current_row}:D{current_row}')
        desc_cell2 = ws.cell(row=current_row, column=1, value="您可以從下方的「錯誤代碼分類」找出最主要的失效原因 (Pareto)。")
        desc_cell2.font = Font(name=font_name, size=11, italic=True, color="555555")
        current_row += 2
        # ------------------

        ws.merge_cells(f'A{current_row}:D{current_row}')
        title_cell = ws.cell(row=current_row, column=1, value="CSV 數據分析報表")
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 2
        
        # 檔案資訊
        ws.cell(row=current_row, column=1, value="檔案名稱:").font = bold_font
        ws.cell(row=current_row, column=2, value=os.path.basename(csv_file)).font = normal_font
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="處理時間:").font = bold_font
        ws.cell(row=current_row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S')).font = normal_font
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="總行數:").font = bold_font
        ws.cell(row=current_row, column=2, value=f"{len(df):,}").font = normal_font
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="總列數:").font = bold_font
        ws.cell(row=current_row, column=2, value=f"{len(df.columns)}").font = normal_font
        current_row += 2
        
        # 尋找時長欄位 (支援多種常見命名)
        time_col = None
        # 優先搜尋明確表示時長的關鍵字，排除可能包含時間戳記的關鍵字
        priority_time_names = ['cycle_time', 'duration', 'time(s)', 'elapsed']
        fallback_time_names = ['test_time', 'total_time', 'time']
        
        for k in priority_time_names:
            for col in df.columns:
                if k in str(col).lower():
                    time_col = col
                    break
            if time_col: break
            
        if not time_col:
            for k in fallback_time_names:
                for col in df.columns:
                    if k in str(col).lower():
                        time_col = col
                        break
                if time_col: break
        
        # 統計數據初始化
        pass_count = 0
        fail_count = 0
        fail_codes = {}  # 錯誤代碼統計 {錯誤代碼: 數量}

        # 一次性遍列數據，計算計數與時長 (優化性能)
        pass_times = []
        fail_times = []
        
        for _, row in df.iterrows():
            is_p = self.is_pass_row(row)
            is_f = self.is_fail_row(row)
            
            # 計數
            if is_p:
                pass_count += 1
            elif is_f:
                fail_count += 1
                # 提取錯誤代碼
                fail_code = self.extract_fail_code(row)
                if fail_code:
                    fail_codes[fail_code] = fail_codes.get(fail_code, 0) + 1
                else:
                    fail_codes['未知錯誤'] = fail_codes.get('未知錯誤', 0) + 1
            
            # 時長採集 (加入防呆，排除大於 10 萬的數值，因為那通常是時間戳記 YYYYMMDD...)
            if time_col and pd.notna(row.get(time_col)):
                try:
                    val = float(row[time_col])
                    if val < 50000: # 排除可能的 Timestamp (如果一個測試超過 50000s = 13小時也太誇張)
                        if is_p: pass_times.append(val)
                        elif is_f: fail_times.append(val)
                except:
                    pass

        # 定義時間轉換函數
        def format_duration(seconds):
            if not isinstance(seconds, (int, float)) or seconds < 0: return "N/A"
            h = int(seconds // 3600)
            m = int((seconds % 3600) // 60)
            s = seconds % 60
            if h > 0: return f"{h}h {m}m {s:.1f}s"
            if m > 0: return f"{m}m {s:.1f}s"
            return f"{s:.2f} s"

        # 預計算平均時長
        pass_avg_time = format_duration(sum(pass_times)/len(pass_times)) if pass_times else "N/A"
        fail_avg_time = format_duration(sum(fail_times)/len(fail_times)) if fail_times else "N/A"

        total_count = len(df)
        other_count = total_count - pass_count - fail_count

        # 統計表格標題
        ws.cell(row=current_row, column=1, value="項目").fill = header_fill
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).border = border
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=current_row, column=2, value="數量").fill = header_fill
        ws.cell(row=current_row, column=2).font = header_font
        ws.cell(row=current_row, column=2).border = border
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=current_row, column=3, value="百分比").fill = header_fill
        ws.cell(row=current_row, column=3).font = header_font
        ws.cell(row=current_row, column=3).border = border
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=current_row, column=4, value="平均測試時長").fill = header_fill
        ws.cell(row=current_row, column=4).font = header_font
        ws.cell(row=current_row, column=4).border = border
        ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # PASS統計
        ws.cell(row=current_row, column=1, value="[PASS]").fill = pass_fill
        ws.cell(row=current_row, column=1).font = bold_font
        ws.cell(row=current_row, column=1).border = border
        ws.cell(row=current_row, column=2, value=pass_count).fill = pass_fill
        ws.cell(row=current_row, column=2).font = normal_font
        ws.cell(row=current_row, column=2).border = border
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
        percentage = (pass_count / total_count * 100) if total_count > 0 else 0
        ws.cell(row=current_row, column=3, value=f"{percentage:.1f}%").fill = pass_fill
        ws.cell(row=current_row, column=3).font = normal_font
        ws.cell(row=current_row, column=3).border = border
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')
        # 平均時長 (PASS)
        ws.cell(row=current_row, column=4, value=pass_avg_time).fill = pass_fill
        ws.cell(row=current_row, column=4).font = normal_font
        ws.cell(row=current_row, column=4).border = border
        ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='center')
        current_row += 1
        
        # FAIL統計
        ws.cell(row=current_row, column=1, value="[FAIL]").fill = fail_fill
        ws.cell(row=current_row, column=1).font = bold_font
        ws.cell(row=current_row, column=1).border = border
        ws.cell(row=current_row, column=2, value=fail_count).fill = fail_fill
        ws.cell(row=current_row, column=2).font = normal_font
        ws.cell(row=current_row, column=2).border = border
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
        percentage = (fail_count / total_count * 100) if total_count > 0 else 0
        ws.cell(row=current_row, column=3, value=f"{percentage:.1f}%").fill = fail_fill
        ws.cell(row=current_row, column=3).font = normal_font
        ws.cell(row=current_row, column=3).border = border
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')
        # 平均時長 (FAIL)
        ws.cell(row=current_row, column=4, value=fail_avg_time).fill = fail_fill
        ws.cell(row=current_row, column=4).font = normal_font
        ws.cell(row=current_row, column=4).border = border
        ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='center')
        current_row += 1
        
        # 其他統計
        if other_count > 0:
            ws.cell(row=current_row, column=1, value="○ 其他").fill = info_fill
            ws.cell(row=current_row, column=1).font = bold_font
            ws.cell(row=current_row, column=1).border = border
            ws.cell(row=current_row, column=2, value=other_count).fill = info_fill
            ws.cell(row=current_row, column=2).font = normal_font
            ws.cell(row=current_row, column=2).border = border
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
            percentage = (other_count / total_count * 100) if total_count > 0 else 0
            ws.cell(row=current_row, column=3, value=f"{percentage:.1f}%").fill = info_fill
            ws.cell(row=current_row, column=3).font = normal_font
            ws.cell(row=current_row, column=3).border = border
            ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')
            current_row += 1
        
        # 總計
        ws.cell(row=current_row, column=1, value="總計").fill = header_fill
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).border = border
        ws.cell(row=current_row, column=2, value=total_count).fill = header_fill
        ws.cell(row=current_row, column=2).font = header_font
        ws.cell(row=current_row, column=2).border = border
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=3, value="100.0%").fill = header_fill
        ws.cell(row=current_row, column=3).font = header_font
        ws.cell(row=current_row, column=3).border = border
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')
        current_row += 2
        
        # 錯誤代碼分類統計（如果有）
        if fail_codes:
            ws.cell(row=current_row, column=1, value="錯誤代碼分類:").font = Font(bold=True, size=11, color="FF0000")
            current_row += 1
            
            # 錯誤代碼表格標題
            ws.cell(row=current_row, column=1, value="錯誤代碼").fill = header_fill
            ws.cell(row=current_row, column=1).font = header_font
            ws.cell(row=current_row, column=1).border = border
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            
            ws.cell(row=current_row, column=2, value="數量").fill = header_fill
            ws.cell(row=current_row, column=2).font = header_font
            ws.cell(row=current_row, column=2).border = border
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
            
            ws.cell(row=current_row, column=3, value="百分比").fill = header_fill
            ws.cell(row=current_row, column=3).font = header_font
            ws.cell(row=current_row, column=3).border = border
            ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # 按數量排序錯誤代碼
            sorted_codes = sorted(fail_codes.items(), key=lambda x: x[1], reverse=True)
            
            for code, count in sorted_codes:
                ws.cell(row=current_row, column=1, value=code).fill = fail_fill
                ws.cell(row=current_row, column=1).font = bold_font
                ws.cell(row=current_row, column=1).border = border
                
                ws.cell(row=current_row, column=2, value=count).fill = fail_fill
                ws.cell(row=current_row, column=2).font = normal_font
                ws.cell(row=current_row, column=2).border = border
                ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
                
                percentage = (count / fail_count * 100) if fail_count > 0 else 0
                ws.cell(row=current_row, column=3, value=f"{percentage:.1f}%").fill = fail_fill
                ws.cell(row=current_row, column=3).font = normal_font
                ws.cell(row=current_row, column=3).border = border
                ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')
                current_row += 1
            
            current_row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
    
    def create_data_sheet(self, ws, df, metadata_info=None):
        """創建數據明細工作表處理 (插入上下限並凍結)"""
        if metadata_info is None: metadata_info = {}
        # 樣式定義 - 統一使用 Calibri (內容 11)
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        font_name = 'Calibri'
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 淺綠色
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 淺紅色
        limit_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # 淺灰色 (上下限用)
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")  # 淺藍色
        header_font = Font(name=font_name, bold=True, size=11, color="FFFFFF")
        normal_font = Font(name=font_name, size=11)
        fail_font = Font(name=font_name, size=11, bold=True)
        limit_font = Font(name=font_name, size=11, italic=True, color="444444")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 寫入標題行（確保正確處理標題）
        headers = []
        for col in df.columns:
            # 清理標題，移除可能的特殊字符和空白
            header_str = str(col).strip()
            # 確保標題不為空
            if not header_str:
                header_str = f"欄位{len(headers)+1}"
            headers.append(header_str)
        
        # 寫入標題行
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # 寫入 UpperLimit (Row 2) 與 LowerLimit (Row 3)
        upper_row = metadata_info.get('limits', {}).get('upper', '').split(',')
        lower_row = metadata_info.get('limits', {}).get('lower', '').split(',')
        
        for row_idx, limit_data in enumerate([upper_row, lower_row], 2):
            for col_num in range(1, len(headers) + 1):
                val = ""
                if col_num - 1 < len(limit_data):
                    val = limit_data[col_num - 1].strip()
                
                cell = ws.cell(row=row_idx, column=col_num, value=val)
                cell.fill = limit_fill
                cell.font = limit_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 寫入資料並設定樣式 (從 Row 4 開始)
        for row_num, (_, row) in enumerate(df.iterrows(), 4):
            row_is_pass = self.is_pass_row(row)
            row_is_fail = self.is_fail_row(row)
            fail_code = None
            
            if row_is_fail:
                fail_code = self.extract_fail_code(row)
            
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                # 處理內容（數值四捨五入到三位，文字保留）
                if pd.notna(value) and str(value).strip():
                    try:
                        # 嘗試轉換為數值進行四捨五入
                        f_val = float(value)
                        # 如果是整數，則直接使用 int，否則 round 到三位
                        if f_val == int(f_val):
                            cell.value = int(f_val)
                        else:
                            cell.value = round(f_val, 3)
                    except:
                        cell.value = str(value)
                else:
                    cell.value = ""
                
                cell.border = border
                
                # 根據內容判斷PASS/FAIL並設定顏色和字體
                if row_is_pass:
                    cell.fill = pass_fill
                    cell.font = normal_font
                elif row_is_fail:
                    cell.fill = fail_fill
                    # FAIL行使用粗體字體以突出顯示
                    cell.font = fail_font
                else:
                    cell.font = normal_font
                
                # 設定文字對齊（數字右對齊，文字左對齊）
                try:
                    float(str(value))
                    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                except (ValueError, TypeError):
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 自動調整欄寬（優化版本）
        self.auto_adjust_column_width_optimized(ws)
        
        # 凍結視窗 (凍結前3行：標題 + 上限 + 下限)
        ws.freeze_panes = 'A4'
        
        # 設定行高 (優化)
        ws.row_dimensions[1].height = 50  # 增加表頭高度以利換行 (從 35 提升到 50)
        ws.row_dimensions[2].height = 20  # UpperLimit
        ws.row_dimensions[3].height = 20  # LowerLimit
        for row_idx in range(4, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 18
        
        # 啟用自動篩選（確保標題行正確）
        if len(df) > 0 and len(headers) > 0:
            last_col = get_column_letter(len(headers))
            last_row = len(df) + 1
            ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    
    def is_pass_row(self, row):
        """判斷是否為PASS行 - 改進版，更精確的判斷"""
        pass_keywords = ['PASS', 'OK', 'SUCCESS', 'TRUE', 'PASSED', 'YES', 'Y']
        pass_phrases = ['TEST PASS', 'TEST OK', 'RESULT PASS', 'STATUS PASS', 'PASSED TEST', 'IS PASS']
        
        row_str = ' '.join([str(v) for v in row if pd.notna(v)]).upper()
        
        # 優先檢查完整短語（避免誤判）
        for phrase in pass_phrases:
            if phrase in row_str:
                return True
        
        # 檢查關鍵字（但排除包含FAIL的情況）
        if 'FAIL' in row_str:
            return False
        
        for value in row:
            if pd.notna(value):
                value_str = str(value).upper().strip()
                # 精確匹配關鍵字（避免部分匹配）
                for keyword in pass_keywords:
                    # 完整單詞匹配或獨立存在
                    if (keyword == value_str or 
                        f' {keyword} ' in f' {value_str} ' or
                        value_str.startswith(keyword + ' ') or
                        value_str.endswith(' ' + keyword)):
                        return True
        
        return False
    
    def is_fail_row(self, row):
        """判斷是否為FAIL行 - 改進版，支持FAIL:錯誤代碼格式"""
        # 優先檢查test_result列（如果存在）
        if hasattr(row, 'index') and 'test_result' in row.index:
            result = str(row['test_result']).strip()
            result_upper = result.upper()
            # 檢查FAIL:格式（如FAIL:BSFI15, FAIL:BSFE01）
            if result_upper.startswith('FAIL'):
                return True
        
        fail_keywords = ['FAIL', 'ERROR', 'NACK', 'TIMEOUT', 'FALSE', 'NO', 'N', 'ABORT', 'ABORTED']
        fail_phrases = ['TEST FAIL', 'TEST ERROR', 'RESULT FAIL', 'STATUS FAIL', 'FAILED TEST', 
                       'IS FAIL', 'EXECUTES FAIL', "DOESN'T MATCH", 'TEST ABORTED']
        
        row_str = ' '.join([str(v) for v in row if pd.notna(v)]).upper()
        
        # 優先檢查完整短語（避免誤判）
        for phrase in fail_phrases:
            if phrase in row_str:
                return True
        
        # 檢查FAIL:格式（支持錯誤代碼）
        if 'FAIL:' in row_str:
            return True
        
        # 檢查關鍵字
        for value in row:
            if pd.notna(value):
                value_str = str(value).upper().strip()
                # 精確匹配關鍵字
                for keyword in fail_keywords:
                    # 完整單詞匹配或獨立存在
                    if (keyword == value_str or 
                        f' {keyword} ' in f' {value_str} ' or
                        value_str.startswith(keyword + ' ') or
                        value_str.endswith(' ' + keyword)):
                        # 排除包含PASS的情況（避免誤判）
                        if 'PASS' not in value_str:
                            return True
        
        # 檢查數值狀態碼（常見的錯誤碼）
        for value in row:
            if pd.notna(value):
                try:
                    num_val = float(str(value))
                    # 如果數值為負數，可能是錯誤狀態
                    if num_val < 0:
                        return True
                except (ValueError, TypeError):
                    pass
        
        return False
    
    def extract_fail_code(self, row):
        """從行中提取FAIL錯誤代碼（如BSFI15, BSFE01）"""
        # 優先檢查test_result列
        if hasattr(row, 'index') and 'test_result' in row.index:
            result = str(row['test_result']).strip()
            if 'FAIL:' in result.upper():
                # 提取FAIL:後面的錯誤代碼
                parts = result.split(':')
                if len(parts) > 1:
                    return parts[1].strip()
        
        # 檢查整行
        row_str = ' '.join([str(v) for v in row if pd.notna(v)])
        if 'FAIL:' in row_str.upper():
            import re
            match = re.search(r'FAIL:\s*([A-Z0-9]+)', row_str, re.IGNORECASE)
            if match:
                return match.group(1)
        
        return None
    
    def auto_adjust_column_width_optimized(self, ws):
        """優化的自動調整欄寬功能：智慧處理長標題換行與數據寬度平衡"""
        import re
        for column in ws.columns:
            column_letter = get_column_letter(column[0].column)
            
            # 分離標題與數據
            header_cell = column[0]
            # 只檢查前 50 行數據以平衡性能與準確度
            data_cells = list(column[1:51]) if len(column) > 51 else list(column[1:])
            
            # --- 1. 計算數據的最大理想寬度 ---
            max_data_width = 0
            for cell in data_cells:
                try:
                    if cell.value:
                        val_str = str(cell.value)
                        text_length = len(val_str)
                        chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', val_str))
                        # 基礎計算：字元數 + 中文補償
                        width = text_length + chinese_chars * 0.8
                        if width > max_data_width:
                            max_data_width = width
                except:
                    pass
            
            # --- 2. 計算標題的理想寬度 ---
            header_val = str(header_cell.value) if header_cell.value else ""
            header_text_len = len(header_val)
            header_chinese = len(re.findall(r'[\u4e00-\u9fff]', header_val))
            header_width = header_text_len + header_chinese * 0.8
            
            # --- 3. 決策邏輯：平衡寬度與換行 ---
            # 如果標題很長但數據很短，我們強制標題換行，而不是把欄位拉到無限寬
            if header_width > 20 and max_data_width < 10:
                # 這種情況下，寬度設定在 15-18 左右，讓標題自動包裝成 2-3 行
                final_width = 16
            else:
                # 正常的取數據與標題的較大值，但標題部分的影響力加權降低
                # 確保標題至少能看到大部分關鍵字
                suggested_width = max(max_data_width + 3, min(header_width, 25))
                final_width = suggested_width
            
            # --- 4. 套用限制與最小寬度 ---
            final_width = max(final_width, 10)  # 最小寬度 10
            final_width = min(final_width, 50)  # 最大寬度 50
            
            ws.column_dimensions[column_letter].width = final_width
    
    def create_analysis_charts_sheet(self, wb, df, metadata_info):
        """創建分析圖表工作表（柏拉圖和趨勢圖）"""
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.styles import Font, PatternFill, Alignment
        
        ws = wb.create_sheet("分析圖表")
        font_name = 'Calibri'
        
        # --- 新增說明區塊 --- (字體調整為 12)
        ws.cell(row=1, column=1, value="【說明】柏拉圖 (Pareto Chart) 顯示前幾大失效來源；趨勢圖追蹤關鍵測量值的波動。").font = Font(name=font_name, size=12, italic=True)
        ws.cell(row=2, column=1, value="若趨勢圖呈現劇烈上下震盪，表示測試不穩定或產品一致性較差。").font = Font(name=font_name, size=12, italic=True)
        # ------------------
        
        start_data_row = 4
        # 1. 錯誤分佈數據 (Pareto)
        fail_codes = {}
        for _, row in df.iterrows():
            if self.is_fail_row(row):
                code = self.extract_fail_code(row) or "未知錯誤"
                fail_codes[code] = fail_codes.get(code, 0) + 1
        
        if fail_codes:
            ws.cell(row=start_data_row+1, column=1, value="錯誤代碼分佈數據").font = Font(name=font_name, bold=True, size=14)
            ws.cell(row=start_data_row+2, column=1, value="錯誤代碼")
            ws.cell(row=start_data_row+2, column=2, value="數量")
            
            sorted_codes = sorted(fail_codes.items(), key=lambda x: x[1], reverse=True)
            for i, (code, count) in enumerate(sorted_codes, start_data_row+3):
                ws.cell(row=i, column=1, value=code)
                ws.cell(row=i, column=2, value=count)
            
            # 創建直方圖 (字體調整為 13)
            chart = BarChart()
            chart.title = "錯誤分佈 (Pareto Chart)"
            
            # 設定圖表文字字體大小 (13pt) - 修正相容性問題
            try:
                from openpyxl.drawing.text import CharacterProperties, Paragraph, ParagraphProperties, RichTextProperties
                cp13 = CharacterProperties(sz=1300) # 13pt = 1300
                # 修正：部分版本不支援 p 參數，改用更安全的方式
                if not hasattr(chart.title, 'txPr') or chart.title.txPr is None:
                    chart.title.txPr = RichTextProperties()
                chart.title.txPr.p = [Paragraph(pPr=ParagraphProperties(defRPr=cp13), endParaRPr=cp13)]
            except Exception as e:
                print(f"[DEBUG] Chart font setting failed: {e}")
            
            # ... 下方代碼 ...
            data = Reference(ws, min_col=2, min_row=start_data_row+2, max_row=start_data_row+2+len(fail_codes))
            cats = Reference(ws, min_col=1, min_row=start_data_row+3, max_row=start_data_row+2+len(fail_codes))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.legend = None
            ws.add_chart(chart, "F4") # 往右移一點，避免蓋到數據表格
            
        # 2. 趨勢圖 (關鍵數值分析)
        # 尋找包含 PixelsShift, Brightness, discontinue 的數值欄位
        numeric_cols = []
        for col in df.columns:
            if any(key in str(col).lower() for key in ['pixelsshift', 'brightness', 'discontinue']):
                # 檢查是否真的含有數值
                try:
                    pd.to_numeric(df[col], errors='raise')
                    numeric_cols.append(col)
                except:
                    continue
        
        if numeric_cols and len(df) > 1:
            start_row = 30
            # 輔助標題
            title = "關鍵指標趨勢分析 (Key Trend Analysis)"
            ws.cell(row=start_row, column=1, value=title).font = Font(name=font_name, bold=True, size=15, color="1565C0")
            
            # 詳細解釋 (字體 12，斜體)
            explanation = "【說明】此區域自動篩選出前 3 個含有 PixelsShift / Brightness / discontinue 關鍵字的數值欄位，"
            ws.cell(row=start_row+1, column=1, value=explanation).font = Font(name=font_name, size=12, italic=True)
            explanation2 = "並依測試序號繪製波動圖。這能幫助您快速識別各個 Camera 通道或感測器數值是否出現異常抖動或規律性偏差。"
            ws.cell(row=start_row+2, column=1, value=explanation2).font = Font(name=font_name, size=12, italic=True)
            
            # 為了避免 Excel 太大，我們只在分析表放前3個關鍵數值的趨勢
            display_cols = numeric_cols[:3]
            start_table_row = start_row + 4
            ws.cell(row=start_table_row-1, column=1, value="分析對象 (Top 3 Items):").font = Font(name=font_name, bold=True)
            for c_idx, col_name in enumerate(display_cols, 1):
                ws.cell(row=start_row+1, column=c_idx, value=str(col_name))
                # 數據已經在 "數據明細" 表，我們可以引用或者複製一部分
                # 這裡為了方便繪圖，簡化處理
            
            # 繪製趨勢圖 (引用 數據明細 表)
            ws_data = wb["數據明細"]
            data_max_row = min(len(df) + 1, 1000) # 限制繪圖點數
            
            for c_idx, col_name in enumerate(display_cols):
                # 找到該列在數據明細中的索引
                df_col_idx = list(df.columns).index(col_name) + 1
                
                lchart = LineChart()
                lchart.title = f"趨勢圖: {col_name}"
                lchart.style = 13
                
                # 設定字體 13pt - 修正相容性問題
                try:
                    cp13 = CharacterProperties(sz=1300)
                    if not hasattr(lchart.title, 'txPr') or lchart.title.txPr is None:
                        lchart.title.txPr = RichTextProperties()
                    lchart.title.txPr.p = [Paragraph(pPr=ParagraphProperties(defRPr=cp13), endParaRPr=cp13)]
                except Exception as e:
                    print(f"[DEBUG] LineChart font setting failed: {e}")
                
                lchart.y_axis.title = "測量值"
                lchart.x_axis.title = "測試序號"
                
                data = Reference(ws_data, min_col=df_col_idx, min_row=1, max_row=data_max_row)
                lchart.add_data(data, titles_from_data=True)
                ws.add_chart(lchart, f"F{start_row + c_idx*15}") # 往右移避免蓋到文字

        # 統一設定字體 (內容大小 11)
        for row in ws.rows:
            for cell in row:
                if cell.font.size != 14: # 不覆蓋標題大小
                    cell.font = Font(name=font_name, size=11)

    def create_spc_analysis_sheet(self, wb, df, metadata_info):
        """創建 SPC & Cpk 分析工作表"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        ws = wb.create_sheet("SPC品質分析")
        font_name = 'Calibri'
        
        # --- 新增說明區塊 ---
        ws.merge_cells('A1:J1')
        ws.cell(row=1, column=1, value="【說明】本表提供測量數據的統計特性。Mean 是平均值，Std 是標準差。").font = Font(name=font_name, size=11, italic=True)
        ws.merge_cells('A2:J2')
        ws.cell(row=2, column=1, value="Cpk 模型評估製程穩健度：Cpk < 1.33 時(黃色標示) 表示生產波動大或容易超出範圍，建議調教測試機。").font = Font(name=font_name, size=11, italic=True)
        # ------------------
        
        headers = ["測量項目", "樣本數", "平均值 (Mean)", "標準差 (Std)", "最大值", "最小值", "LSL", "USL", "Cp", "Cpk"]
        
        # 樣式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name=font_name, bold=True, size=11, color="FFFFFF")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=c, value=h) # 從第 4 行開始
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            
        # 獲取 Limit 信息
        limits_raw = metadata_info.get('limits', {})
        upper_limit_row = limits_raw.get('upper', '').split(',') if limits_raw.get('upper') else []
        lower_limit_row = limits_raw.get('lower', '').split(',') if limits_raw.get('lower') else []
        
        current_row = 5
        for col_idx, col in enumerate(df.columns):
            # 針對數值型且非 Metadata 的列 (包含 PixelsShift, Brightness, discontinue, cycle_time)
            col_search = str(col).lower()
            if any(key in col_search for key in ['pixelsshift', 'brightness', 'discontinue', 'cycle_time']):
                try:
                    data = pd.to_numeric(df[col], errors='coerce').dropna()
                    if len(data) < 2: continue
                    
                    mean = data.mean()
                    std = data.std()
                    max_v = data.max()
                    min_v = data.min()
                    count = len(data)
                    
                    # 嘗試從 Metadata 獲取 Limit (假設索引一致)
                    # 這裡需要找到標題在 CSV 中的實際索引
                    # 目前簡化：如果 metadata_rows 的長度匹配 columns
                    lsl = None
                    usl = None
                    try:
                        # 補償 metadata 行前面的 header_row 偏差
                        # 實際上應該要有更好的對齊邏輯，這裡先嘗試匹配
                        if len(upper_limit_row) > col_idx:
                            val = upper_limit_row[col_idx].strip()
                            if val and val != 'Upperlimit' and val != 'NULL': usl = float(val)
                        if len(lower_limit_row) > col_idx:
                            val = lower_limit_row[col_idx].strip()
                            if val and val != 'Lowerlimit' and val != 'NULL': lsl = float(val)
                    except: pass
                    
                    cp = ""
                    cpk = ""
                    if std > 0:
                        if lsl is not None and usl is not None:
                            cp = (usl - lsl) / (6 * std)
                            cpk = min((usl - mean) / (3 * std), (mean - lsl) / (3 * std))
                        elif usl is not None:
                            cpk = (usl - mean) / (3 * std)
                        elif lsl is not None:
                            cpk = (mean - lsl) / (3 * std)
                    
                    # 寫入資料
                    vals = [str(col), count, round(mean, 4), round(std, 4), max_v, min_v, lsl, usl, 
                            round(cp, 3) if isinstance(cp, float) else cp, 
                            round(cpk, 3) if isinstance(cpk, float) else cpk]
                    
                    for c, v in enumerate(vals, 1):
                        cell = ws.cell(row=current_row, column=c, value=v)
                        cell.border = border
                        cell.font = Font(name=font_name, size=11)
                    
                    # Cpk 著色 (Cpk < 1.33 警告)
                    if isinstance(cpk, float) and cpk < 1.33:
                        ws.cell(row=current_row, column=10).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    
                    current_row += 1
                except: continue
        
        # --- 新增 CPK 圖表 ---
        if current_row > 5: # 表示有數據
            from openpyxl.chart import BarChart, Reference
            chart = BarChart()
            chart.type = "col"
            chart.title = "Cpk 品質指標分佈"
            chart.y_axis.title = "Cpk 數值"
            chart.x_axis.title = "測量項目"
            
            # 設定圖表文字字體大小 (13pt)
            try:
                from openpyxl.drawing.text import CharacterProperties, Paragraph, ParagraphProperties, RichTextProperties
                cp13 = CharacterProperties(sz=1300)
                chart.title.txPr = RichTextProperties(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp13), endParaRPr=cp13)])
            except: pass
            
            data = Reference(ws, min_col=11, min_row=4, max_row=current_row-1) # 第11欄是原代碼邏輯中寫入 Cpk 的位置?
            # 不對，我在 vals 裡 Cpk 是第 11 個元素 (索引 10)，所以對應 Excel 第 11 欄 (K欄)
            # 檢查 vals: [str(col), count, mean, std, max, min, lsl, usl, cp, cpk]
            # 1:項目, 2:樣本數, 3:Mean, 4:Std, 5:Max, 6:Min, 7:LSL, 8:USL, 9:Cp, 10:Cpk
            # 所以 Cpk 在第 10 欄 (J 欄)
            
            data = Reference(ws, min_col=10, min_row=4, max_row=current_row-1)
            cats = Reference(ws, min_col=1, min_row=5, max_row=current_row-1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.legend = None
            
            # 放置圖表
            ws.add_chart(chart, "L4") # 放在表格右側
        # ------------------
        
        # 調整欄寬
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 15
        ws.column_dimensions['A'].width = 35

    def create_progress_window(self):
        """創建進度視窗"""
        return ProgressWindow(self.app.root, "處理CSV檔案")

class ProgressWindow:
    def __init__(self, parent, title):
        self.window = tk.Toplevel(parent)
        self.window.title(title)
        self.window.configure(bg='white')
        self.window.geometry("420x150")
        
        # 🟢 強制置頂與居中
        self.window.transient(parent)
        self.window.attributes("-topmost", True)
        self.window.grab_set()
        
        try:
            self.window.update_idletasks()
            sw = self.window.winfo_screenwidth()
            sh = self.window.winfo_screenheight()
            x = (sw - 420) // 2
            y = (sh - 150) // 2
            self.window.geometry(f"420x150+{(sw-420)//2}+{(sh-150)//2}")
        except:
            pass
            
        self.window.resizable(False, False)
        
        # 進度條
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.window, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(pady=20, padx=20, fill=tk.X)
        
        # 狀態標籤
        self.status_label = tk.Label(self.window, text="準備中...", font=('Calibri', 10))
        self.status_label.pack(pady=10)
        
        # 檔案名稱標籤
        self.file_label = tk.Label(self.window, text="", font=('Calibri', 9), fg='gray')
        self.file_label.pack()
        
        self.window.update()
    
    def update_progress(self, current, total, filename):
        """更新進度"""
        progress = (current / total) * 100
        self.progress_var.set(progress)
        self.status_label.config(text=f"處理中... ({current}/{total})")
        self.file_label.config(text=f"當前檔案: {filename}")
        self.window.update()
    
    def complete(self):
        """完成處理"""
        self.progress_var.set(100)
        self.status_label.config(text="處理完成！")
        self.file_label.config(text="所有選中的CSV檔案已處理完成")
        
        # 3秒後自動關閉
        self.window.after(3000, self.window.destroy)

