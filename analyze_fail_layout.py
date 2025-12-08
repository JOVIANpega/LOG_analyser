# -*- coding: utf-8 -*-
import openpyxl
import sys

# 設定輸出編碼為 UTF-8
sys.stdout.reconfigure(encoding='utf-8')

file_path = r'MINE\result\FAIL匯總.xlsx'

print(f"正在分析 Excel 佈局: {file_path}")

wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"工作表名稱: {ws.title}")
print(f"最大行數: {ws.max_row}")
print(f"最大列數: {ws.max_column}")

print("\n" + "=" * 80)
print("前 20 行儲存格內容分析")
print("=" * 80)

for row in range(1, min(21, ws.max_row + 1)):
    row_content = []
    for col in range(1, min(6, ws.max_column + 1)): # 只看前 5 列
        cell = ws.cell(row=row, column=col)
        val = cell.value
        if val:
            # 簡化顯示，過長的文字截斷
            val_str = str(val).replace('\n', '\\n')
            if len(val_str) > 50:
                val_str = val_str[:50] + "..."
            row_content.append(f"[{col}]: {val_str}")
    
    if row_content:
        print(f"Row {row}: {', '.join(row_content)}")
    else:
        print(f"Row {row}: (空)")

print("\n" + "=" * 80)
print("合併儲存格分析")
print("=" * 80)
if ws.merged_cells.ranges:
    print(f"發現 {len(ws.merged_cells.ranges)} 個合併儲存格區域")
    for i, merge_range in enumerate(ws.merged_cells.ranges):
        if i < 5: # 只顯示前 5 個
            print(f"  - {merge_range}")
else:
    print("無合併儲存格")
