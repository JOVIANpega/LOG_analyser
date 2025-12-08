# -*- coding: utf-8 -*-
import openpyxl
import sys

# 設定輸出編碼為 UTF-8
sys.stdout.reconfigure(encoding='utf-8')

# 讀取 Excel 檔案
wb = openpyxl.load_workbook('Analysis_CSV_FILE/VALO360_AOCI_4Cam_Stitching_test1_log_Analysis_CSV.xlsx')

print("=" * 100)
print("詳細分析 - 數據明細工作表")
print("=" * 100)

ws = wb['📋 數據明細']

# 讀取第一行的實際標題
print("\n第一行 (標題行) 的所有欄位:")
print("-" * 100)
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(1, col).value
    print(f"欄位 {col}: {cell_value}")

# 讀取前 30 行的完整資料
print("\n\n前 30 行的完整資料:")
print("=" * 100)

for row in range(1, min(31, ws.max_row + 1)):
    print(f"\n--- 第 {row} 行 ---")
    row_data = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row, col).value
        if cell_value is not None:
            row_data.append(f"Col{col}: {cell_value}")
    
    if row_data:
        for data in row_data:
            print(f"  {data}")
    else:
        print("  (空白行)")

# 檢查是否有 Send, Reply, Ref, Validation 等欄位
print("\n\n" + "=" * 100)
print("尋找測試腳本相關欄位 (Send, Reply, Ref, Validation):")
print("=" * 100)

# 檢查前 10 行中是否有這些關鍵字
keywords = ['Send', 'Reply', 'Ref', 'Validation', 'send', 'reply', 'ref', 'validation']
found_keywords = {}

for row in range(1, min(50, ws.max_row + 1)):
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row, col).value
        if cell_value:
            cell_str = str(cell_value)
            for keyword in keywords:
                if keyword in cell_str:
                    if keyword not in found_keywords:
                        found_keywords[keyword] = []
                    found_keywords[keyword].append(f"Row {row}, Col {col}: {cell_str[:100]}")

if found_keywords:
    for keyword, locations in found_keywords.items():
        print(f"\n找到關鍵字 '{keyword}':")
        for loc in locations[:5]:  # 只顯示前 5 個
            print(f"  {loc}")
else:
    print("\n未找到測試腳本相關欄位")

print("\n" + "=" * 100)
print("分析完成")
print("=" * 100)
