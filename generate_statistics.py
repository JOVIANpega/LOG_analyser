# -*- coding: utf-8 -*-
import openpyxl
import sys

# 設定輸出編碼為 UTF-8
sys.stdout.reconfigure(encoding='utf-8')

# 讀取測試腳本 Excel 檔案
file_path = 'MINE/VALO360_TestFlow_AOCI_4Cam_Stitching_test1_FATP_250924B.xlsx'
wb = openpyxl.load_workbook(file_path)

# 分析測試腳本工作表
sheet_name = '4cam stitching1 test'
ws = wb[sheet_name]

# 讀取標題列
headers = []
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    headers.append(header if header else f"Col_{col}")

# 尋找關鍵欄位
key_columns = {}
for i, h in enumerate(headers, 1):
    if h in ['Send', 'Reply', 'Ref', 'Validation']:
        key_columns[h] = i

# 統計分析
total_rows = ws.max_row - 1  # 扣除標題行

ref_count = 0
validation_count = 0
both_count = 0
neither_count = 0

ref_col = key_columns.get('Ref')
val_col = key_columns.get('Validation')

for row in range(2, ws.max_row + 1):
    has_ref = False
    has_val = False
    
    if ref_col:
        ref_value = ws.cell(row, ref_col).value
        if ref_value and str(ref_value).strip():
            has_ref = True
            ref_count += 1
    
    if val_col:
        val_value = ws.cell(row, val_col).value
        if val_value and str(val_value).strip():
            has_val = True
            validation_count += 1
    
    if has_ref and has_val:
        both_count += 1
    elif not has_ref and not has_val:
        neither_count += 1

# 計算只有 Ref 但沒有 Validation 的行數
only_ref = ref_count - both_count

print("=" * 100)
print("Excel 測試腳本統計報告")
print("=" * 100)

print(f"\n📊 基本統計:")
print(f"  總測試步驟: {total_rows:,} 行")
print(f"  總欄位數: {ws.max_column} 列")

print(f"\n📈 Ref 欄位統計:")
print(f"  有使用 Ref: {ref_count:,} 行 ({ref_count/total_rows*100:.1f}%)")
print(f"  未使用 Ref: {total_rows - ref_count:,} 行 ({(total_rows - ref_count)/total_rows*100:.1f}%)")

print(f"\n📉 Validation 欄位統計:")
print(f"  有使用 Validation: {validation_count:,} 行 ({validation_count/total_rows*100:.1f}%)")
print(f"  未使用 Validation: {total_rows - validation_count:,} 行 ({(total_rows - validation_count)/total_rows*100:.1f}%)")

print(f"\n🔍 交叉分析:")
print(f"  同時有 Ref 和 Validation: {both_count:,} 行 ({both_count/total_rows*100:.1f}%)")
print(f"  只有 Ref 沒有 Validation: {only_ref:,} 行 ({only_ref/total_rows*100:.1f}%) ⚠️")
print(f"  兩者都沒有: {neither_count:,} 行 ({neither_count/total_rows*100:.1f}%)")

print(f"\n⚠️  關鍵發現:")
print(f"  有 {only_ref:,} 行 ({only_ref/ref_count*100:.1f}% 的 Ref) 提取了變數但沒有驗證!")
print(f"  這代表 {only_ref/ref_count*100:.1f}% 的數據提取缺少品質檢查")

print(f"\n💡 改進潛力:")
if only_ref > 0:
    print(f"  如果為這 {only_ref:,} 行添加 Validation:")
    new_val_count = validation_count + only_ref
    print(f"  - Validation 使用率將從 {validation_count/total_rows*100:.1f}% 提升到 {new_val_count/total_rows*100:.1f}%")
    print(f"  - 提升幅度: {(new_val_count - validation_count)/total_rows*100:.1f} 個百分點")
    print(f"  - 測試覆蓋率提升: {(new_val_count - validation_count)/validation_count*100:.0f}%")

# 視覺化統計
print(f"\n📊 視覺化統計:")
print("=" * 100)

def print_bar(label, count, total, width=50):
    percentage = count / total * 100
    filled = int(width * count / total)
    bar = "█" * filled + "░" * (width - filled)
    print(f"  {label:30s} [{bar}] {count:5,} ({percentage:5.1f}%)")

print("\nRef 欄位使用情況:")
print_bar("有使用 Ref", ref_count, total_rows)
print_bar("未使用 Ref", total_rows - ref_count, total_rows)

print("\nValidation 欄位使用情況:")
print_bar("有使用 Validation", validation_count, total_rows)
print_bar("未使用 Validation", total_rows - validation_count, total_rows)

print("\n交叉分析:")
print_bar("Ref + Validation", both_count, total_rows)
print_bar("只有 Ref", only_ref, total_rows)
print_bar("兩者都沒有", neither_count, total_rows)

print("\n" + "=" * 100)
print("報告完成")
print("=" * 100)
