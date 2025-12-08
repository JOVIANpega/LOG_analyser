# -*- coding: utf-8 -*-
import openpyxl
import sys

# 設定輸出編碼為 UTF-8
sys.stdout.reconfigure(encoding='utf-8')

# 讀取測試腳本 Excel 檔案
file_path = 'MINE/VALO360_TestFlow_AOCI_4Cam_Stitching_test1_FATP_250924B.xlsx'
wb = openpyxl.load_workbook(file_path)

# 分析測試腳本工作表
sheet_name = '4cam stitching1 test'
ws = wb[sheet_name]

print("=" * 120)
print(f"測試腳本工作表分析: {sheet_name}")
print("=" * 120)

print(f"\n總行數: {ws.max_row}")
print(f"總列數: {ws.max_column}")

# 讀取標題列
headers = []
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    headers.append(header if header else f"Col_{col}")

print(f"\n所有欄位標題 ({len(headers)} 個):")
for i, h in enumerate(headers, 1):
    print(f"  {i:2d}. {h}")

# 尋找關鍵欄位
key_columns = {}
for i, h in enumerate(headers, 1):
    if h in ['Send', 'Reply', 'Ref', 'Validation', 'Expect', 'send', 'reply', 'ref', 'validation', 'expect']:
        key_columns[h] = i

print(f"\n找到的關鍵欄位:")
for col_name, col_idx in key_columns.items():
    print(f"  {col_name}: 第 {col_idx} 列")

# 顯示前 50 行資料
print(f"\n\n前 50 行資料預覽:")
print("=" * 120)

for row_idx in range(1, min(51, ws.max_row + 1)):
    # 讀取整行資料
    row_data = {}
    has_content = False
    
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row_idx, col_idx).value
        if cell_value is not None and str(cell_value).strip():
            has_content = True
            row_data[headers[col_idx-1]] = cell_value
    
    if has_content:
        print(f"\n--- 第 {row_idx} 行 ---")
        
        # 優先顯示關鍵欄位
        priority_fields = ['Send', 'Reply', 'Ref', 'Validation', 'Expect']
        for field in priority_fields:
            if field in row_data:
                value = str(row_data[field])
                if len(value) > 100:
                    value = value[:100] + "..."
                print(f"  {field}: {value}")
        
        # 顯示其他欄位
        for field, value in row_data.items():
            if field not in priority_fields:
                str_value = str(value)
                if len(str_value) > 80:
                    str_value = str_value[:80] + "..."
                print(f"  {field}: {str_value}")

# 統計分析
print(f"\n\n{'=' * 120}")
print("統計分析")
print("=" * 120)

if 'Ref' in key_columns:
    ref_col = key_columns['Ref']
    ref_count = 0
    ref_examples = []
    
    for row in range(2, ws.max_row + 1):
        ref_value = ws.cell(row, ref_col).value
        if ref_value and str(ref_value).strip():
            ref_count += 1
            if len(ref_examples) < 15:
                ref_examples.append((row, str(ref_value)))
    
    print(f"\nRef 欄位統計:")
    print(f"  總行數: {ws.max_row - 1}")
    print(f"  有值的行數: {ref_count}")
    print(f"  使用率: {ref_count / (ws.max_row - 1) * 100:.1f}%")
    
    if ref_examples:
        print(f"\n  範例 (前 15 個):")
        for row_num, ex in ref_examples:
            print(f"    Row {row_num}: {ex}")

if 'Validation' in key_columns:
    val_col = key_columns['Validation']
    val_count = 0
    val_examples = []
    
    for row in range(2, ws.max_row + 1):
        val_value = ws.cell(row, val_col).value
        if val_value and str(val_value).strip():
            val_count += 1
            if len(val_examples) < 15:
                val_examples.append((row, str(val_value)))
    
    print(f"\n\nValidation 欄位統計:")
    print(f"  總行數: {ws.max_row - 1}")
    print(f"  有值的行數: {val_count}")
    print(f"  使用率: {val_count / (ws.max_row - 1) * 100:.1f}%")
    
    if val_examples:
        print(f"\n  範例 (前 15 個):")
        for row_num, ex in val_examples:
            print(f"    Row {row_num}: {ex}")

print("\n" + "=" * 120)
print("分析完成")
print("=" * 120)
