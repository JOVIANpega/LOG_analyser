#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSV檔案整理工具
功能：選擇目錄 → 搜尋CSV → 顯示checkbox列表 → 複製到Analysis_CSV_FILE → 整理並重新命名
"""

import os
import shutil
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
import glob
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import re

class CSVProcessor:
    def __init__(self, app):
        self.app = app
        self.csv_files = []
        self.analysis_dir = None
        self.max_display_length = 50  # 預設最大顯示長度
        self.selected_files = []  # 使用者選擇的檔案
        
    def select_directory(self):
        """讓使用者選擇包含CSV檔案的目錄"""
        directory = filedialog.askdirectory(title="選擇包含CSV檔案的目錄")
        if directory:
            self.scan_csv_files(directory)
            return True
        return False
    
    def select_files(self):
        """讓使用者直接選擇CSV檔案（支援多選）"""
        file_paths = filedialog.askopenfilenames(
            title="選擇CSV檔案（可多選）",
            filetypes=[
                ("CSV檔案", "*.csv"),
                ("所有檔案", "*.*")
            ]
        )
        if file_paths:
            self.csv_files = list(file_paths)
            self.show_csv_selection_window()
            return True
        return False
    
    def scan_csv_files(self, directory):
        """掃描目錄中的CSV檔案"""
        self.csv_files = []
        patterns = ['*.csv', '*.CSV']
        
        for pattern in patterns:
            files = glob.glob(os.path.join(directory, pattern))
            self.csv_files.extend(files)
        
        # 顯示找到的CSV檔案
        if self.csv_files:
            self.show_csv_selection_window()
        else:
            messagebox.showinfo("提示", "在選擇的目錄中沒有找到CSV檔案")
    
    def show_csv_selection_window(self):
        """顯示CSV檔案選擇視窗（含checkbox）"""
        # 創建新視窗顯示CSV檔案列表
        self.selection_window = tk.Toplevel(self.app.root)
        self.selection_window.title("選擇要處理的CSV檔案")
        self.selection_window.geometry("700x500")
        self.selection_window.resizable(True, True)
        
        # 主框架
        main_frame = tk.Frame(self.selection_window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 標題
        title_label = tk.Label(main_frame, text=f"找到 {len(self.csv_files)} 個CSV檔案", 
                              font=('Arial', 14, 'bold'))
        title_label.pack(pady=(0, 10))
        
        # 說明文字
        info_label = tk.Label(main_frame, 
                             text="請選擇要處理的CSV檔案（至少選擇一個）", 
                             font=('Arial', 10), fg='blue')
        info_label.pack(pady=(0, 10))
        
        # 全選/全不選按鈕框架
        button_frame = tk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Button(button_frame, text="全選", 
                 command=self.select_all_files,
                 bg='#4CAF50', fg='white', font=('Arial', 9)).pack(side=tk.LEFT, padx=5)
        
        tk.Button(button_frame, text="全不選", 
                 command=self.deselect_all_files,
                 bg='#f44336', fg='white', font=('Arial', 9)).pack(side=tk.LEFT, padx=5)
        
        # 檔案列表框架
        list_frame = tk.Frame(main_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        # 創建Treeview with checkboxes
        columns = ('選擇', '檔案名稱', '檔案大小', '修改時間')
        self.tree = ttk.Treeview(list_frame, columns=columns, show='headings')
        
        # 設定欄位寬度
        self.tree.heading('選擇', text='選擇')
        self.tree.heading('檔案名稱', text='檔案名稱')
        self.tree.heading('檔案大小', text='檔案大小')
        self.tree.heading('修改時間', text='修改時間')
        
        self.tree.column('選擇', width=60, anchor='center')
        self.tree.column('檔案名稱', width=300)
        self.tree.column('檔案大小', width=100)
        self.tree.column('修改時間', width=150)
        
        # 添加檔案資訊
        self.checkbox_vars = []  # 儲存checkbox變數
        
        for i, file_path in enumerate(self.csv_files):
            file_name = os.path.basename(file_path)
            file_size = os.path.getsize(file_path)
            mod_time = os.path.getmtime(file_path)
            
            # 創建checkbox變數
            var = tk.BooleanVar()
            var.set(i == 0)  # 預設第一個打勾
            self.checkbox_vars.append(var)
            
            # 插入資料
            self.tree.insert('', 'end', values=(
                '☑' if var.get() else '☐',  # checkbox顯示
                file_name,
                f"{file_size:,} bytes",
                pd.Timestamp.fromtimestamp(mod_time).strftime('%Y-%m-%d %H:%M:%S')
            ))
            
            # 綁定checkbox變更事件
            var.trace('w', lambda *args, idx=i: self.update_checkbox_display(idx))
        
        # 滾動條
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 綁定點擊事件
        self.tree.bind('<Button-1>', self.on_tree_click)
        
        # 底部按鈕框架
        bottom_frame = tk.Frame(main_frame)
        bottom_frame.pack(fill=tk.X, pady=(10, 0))
        
        # 狀態標籤
        self.status_label = tk.Label(bottom_frame, text="已選擇 1 個檔案", 
                                    font=('Arial', 10), fg='green')
        self.status_label.pack(side=tk.LEFT)
        
        # 按鈕
        tk.Button(bottom_frame, text="開始處理", 
                 command=self.start_processing,
                 bg='#4CAF50', fg='white', font=('Arial', 10, 'bold')).pack(side=tk.RIGHT, padx=5)
        
        tk.Button(bottom_frame, text="取消", 
                 command=self.selection_window.destroy,
                 bg='#f44336', fg='white').pack(side=tk.RIGHT, padx=5)
        
        # 更新狀態
        self.update_status_label()
    
    def on_tree_click(self, event):
        """處理Treeview點擊事件"""
        item = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        
        if item and column == '#1':  # 點擊選擇欄位
            # 找到對應的檔案索引
            item_index = self.tree.index(item)
            if 0 <= item_index < len(self.checkbox_vars):
                # 切換checkbox狀態
                current_value = self.checkbox_vars[item_index].get()
                self.checkbox_vars[item_index].set(not current_value)
    
    def update_checkbox_display(self, index):
        """更新checkbox顯示"""
        if 0 <= index < len(self.checkbox_vars):
            var = self.checkbox_vars[index]
            items = list(self.tree.get_children())
            if 0 <= index < len(items):
                item = items[index]
                values = list(self.tree.item(item, 'values'))
                values[0] = '☑' if var.get() else '☐'
                self.tree.item(item, values=values)
                self.update_status_label()
    
    def select_all_files(self):
        """全選所有檔案"""
        for var in self.checkbox_vars:
            var.set(True)
    
    def deselect_all_files(self):
        """全不選所有檔案"""
        for var in self.checkbox_vars:
            var.set(False)
    
    def update_status_label(self):
        """更新狀態標籤"""
        selected_count = sum(1 for var in self.checkbox_vars if var.get())
        self.status_label.config(text=f"已選擇 {selected_count} 個檔案")
        
        # 更新開始處理按鈕狀態
        if selected_count == 0:
            self.status_label.config(fg='red')
        else:
            self.status_label.config(fg='green')
    
    def start_processing(self):
        """開始處理選中的CSV檔案"""
        # 檢查是否至少選擇一個檔案
        selected_files = []
        for i, var in enumerate(self.checkbox_vars):
            if var.get():
                selected_files.append(self.csv_files[i])
        
        if not selected_files:
            messagebox.showwarning("警告", "請至少選擇一個CSV檔案進行處理")
            return
        
        self.selected_files = selected_files
        self.selection_window.destroy()
        
        # 創建Analysis_CSV_FILE目錄
        self.create_analysis_directory()
        
        # 複製檔案並處理
        self.copy_and_process_files()
    
    def create_analysis_directory(self):
        """創建Analysis_CSV_FILE目錄"""
        current_dir = os.path.dirname(os.path.abspath(__file__))
        self.analysis_dir = os.path.join(current_dir, "Analysis_CSV_FILE")
        
        if not os.path.exists(self.analysis_dir):
            os.makedirs(self.analysis_dir)
            print(f"已創建目錄: {self.analysis_dir}")
    
    def copy_and_process_files(self):
        """複製檔案並逐一處理"""
        progress_window = self.create_progress_window()
        
        for i, csv_file in enumerate(self.selected_files):
            try:
                # 更新進度
                progress_window.update_progress(i, len(self.selected_files), os.path.basename(csv_file))
                
                # 複製檔案
                copied_file = self.copy_file(csv_file)
                
                # 處理CSV檔案
                self.process_csv_file(copied_file)
                
            except Exception as e:
                print(f"處理檔案 {csv_file} 時發生錯誤: {e}")
        
        progress_window.complete()
        
        # 處理完成後詢問是否開啟檔案
        self.ask_open_files()
    
    def ask_open_files(self):
        """詢問使用者是否要開啟處理後的檔案"""
        # 獲取所有生成的Excel檔案
        excel_files = []
        for file in os.listdir(self.analysis_dir):
            if file.endswith('.xlsx'):
                excel_files.append(os.path.join(self.analysis_dir, file))
        
        if not excel_files:
            messagebox.showinfo("完成", "CSV檔案處理完成！")
            return
        
        # 詢問是否開啟檔案
        result = messagebox.askyesno(
            "處理完成", 
            f"CSV檔案處理完成！\n\n已生成 {len(excel_files)} 個Excel檔案。\n\n是否要開啟Analysis_CSV_FILE資料夾？"
        )
        
        if result:
            # 開啟資料夾
            import subprocess
            import platform
            
            if platform.system() == "Windows":
                subprocess.run(["explorer", self.analysis_dir])
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", self.analysis_dir])
            else:  # Linux
                subprocess.run(["xdg-open", self.analysis_dir])
    
    def copy_file(self, source_file):
        """複製檔案到Analysis_CSV_FILE目錄"""
        file_name = os.path.basename(source_file)
        dest_file = os.path.join(self.analysis_dir, file_name)
        shutil.copy2(source_file, dest_file)
        return dest_file
    
    def process_csv_file(self, csv_file):
        """處理單個CSV檔案"""
        try:
            # 讀取CSV檔案
            df = pd.read_csv(csv_file, encoding='utf-8')
            
            # 創建新的Excel檔案
            output_file = self.create_output_filename(csv_file)
            self.create_formatted_excel(df, output_file)
            
        except Exception as e:
            print(f"處理CSV檔案 {csv_file} 失敗: {e}")
    
    def create_output_filename(self, csv_file):
        """創建輸出檔案名稱"""
        base_name = os.path.splitext(os.path.basename(csv_file))[0]
        return os.path.join(self.analysis_dir, f"{base_name}_Analysis_CSV.xlsx")
    
    def create_formatted_excel(self, df, output_file):
        """創建格式化的Excel檔案"""
        wb = Workbook()
        ws = wb.active
        ws.title = "CSV Analysis"
        
        # 設定樣式
        pass_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 淺綠色
        fail_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # 淺紅色
        header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # 淺藍色
        header_font = Font(bold=True, size=11)
        normal_font = Font(size=10)
        
        # 寫入標題行
        headers = list(df.columns)
        ws.append(headers)
        
        # 設定標題行樣式
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 寫入資料並設定樣式
        for row_num, (_, row) in enumerate(df.iterrows(), 2):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                # 處理文字內容
                text_value = str(value) if pd.notna(value) else ""
                cell.value = text_value[:self.max_display_length]  # 限制顯示長度
                cell.font = normal_font
                
                # 根據內容判斷PASS/FAIL並設定顏色
                if self.is_pass_row(row):
                    cell.fill = pass_fill
                elif self.is_fail_row(row):
                    cell.fill = fail_fill
                
                # 設定文字對齊
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 自動調整欄寬（優化版本）
        self.auto_adjust_column_width_optimized(ws)
        
        # 設定行高
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 20
        
        # 儲存檔案
        wb.save(output_file)
        print(f"已處理並儲存: {output_file}")
    
    def is_pass_row(self, row):
        """判斷是否為PASS行"""
        pass_keywords = ['PASS', 'pass', 'OK', 'ok', 'SUCCESS', 'success', 'TRUE', 'true']
        for value in row:
            if pd.notna(value) and any(keyword in str(value).upper() for keyword in pass_keywords):
                return True
        return False
    
    def is_fail_row(self, row):
        """判斷是否為FAIL行"""
        fail_keywords = ['FAIL', 'fail', 'ERROR', 'error', 'NACK', 'nack', 'TIMEOUT', 'timeout', 'FALSE', 'false']
        for value in row:
            if pd.notna(value) and any(keyword in str(value).upper() for keyword in fail_keywords):
                return True
        return False
    
    def auto_adjust_column_width_optimized(self, ws):
        """優化的自動調整欄寬功能"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        # 計算文字長度（考慮中文字符）
                        text_length = len(str(cell.value))
                        
                        # 中文字符權重調整
                        chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', str(cell.value)))
                        adjusted_length = text_length + chinese_chars * 0.5
                        
                        if adjusted_length > max_length:
                            max_length = adjusted_length
                except:
                    pass
            
            # 設定欄寬（優化計算）
            if max_length == 0:
                adjusted_width = 10
            elif max_length < 5:
                adjusted_width = max_length + 2
            elif max_length <= 15:
                adjusted_width = max_length + 3
            elif max_length <= 30:
                adjusted_width = max_length + 4
            else:
                adjusted_width = min(max_length + 5, self.max_display_length + 5)
            
            # 設定最小和最大寬度
            adjusted_width = max(adjusted_width, 8)  # 最小寬度
            adjusted_width = min(adjusted_width, 60)  # 最大寬度
            
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_progress_window(self):
        """創建進度視窗"""
        return ProgressWindow(self.app.root, "處理CSV檔案")

class ProgressWindow:
    def __init__(self, parent, title):
        self.window = tk.Toplevel(parent)
        self.window.title(title)
        self.window.geometry("400x150")
        self.window.resizable(False, False)
        
        # 進度條
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.window, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(pady=20, padx=20, fill=tk.X)
        
        # 狀態標籤
        self.status_label = tk.Label(self.window, text="準備中...", font=('Arial', 10))
        self.status_label.pack(pady=10)
        
        # 檔案名稱標籤
        self.file_label = tk.Label(self.window, text="", font=('Arial', 9), fg='gray')
        self.file_label.pack()
        
        self.window.update()
    
    def update_progress(self, current, total, filename):
        """更新進度"""
        progress = (current / total) * 100
        self.progress_var.set(progress)
        self.status_label.config(text=f"處理中... ({current}/{total})")
        self.file_label.config(text=f"當前檔案: {filename}")
        self.window.update()
    
    def complete(self):
        """完成處理"""
        self.progress_var.set(100)
        self.status_label.config(text="處理完成！")
        self.file_label.config(text="所有選中的CSV檔案已處理完成")
        
        # 3秒後自動關閉
        self.window.after(3000, self.window.destroy)

