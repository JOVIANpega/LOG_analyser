# -*- coding: utf-8 -*-
"""
新格式FAIL工作簿构建器 - 添加到excel_writer.py末尾的方法
"""

def _build_fail_workbook_new_format(self, output_path: str, logs: list):
    """
    构建新格式的FAIL工作簿 (Dashboard + FAIL_LIST + 原始LOG工作表)
    """
    from .excel_fail_list_builder import FailListBuilder
    from openpyxl import Workbook
    
    wb = Workbook()
    # 移除默认工作表
    if wb.active:
        wb.remove(wb.active)
    
    # 创建FailListBuilder实例
    builder = FailListBuilder()
    
    # 准备数据 - 转换logs格式以适配FailListBuilder
    processed_logs = []
    sheet_map = {}
    
    for entry in logs:
        # 提取错误文本 (使用4级优先级逻辑)
        error_text = self._extract_primary_error(entry)
        
        # 准备LOG条目
        log_data = {
            'filename': entry.get('file_name', ''),
            'error_text': error_text,
            'summary': entry.get('summary', {}),
            'fail_items': entry.get('fail_items', []),
            'retry_count': self._count_retries(entry)
        }
        
        # 创建原始LOG工作表
        fname = entry.get('file_name', 'LOG')
        isn = self._extract_isn_from_filename(fname)
        sheet_name_base = self._sanitize_sheet_title(isn if isn else fname)
        sheet_name = self._unique_sheet_name(wb, sheet_name_base)
        sheet_map[fname] = sheet_name
        log_data['sheet_name'] = sheet_name
        
        # 创建工作表
        ws_log = wb.create_sheet(title=sheet_name)
        
        # 添加返回Dashboard的链接
        back_cell = ws_log.cell(row=1, column=1, value='🔙 回到 Dashboard')
        back_cell.font = Font(name='Calibri', size=11, bold=True, color='FF0000FF', underline='single')
        back_cell.hyperlink = "#Dashboard!A1"
        
        # 档名标题
        title_cell = ws_log.cell(row=3, column=1, value=self._sanitize_cell_text(fname))
        title_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill('solid', fgColor='FFC00000')
        
        # 写入原始LOG
        current_row = 5
        self._write_raw_log_with_annotations(
            ws_log,
            start_row=current_row,
            raw_lines=entry.get('raw_lines', []),
            annotations=entry.get('ui_annotations', []),
            font=Font(name='Calibri', size=11),
            step_marks=entry.get('step_marks')
        )
        
        # 调整列宽
        self._auto_fit_columns(ws_log, min_widths={1: 150})
        
        processed_logs.append(log_data)
    
    # 生成Dashboard工作表
    builder.build_dashboard_sheet(wb, processed_logs)
    
    # 生成FAIL_LIST工作表
    builder.build_fail_list_sheet(wb, processed_logs)
    
    # 保存工作簿
    try:
        wb.save(output_path)
        print(f"新格式FAIL报告已生成: {output_path}")
    finally:
        try:
            wb.close()
        except:
            pass

def _extract_primary_error(self, entry):
    """
    使用4级优先级逻辑提取主要错误
    1. "doesn't match" (最后一次出现)
    2. "is Fail" (最后一次出现)
    3. "FAIL" (最后一次出现)
    4. "ERROR" (最后一次出现)
    """
    raw_lines = entry.get('raw_lines', [])
    if not raw_lines:
        return entry.get('summary', {}).get('FAIL原因', 'Unknown Error')
    
    # 转换为字符串列表
    lines = [str(line) for line in raw_lines]
    full_text = '\n'.join(lines)
    
    # 优先级1: "doesn't match"
    matches = [i for i, line in enumerate(lines) if "doesn't match" in line.lower()]
    if matches:
        return lines[matches[-1]].strip()
    
    # 优先级2: "is Fail"
    matches = [i for i, line in enumerate(lines) if "is fail" in line.lower()]
    if matches:
        return lines[matches[-1]].strip()
    
    # 优先级3: "FAIL"
    matches = [i for i, line in enumerate(lines) if "fail" in line.lower()]
    if matches:
        return lines[matches[-1]].strip()
    
    # 优先级4: "ERROR"
    matches = [i for i, line in enumerate(lines) if "error" in line.lower()]
    if matches:
        return lines[matches[-1]].strip()
    
    return entry.get('summary', {}).get('FAIL原因', 'Unknown Error')

def _count_retries(self, entry):
    """统计retry次数"""
    retry_count = 0
    for item in entry.get('fail_items', []):
        retry_count += item.get('retry', 0)
    return retry_count
