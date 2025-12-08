# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import sys
import os

# 設定輸出編碼為 UTF-8
sys.stdout.reconfigure(encoding='utf-8')

input_file = r'MINE\result\FAIL匯總.xlsx'
output_file = r'MINE\result\FAIL匯總_NewFormat.xlsx'

print(f"正在讀取原始檔案: {input_file}")

try:
    # 讀取原始 Excel
    wb = openpyxl.load_workbook(input_file)
    
    # 取得原始 Summary 工作表來提取資料
    if 'Summary' not in wb.sheetnames:
        print("錯誤: 找不到 Summary 工作表")
        sys.exit(1)
        
    ws_orig = wb['Summary']
    
    extracted_data = []
    
    # 遍歷 Summary 的每一行提取資料
    for row in range(1, ws_orig.max_row + 1):
        col1_val = ws_orig.cell(row=row, column=1).value
        col2_val = ws_orig.cell(row=row, column=2).value # 詳細錯誤原因
        
        if not col1_val or not isinstance(col1_val, str):
            continue
            
        # 跳過無關行
        if "錯誤原因統計" in col1_val or "──" in col1_val or "回到 Summary" in col1_val or "工作表快速連結" in col1_val:
            continue
            
        # 解析 Log 資訊 (ISN, Station)
        log_info = col1_val
        
        # 預設值
        isn = ""
        station = ""
        fail_item = ""
        fail_reason = ""
        suggestion = "" # 目前資料可能沒有，先留空
        
        # 嘗試解析 ISN 和 Station
        # 格式範例: 1+4cam stitching1 test-WE253600014-2025-0923-104358...
        parts = log_info.split('-')
        
        # Station (測項)
        if '+' in parts[0]:
            station = parts[0].split('+')[1].strip()
        else:
            station = parts[0].strip()
            
        # ISN (序號) - 尋找 WE 開頭的
        for part in parts:
            if part.startswith('WE') and len(part) > 8:
                isn = part
                break
        if not isn and len(parts) > 1:
            isn = parts[1] # 如果找不到 WE 開頭，暫取第二段
            
        # 解析錯誤訊息 (FAIL Item vs FAIL Reason)
        full_error = str(col2_val) if col2_val else ""
        
        # 清理雜訊
        full_error = full_error.replace("===============錯誤原因====================\n\n", "")
        full_error = full_error.replace("==================================================\n\n", "")
        
        lines = full_error.split('\n')
        clean_lines = [l.strip() for l in lines if l.strip()]
        
        if clean_lines:
            # 啟發式規則：第一行通常是錯誤概或是 FAIL Item
            first_line = clean_lines[0]
            
            # 嘗試從第一行提取 FAIL Item
            # 例如: "48:12 [1] B7PL011-202:TRY_TEXT is Fail" -> "TRY_TEXT is Fail"
            if "is Fail" in first_line:
                fail_item = first_line.split(':')[-1].strip()
            elif "doesn't match" in first_line:
                fail_item = "doesn't match SPEC"
                fail_reason = first_line
            else:
                fail_item = first_line
            
            # 剩下的作為 FAIL Reason
            # 尋找包含數值比較的行，例如 "AVE_SNR = 28.8..." 或 "MAX_PixelsShift_0=7.0"
            reason_lines = []
            for line in clean_lines:
                if "=" in line or "expected" in line or "got" in line or "fail" in line.lower():
                    if line != first_line:
                        reason_lines.append(line)
            
            if reason_lines:
                fail_reason = "\n".join(reason_lines[:3]) # 取前3行重要資訊
            elif len(clean_lines) > 1:
                fail_reason = clean_lines[1]
            else:
                if not fail_reason:
                    fail_reason = first_line # 如果只有一行，就都放這裡
        
        # 建立資料物件
        extracted_data.append({
            "ISN": isn,
            "Station": station,
            "FAIL Item": fail_item,
            "FAIL Reason": fail_reason,
            "suggestion": "提供圖片給PEGA 看", # 依照截圖範例填入預設值，或者留空
            "Original_Log_Name": log_info # 用於建立連結
        })

    # 建立新的工作表 'Dashboard'
    if 'Dashboard' in wb.sheetnames:
        wb.remove(wb['Dashboard'])
    ws_new = wb.create_sheet('Dashboard', 0) # 放在最前面
    
    # 設定標題
    headers = ['ISN', 'Station', 'FAIL Item', 'FAIL Reason', 'suggestion', 'Full Log']
    ws_new.append(headers)
    
    # 設定樣式
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid") # 綠色標題
    header_font = Font(name='Microsoft JhengHei', size=12, color="FFFFFF", bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 寫入標題樣式
    for col_num, header in enumerate(headers, 1):
        cell = ws_new.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # 寫入資料
    for row_idx, data in enumerate(extracted_data, 2):
        # ISN
        c1 = ws_new.cell(row=row_idx, column=1, value=data['ISN'])
        c1.alignment = center_align
        c1.border = border
        
        # Station
        c2 = ws_new.cell(row=row_idx, column=2, value=data['Station'])
        c2.alignment = center_align
        c2.border = border
        
        # FAIL Item
        c3 = ws_new.cell(row=row_idx, column=3, value=data['FAIL Item'])
        c3.alignment = left_align
        c3.border = border
        
        # FAIL Reason
        c4 = ws_new.cell(row=row_idx, column=4, value=data['FAIL Reason'])
        c4.alignment = left_align
        c4.border = border
        
        # suggestion
        c5 = ws_new.cell(row=row_idx, column=5, value=data['suggestion'])
        c5.alignment = center_align
        c5.border = border
        
        # Full Log Link
        c6 = ws_new.cell(row=row_idx, column=6, value="查看 Log")
        c6.font = Font(color="0000FF", underline="single")
        c6.alignment = center_align
        c6.border = border
        
        # 嘗試建立連結到對應的 Sheet
        # 原始程式碼中 sheet 名稱可能被截斷或處理過，這裡嘗試簡單匹配
        # 通常是檔名的一部分
        target_sheet = None
        for sheet_name in wb.sheetnames:
            if sheet_name != 'Summary' and sheet_name != 'Dashboard':
                # 簡單比對：如果 sheet name 在原始檔名中
                clean_sheet_name = sheet_name.split('(')[0].strip() # 移除可能的 (1) 後綴
                if clean_sheet_name in data['Original_Log_Name']:
                    target_sheet = sheet_name
                    break
        
        if target_sheet:
            c6.hyperlink = f"#'{target_sheet}'!A1"
        else:
            c6.value = "無連結"

        # 隔行變色 (模仿截圖的淺藍色背景)
        if row_idx % 2 == 1: # 奇數行 (Excel row 3, 5...)
            fill_color = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            for col in range(1, 7):
                ws_new.cell(row=row_idx, column=col).fill = fill_color

    # 設定欄寬
    ws_new.column_dimensions['A'].width = 20 # ISN
    ws_new.column_dimensions['B'].width = 20 # Station
    ws_new.column_dimensions['C'].width = 40 # FAIL Item
    ws_new.column_dimensions['D'].width = 60 # FAIL Reason
    ws_new.column_dimensions['E'].width = 25 # suggestion
    ws_new.column_dimensions['F'].width = 15 # Link

    # 凍結首列
    ws_new.freeze_panes = 'A2'

    # 儲存新檔案
    wb.save(output_file)
    print(f"已生成新格式檔案: {output_file}")

except Exception as e:
    print(f"發生錯誤: {e}")
    import traceback
    traceback.print_exc()
